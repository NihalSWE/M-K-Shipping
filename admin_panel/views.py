from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.admin.views.decorators import staff_member_required
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.contrib.auth.decorators import permission_required
from django.contrib.auth import update_session_auth_hash
from django.views.decorators.clickjacking import xframe_options_sameorigin
import sys
import re
from django.http import JsonResponse
from django.contrib import messages
from django.db.models import F
from io import BytesIO
from datetime import timedelta, datetime
from django.utils.crypto import get_random_string
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.utils.dateparse import parse_datetime
from django.utils import timezone

from portal.booking_invoice.qr_generation import ensure_booking_qr            # <--- NEW IMPORT
from .tasks import send_sms_task, auto_cancel_booking  # <--- NEW IMPORT
from django.core.validators import URLValidator
from django.core.exceptions import ValidationError
import csv
from decimal import Decimal
from django.db.models import Sum, Q, Count, F, DecimalField, ExpressionWrapper, Value
from django.db.models.functions import Coalesce, TruncDate
from django.http import HttpResponse
import uuid 
import json
from .models import *
from accounts.models import UserProfile
import logging
from django.db.models import Max
from django.views.decorators.http import require_POST
from django.db.models import ProtectedError
from django.db import transaction
from django.db import IntegrityError
from django.contrib.auth import get_user_model, logout
from .services import sync_route_prices
from .services import generate_smart_trips
from .forms import BlogPostForm, BlogBannerForm, AdminUserAddForm, TripSearchForm, AdminUserPermissionsForm
from accounts.forms import AdminUserEditForm
from django.urls import reverse
from .utils import send_booking_sms, get_logged_in_counter
from portal.utils import seat_hold_key, get_holder_id
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer
from PIL import Image
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from django.db.models import Sum, Q









User = get_user_model()
logger = logging.getLogger(__name__)



@login_required
def dashboard(request):
    now = timezone.now()
    today = now.date()

    # -----------------------------
    # Date ranges
    # -----------------------------
    start_today = timezone.make_aware(timezone.datetime.combine(today, timezone.datetime.min.time()))
    end_today = timezone.make_aware(timezone.datetime.combine(today, timezone.datetime.max.time()))

    first_day_of_month = today.replace(day=1)
    start_month = timezone.make_aware(timezone.datetime.combine(first_day_of_month, timezone.datetime.min.time()))
    end_month = end_today

    # -----------------------------
    # 1) KPI cards
    # -----------------------------
    bookings_today = Booking.objects.filter(created_at__date=today).count()

    paid_revenue_month = Booking.objects.filter(
        payment_status='PAID',
        created_at__gte=start_month,
        created_at__lte=end_month
    ).aggregate(total=Sum('paid_amount'))['total'] or Decimal('0.00')

    active_trips_today = Trip.objects.filter(
        departure_datetime__date=today,
        is_published=True
    ).count()

    active_seat_holds = SeatHold.objects.filter(expires_at__gt=now).count()

    pending_bookings = Booking.objects.filter(status='PENDING').count()

    expired_bookings = Booking.objects.filter(status='EXPIRED').count()

    # -----------------------------
    # 2) Occupancy (today average)
    # -----------------------------
    today_trips = Trip.objects.filter(departure_datetime__date=today).select_related('ship', 'route')
    today_trip_ids = list(today_trips.values_list('id', flat=True))

    booked_tickets_counts = (
        Ticket.objects.filter(
            trip_id__in=today_trip_ids,
            status='BOOKED'
        )
        .values('trip_id')
        .annotate(cnt=Count('id'))
    )
    booked_map = {item['trip_id']: item['cnt'] for item in booked_tickets_counts}

    occupancy_values = []
    for trip in today_trips:
        capacity = trip.ship.total_capacity or 0
        if capacity > 0:
            sold = booked_map.get(trip.id, 0)
            occupancy_values.append((sold / capacity) * 100)

    avg_occupancy_today = round(sum(occupancy_values) / len(occupancy_values), 2) if occupancy_values else 0

    # -----------------------------
    # 3) Booking trend chart (last 14 days)
    # -----------------------------
    last_14_days_start = today - timezone.timedelta(days=13)

    bookings_trend_qs = (
        Booking.objects.filter(created_at__date__gte=last_14_days_start, created_at__date__lte=today)
        .annotate(day=TruncDate('created_at'))
        .values('day')
        .annotate(total=Count('id'))
        .order_by('day')
    )
    bookings_trend_map = {item['day']: item['total'] for item in bookings_trend_qs}

    revenue_trend_qs = (
        Booking.objects.filter(
            payment_status='PAID',
            created_at__date__gte=last_14_days_start,
            created_at__date__lte=today
        )
        .annotate(day=TruncDate('created_at'))
        .values('day')
        .annotate(total=Sum('paid_amount'))
        .order_by('day')
    )
    revenue_trend_map = {item['day']: float(item['total'] or 0) for item in revenue_trend_qs}

    trend_labels = []
    trend_bookings = []
    trend_revenue = []

    for i in range(14):
        d = last_14_days_start + timezone.timedelta(days=i)
        trend_labels.append(d.strftime('%d %b'))
        trend_bookings.append(bookings_trend_map.get(d, 0))
        trend_revenue.append(revenue_trend_map.get(d, 0))

    # -----------------------------
    # 4) Booking status chart
    # -----------------------------
    booking_status_counts = Booking.objects.values('status').annotate(total=Count('id'))
    status_map = {item['status']: item['total'] for item in booking_status_counts}

    booking_status_labels = ['PENDING', 'CONFIRMED', 'CANCELLED', 'EXPIRED', 'LOCKED']
    booking_status_series = [status_map.get(label, 0) for label in booking_status_labels]

    # -----------------------------
    # 5) Sales channel chart
    # -----------------------------
    sales_channel_counts = Booking.objects.values('sales_channel').annotate(total=Count('id'))
    channel_map = {item['sales_channel']: item['total'] for item in sales_channel_counts}

    sales_channel_labels = ['ONLINE', 'COUNTER']
    sales_channel_series = [channel_map.get(label, 0) for label in sales_channel_labels]

    # -----------------------------
    # 6) Top routes (by booked tickets)
    # -----------------------------
    top_routes = (
        Ticket.objects.filter(status='BOOKED')
        .values('trip__route__name')
        .annotate(
            tickets_sold=Count('id'),
            revenue=Sum('fare_amount')
        )
        .order_by('-tickets_sold')[:5]
    )

    top_routes_data = [
        {
            'route_name': item['trip__route__name'] or 'N/A',
            'tickets_sold': item['tickets_sold'],
            'revenue': float(item['revenue'] or 0),
        }
        for item in top_routes
    ]

    # -----------------------------
    # 7) Upcoming trips (next 6)
    # -----------------------------
    upcoming_trips_qs = (
        Trip.objects.filter(departure_datetime__gte=now, is_published=True)
        .select_related('ship', 'route')
        .order_by('departure_datetime')[:6]
    )

    upcoming_trip_ids = [trip.id for trip in upcoming_trips_qs]
    upcoming_ticket_counts = (
        Ticket.objects.filter(trip_id__in=upcoming_trip_ids, status='BOOKED')
        .values('trip_id')
        .annotate(cnt=Count('id'))
    )
    upcoming_ticket_map = {item['trip_id']: item['cnt'] for item in upcoming_ticket_counts}

    upcoming_trips = []
    for trip in upcoming_trips_qs:
        sold = upcoming_ticket_map.get(trip.id, 0)
        capacity = trip.ship.total_capacity or 0
        occupancy = round((sold / capacity) * 100, 1) if capacity > 0 else 0

        upcoming_trips.append({
            'ship': trip.ship.name,
            'route': trip.route.name,
            'departure': timezone.localtime(trip.departure_datetime).strftime('%d %b %Y, %I:%M %p'),
            'sold': sold,
            'capacity': capacity,
            'occupancy': occupancy,
        })

    # -----------------------------
    # 8) Top ships (by revenue)
    # -----------------------------
    top_ships = (
        Ticket.objects.filter(status='BOOKED')
        .values('trip__ship__name')
        .annotate(
            tickets_sold=Count('id'),
            revenue=Sum('fare_amount')
        )
        .order_by('-revenue')[:5]
    )

    top_ships_data = [
        {
            'ship_name': item['trip__ship__name'] or 'N/A',
            'tickets_sold': item['tickets_sold'],
            'revenue': float(item['revenue'] or 0),
        }
        for item in top_ships
    ]

    # -----------------------------
    # 9) Alert module (small operational block)
    # -----------------------------
    trips_departing_6h = Trip.objects.filter(
        departure_datetime__gte=now,
        departure_datetime__lte=now + timezone.timedelta(hours=6),
        is_published=True
    ).count()

    overdue_pending = Booking.objects.filter(
        status='PENDING',
        expiry_at__isnull=False,
        expiry_at__lt=now
    ).count()

    context = {
        # KPI
        'bookings_today': bookings_today,
        'paid_revenue_month': float(paid_revenue_month),
        'active_trips_today': active_trips_today,
        'active_seat_holds': active_seat_holds,
        'pending_bookings': pending_bookings,
        'expired_bookings': expired_bookings,
        'avg_occupancy_today': avg_occupancy_today,

        # Modules
        'top_routes_data': top_routes_data,
        'upcoming_trips': upcoming_trips,
        'top_ships_data': top_ships_data,
        'trips_departing_6h': trips_departing_6h,
        'overdue_pending': overdue_pending,

        # Charts (JSON)
        'trend_labels_json': json.dumps(trend_labels),
        'trend_bookings_json': json.dumps(trend_bookings),
        'trend_revenue_json': json.dumps(trend_revenue),
        'booking_status_labels_json': json.dumps(booking_status_labels),
        'booking_status_series_json': json.dumps(booking_status_series),
        'sales_channel_labels_json': json.dumps(sales_channel_labels),
        'sales_channel_series_json': json.dumps(sales_channel_series),
    }
    return render(request, 'admin_panel/dashboard/dashboard.html', context)


@require_POST
def admin_logout(request):
    logout(request)
    return redirect(reverse("signin"))  # -> /accounts/signin/


import base64
from django.core.files.base import ContentFile
@login_required
@csrf_exempt
def ships(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            action = data.get('action')
            
            if action == 'add':
                name = data.get('name')
                code = data.get('code')
                total_capacity = data.get('total_capacity', 0)
                
                # Create ship
                ship = Ship.objects.create(name=name, code=code, total_capacity=total_capacity)
                
                # Handle base64 image if provided
                if 'image' in data and data['image']:
                    # Decode base64 image
                    format, imgstr = data['image'].split(';base64,') 
                    ext = format.split('/')[-1] 
                    image_data = ContentFile(
                        base64.b64decode(imgstr), 
                        name=f"{slugify(name)}.{ext}"
                    )
                    ship.image.save(f"{slugify(name)}.{ext}", image_data, save=True)
                
                return JsonResponse({'status': 'success', 'message': 'Ship added successfully!'})
            
            elif action == 'edit':
                ship_id = data.get('id')
                ship = get_object_or_404(Ship, id=ship_id)
                ship.name = data.get('name')
                ship.code = data.get('code')
                ship.total_capacity = data.get('total_capacity', 0)
                
                # Handle base64 image if provided
                if 'image' in data and data['image']:
                    # Decode base64 image
                    format, imgstr = data['image'].split(';base64,') 
                    ext = format.split('/')[-1] 
                    image_data = ContentFile(
                        base64.b64decode(imgstr), 
                        name=f"{slugify(ship.name)}.{ext}"
                    )
                    ship.image.save(f"{slugify(ship.name)}.{ext}", image_data, save=True)
                
                ship.save()
                return JsonResponse({'status': 'success', 'message': 'Ship updated successfully!'})
            
            elif action == 'delete':
                ship_id = data.get('id')
                ship = get_object_or_404(Ship, id=ship_id)
                ship.delete()
                return JsonResponse({'status': 'success', 'message': 'Ship deleted successfully!'})
                
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    ships = Ship.objects.all().order_by('-id')
    return render(request, 'admin_panel/ships/ship.html', {'ships': ships})

@login_required
@csrf_exempt
def ship_details(request, ship_id):
    ship = get_object_or_404(Ship, id=ship_id)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            action = data.get('action')

            # --- SHIP ACTIONS ---
            if action == 'update_ship':
                ship.name = data.get('name')
                ship.code = data.get('code')
                ship.total_capacity = data.get('total_capacity', 0)
                ship.save()
                return JsonResponse({'status': 'success', 'message': 'Ship updated successfully'})

            # --- DECK ACTIONS ---
            elif action == 'add_deck':
                Deck.objects.create(
                    ship=ship,
                    name=data.get('name'),
                    level_order=data.get('level_order', 1),
                    grid_cols=data.get('grid_cols', 24),
                    total_rows=data.get('total_rows', 20)
                )
                return JsonResponse({'status': 'success', 'message': 'Deck added successfully'})

            elif action == 'edit_deck':
                deck = get_object_or_404(Deck, id=data.get('id'), ship=ship)
                deck.name = data.get('name')
                deck.level_order = data.get('level_order')
                deck.grid_cols = data.get('grid_cols')
                deck.total_rows = data.get('total_rows')
                deck.save()
                return JsonResponse({'status': 'success', 'message': 'Deck updated successfully'})

            elif action == 'delete_deck':
                deck = get_object_or_404(Deck, id=data.get('id'), ship=ship)
                deck.delete()
                return JsonResponse({'status': 'success', 'message': 'Deck deleted successfully'})

            return JsonResponse({'status': 'error', 'message': 'Invalid action'}, status=400)

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    decks = ship.decks.all().order_by('level_order')
    return render(request, 'admin_panel/ships/ship_details.html', {
        'ship': ship,
        'decks': decks
    })
    
    
# @login_required
@login_required
def manage_structures(request):
    """
    Manages non-bookable layout structures (Corridors, Walls, Labels).
    """
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            action = data.get('action')
            
            icon_id = data.get('icon_id')
            icon_obj = SeatIcon.objects.filter(id=icon_id).first() if icon_id else None

            if action == 'add':
                SeatCategory.objects.create(
                    name=data.get('name'),
                    description=data.get('description', ''),
                    color_code=data.get('color_code', '#FFFFFF'),
                    icon=icon_obj,
                    is_bookable=False,  # Enforce Non-Bookable
                    capacity=0          # Structures usually have 0 capacity
                )
                return JsonResponse({'status': 'success', 'message': 'Structure added successfully!'})

            elif action == 'edit':
                category = get_object_or_404(SeatCategory, id=data.get('id'))
                category.name = data.get('name')
                category.description = data.get('description', '')
                category.color_code = data.get('color_code', '#FFFFFF')
                category.icon = icon_obj
                # Ensure we don't accidentally make it bookable via this view
                category.is_bookable = False 
                category.save()
                return JsonResponse({'status': 'success', 'message': 'Structure updated successfully!'})

            elif action == 'delete':
                category = get_object_or_404(SeatCategory, id=data.get('id'))
                category.delete()
                return JsonResponse({'status': 'success', 'message': 'Structure deleted successfully!'})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    # GET Request: Only fetch non-bookable categories
    structures = SeatCategory.objects.filter(is_bookable=False).order_by('name')
    all_icons = SeatIcon.objects.all().order_by('name')
    
    context = {
        'structures': structures,
        'all_icons': all_icons,
    }
    return render(request, 'admin_panel/seat_layout/manage_structures.html', context)


@login_required
def manage_bookable_categories(request):
    """
    Manages Bookable Seat Categories (Cabins, VIP Seats, Economy).
    Enforces is_bookable = True.
    """
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            action = data.get('action')
            
            # Get the Icon instance from the ID sent by the dropdown
            icon_id = data.get('icon_id')
            icon_obj = SeatIcon.objects.filter(id=icon_id).first() if icon_id else None

            if action == 'add':
                SeatCategory.objects.create(
                    name=data.get('name'),
                    description=data.get('description', ''),
                    color_code=data.get('color_code', '#000000'),
                    icon=icon_obj,
                    capacity=int(data.get('capacity', 1)),
                    is_bookable=True
                )
                return JsonResponse({'status': 'success', 'message': 'Category added successfully!'})

            elif action == 'edit':
                category = get_object_or_404(SeatCategory, id=data.get('id'))
                category.name = data.get('name')
                category.description = data.get('description', '')
                category.color_code = data.get('color_code', '#000000')
                category.icon = icon_obj
                category.capacity = int(data.get('capacity', 1))
                category.save()
                return JsonResponse({'status': 'success', 'message': 'Category updated successfully!'})

            elif action == 'delete':
                category = get_object_or_404(SeatCategory, id=data.get('id'))
                category.delete()
                return JsonResponse({'status': 'success', 'message': 'Category deleted successfully!'})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    # GET: Only fetch bookable categories
    categories = SeatCategory.objects.filter(is_bookable=True).order_by('name')
    all_icons = SeatIcon.objects.all().order_by('name')
    
    context = {
        'categories': categories,
        'all_icons': all_icons,
    }
    return render(request, 'admin_panel/seat_layout/manage_bookable_categories.html', context)

@login_required
def manage_seat_features(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            action = data.get("action")

            if action == "add":
                SeatFeature.objects.create(
                    name=data.get("name"),
                    description=data.get("description")
                )
                return JsonResponse({"status": "success", "message": "Feature added successfully!"})

            elif action == "edit":
                obj = SeatFeature.objects.get(id=data.get("id"))
                obj.name = data.get("name")
                obj.description = data.get("description")
                obj.save()
                return JsonResponse({"status": "success", "message": "Feature updated successfully!"})

            elif action == "delete":
                SeatFeature.objects.get(id=data.get("id")).delete()
                return JsonResponse({"status": "success", "message": "Feature deleted successfully!"})

        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)})

    # GET request
    context = {
        "features": SeatFeature.objects.all().order_by("-id"),
    }
    return render(request, "admin_panel/seat_layout/manage_seat_features.html", context)

@login_required
def seat_icon_management(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            action = data.get("action")
            
            if action == "add":
                SeatIcon.objects.create(
                    name=data.get("name"),
                    iconify_code=data.get("iconify_code")
                )
                return JsonResponse({"status": "success", "message": "Icon added successfully!"})

            elif action == "edit":
                icon = SeatIcon.objects.get(id=data.get("id"))
                icon.name = data.get("name")
                icon.iconify_code = data.get("iconify_code")
                icon.save()
                return JsonResponse({"status": "success", "message": "Icon updated successfully!"})

            elif action == "delete":
                SeatIcon.objects.get(id=data.get("id")).delete()
                return JsonResponse({"status": "success", "message": "Icon deleted successfully!"})

        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)})

    icons = SeatIcon.objects.all().order_by('name')
    return render(request, 'admin_panel/seat_layout/seat_icons.html', {'icons': icons})
    

@login_required
def seat_plan_editor(request, deck_id):
    deck = get_object_or_404(Deck, id=deck_id)
    
    # 1. Fetch Categories (The Brushes)
    categories = list(SeatCategory.objects.values(
        'id', 
        'name', 
        'color_code', 
        'is_bookable', 
        'icon__iconify_code'
    ))
    
    # 2. Fetch Features (The Tags like "River Side")
    features = list(SeatFeature.objects.values('id', 'name'))

    # 3. Fetch Existing Layout (If any)
    existing_objects = LayoutObject.objects.filter(deck=deck).select_related('category')
    layout_data = []
    for obj in existing_objects:
        layout_data.append({
            'row': obj.row_index,
            'col': obj.col_index,
            'row_span': obj.row_span,
            'col_span': obj.col_span,
            'category_id': obj.category.id,
            'label': obj.label,
            'features': list(obj.features.values_list('id', flat=True))
        })

    context = {
        'deck': deck,
        'categories_json': json.dumps(categories),
        'features_json': json.dumps(features),
        'layout_json': json.dumps(layout_data),
    }
    # Point this to your CUSTOM template location
    return render(request, 'admin_panel/seat_layout/seat_plan_editor.html', context)

# @csrf_exempt
# # @staff_member_required
# def save_seat_layout(request, deck_id):
#     if request.method == "POST":
#         data = json.loads(request.body)
#         deck = get_object_or_404(Deck, id=deck_id)
        
#         # 1. Clear old layout for this deck (Simple approach)
#         LayoutObject.objects.filter(deck=deck).delete()
        
#         # 2. Bulk Create new objects
#         new_objects = []
#         feature_relations = [] # To handle ManyToMany
        
#         for item in data.get('layout', []):
#             obj = LayoutObject(
#                 deck=deck,
#                 row_index=item['row'],
#                 col_index=item['col'],
#                 row_span=item['row_span'],
#                 col_span=item['col_span'],
#                 category_id=item['category_id'],
#                 label=item.get('label', ''),
#                 seat_identifier=item.get('seat_id', None)
#             )
#             # We must save to get an ID before adding M2M
#             obj.save() 
            
#             # Add features
#             if 'feature_ids' in item:
#                 obj.features.set(item['feature_ids'])
        
#         return JsonResponse({'status': 'success', 'message': 'Layout saved successfully'})
#     return JsonResponse({'status': 'error'}, status=400)


# @staff_member_required
# @require_POST
@login_required
@csrf_exempt
def update_deck_rows(request, deck_id):
    if request.method == 'POST':
        try:
            deck = Deck.objects.get(pk=deck_id)
            data = json.loads(request.body)
            action = data.get('action')
            
            # Rows
            if action == 'add': # Keep existing key for rows
                deck.total_rows += 1
            elif action == 'remove' and deck.total_rows > 1:
                deck.total_rows -= 1
            
            # Columns (New)
            elif action == 'add_col':
                deck.grid_cols += 1
            elif action == 'remove_col' and deck.grid_cols > 1:
                deck.grid_cols -= 1
                
            deck.save()
            return JsonResponse({
                'status': 'success', 
                'new_rows': deck.total_rows,
                'new_cols': deck.grid_cols
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error'}, status=400)

@login_required
@csrf_exempt
def save_seat_layout(request, deck_id):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)

    deck = get_object_or_404(Deck, id=deck_id)

    try:
        data = json.loads(request.body)
        layout_items = data.get('layout', [])
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'Invalid JSON data'}, status=400)

    try:
        with transaction.atomic():
            # --- STEP 1: MAP EXISTING OBJECTS ---
            # We map (row, col) -> object so we can find them quickly.
            # This is crucial to PRESERVE IDs for existing bookings.
            existing_objects = {
                (obj.row_index, obj.col_index): obj 
                for obj in LayoutObject.objects.filter(deck=deck)
            }
            
            # Track which coordinates we have processed in this update
            processed_coords = set()

            # --- STEP 2: CREATE OR UPDATE ---
            for item in layout_items:
                row = int(item['row'])
                col = int(item['col'])
                coords = (row, col)
                processed_coords.add(coords)
                
                # Extract data from JSON
                category_id = item['category_id']
                label = item.get('label', '')
                row_span = int(item.get('row_span', 1))
                col_span = int(item.get('col_span', 1))
                feature_ids = item.get('feature_ids', [])

                # Check if we are updating an existing block or creating a new one
                obj = existing_objects.get(coords)
                
                if obj:
                    # UPDATE existing (Preserves Booking History)
                    obj.category_id = category_id
                    obj.label = label
                    obj.row_span = row_span
                    obj.col_span = col_span
                    # We save immediately to update basic fields
                    obj.save() 
                else:
                    # CREATE new
                    obj = LayoutObject.objects.create(
                        deck=deck,
                        row_index=row,
                        col_index=col,
                        category_id=category_id,
                        label=label,
                        row_span=row_span,
                        col_span=col_span
                    )
                
                # Update Many-to-Many Features (Tags)
                # .set() automatically handles add/remove of tags
                if feature_ids:
                    obj.features.set(feature_ids)
                else:
                    obj.features.clear()

            # --- STEP 3: DELETE REMOVED ITEMS ---
            # Any object that was in the DB but NOT in the new JSON payload must be deleted.
            for coords, obj in existing_objects.items():
                if coords not in processed_coords:
                    try:
                        obj.delete()
                    except ProtectedError:
                        # This happens if you try to delete a seat that has a Ticket.
                        # We fail safely and warn the admin.
                        raise Exception(f"Cannot delete '{obj.label}' at R{coords[0]}:C{coords[1]} because it has existing bookings.")

        return JsonResponse({'status': 'success', 'message': 'Layout saved successfully'})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    

# @staff_member_required
@login_required
def view_seat_plan(request, deck_id):
    deck = get_object_or_404(Deck, id=deck_id)
    
    # Fetch all objects. 'select_related' optimizes database access.
    layout_objects = LayoutObject.objects.filter(deck=deck).select_related('category')

    context = {
        'deck': deck,
        'layout_objects': layout_objects,
    }
    return render(request, 'admin_panel/seat_layout/view_seat_plan.html', context)


@login_required
@csrf_exempt
def locations(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            action = data.get('action')
            
            if action == 'add':
                name = data.get('name', '').strip()
                
                if not name:
                    return JsonResponse({'status': 'error', 'message': 'Location name is required!'}, status=400)
                
                # Validation: Name cannot be the same
                if Location.objects.filter(name__iexact=name).exists():
                     return JsonResponse({'status': 'error', 'message': 'This location name already exists!'}, status=400)

                # Slug is handled automatically in the model's save() method
                Location.objects.create(name=name)
                return JsonResponse({'status': 'success', 'message': 'Location added successfully!'})
            
            elif action == 'edit':
                loc_id = data.get('id')
                new_name = data.get('name', '').strip()
                
                if not new_name:
                    return JsonResponse({'status': 'error', 'message': 'Location name is required!'}, status=400)
                
                # Check for duplicates excluding the current location
                if Location.objects.filter(name__iexact=new_name).exclude(id=loc_id).exists():
                     return JsonResponse({'status': 'error', 'message': 'This location name already exists!'}, status=400)

                loc = get_object_or_404(Location, id=loc_id)
                loc.name = new_name
                loc.save() # The slug will auto-update if the name changed due to our model logic
                return JsonResponse({'status': 'success', 'message': 'Location updated successfully!'})
            
            elif action == 'delete':
                loc_id = data.get('id')
                loc = get_object_or_404(Location, id=loc_id)
                loc.delete()
                return JsonResponse({'status': 'success', 'message': 'Location deleted successfully!'})
                
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    locations = Location.objects.all().order_by('name')
    # Removed divisions as it's no longer needed for the template
    return render(request, 'admin_panel/routes/locations.html', {
        'locations': locations,
    })
    
    
    

@login_required
@csrf_exempt
def counters(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            action = data.get('action')

            if action == 'add':
                name = (data.get('name') or '').strip()
                location_id = data.get('location_id')
                user_ids = data.get('user_ids', [])  # NEW

                if not name:
                    return JsonResponse({'status': 'error', 'message': 'Counter name is required.'}, status=400)
                if not location_id:
                    return JsonResponse({'status': 'error', 'message': 'Location is required.'}, status=400)

                with transaction.atomic():
                    counter = Counter.objects.create(name=name, location_id=location_id)

                    # Assign selected users to this counter
                    if user_ids:
                        User.objects.filter(
                            id__in=user_ids,
                            user_type__in=[0, 2]  # Admin / Staff only
                        ).update(assigned_counter=counter)

                return JsonResponse({'status': 'success', 'message': 'Counter added successfully!'})

            elif action == 'edit':
                c_id = data.get('id')
                name = (data.get('name') or '').strip()
                location_id = data.get('location_id')
                user_ids = data.get('user_ids', [])  # NEW

                if not c_id:
                    return JsonResponse({'status': 'error', 'message': 'Counter ID missing.'}, status=400)
                if not name:
                    return JsonResponse({'status': 'error', 'message': 'Counter name is required.'}, status=400)
                if not location_id:
                    return JsonResponse({'status': 'error', 'message': 'Location is required.'}, status=400)

                with transaction.atomic():
                    counter = get_object_or_404(Counter, id=c_id)
                    counter.name = name
                    counter.location_id = location_id
                    counter.save()

                    # Clear old assignments from this counter first
                    User.objects.filter(assigned_counter=counter).update(assigned_counter=None)

                    # Reassign selected users
                    if user_ids:
                        User.objects.filter(
                            id__in=user_ids,
                            user_type__in=[0, 2]
                        ).update(assigned_counter=counter)

                return JsonResponse({'status': 'success', 'message': 'Counter updated successfully!'})

            elif action == 'delete':
                c_id = data.get('id')
                counter = get_object_or_404(Counter, id=c_id)

                with transaction.atomic():
                    # Unassign users first (safe + clean)
                    User.objects.filter(assigned_counter=counter).update(assigned_counter=None)
                    counter.delete()

                return JsonResponse({'status': 'success', 'message': 'Counter deleted successfully!'})

            return JsonResponse({'status': 'error', 'message': 'Invalid action.'}, status=400)

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    counters = Counter.objects.select_related('location').prefetch_related('users').all().order_by('location__name', 'name')
    locations = Location.objects.all().order_by('name')

    # Assignable users = Admin + Staff only (exclude customers)
    assignable_users = User.objects.filter(user_type__in=[0, 2]).order_by('first_name', 'phone_number')

    # Build JSON for frontend edit modal prefill
    counters_users_map = {
        str(counter.id): list(counter.users.values_list('id', flat=True))
        for counter in counters
    }

    context = {
        'counters': counters,
        'locations': locations,
        'assignable_users': assignable_users,  # NEW
        'counters_users_map': json.dumps(counters_users_map),  # NEW
    }
    return render(request, 'admin_panel/routes/counters.html', context)



@login_required
def routes(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            action = data.get('action')
            
            # --- ACTION: ADD ---
            if action == 'add':
                name = (data.get('name') or '').strip()
                source_id = int(data.get('source_id'))
                dest_id = int(data.get('destination_id'))
                arrival_offset_minutes = int(data.get('arrival_offset_minutes') or 0)
                
                # Validation
                if not name:
                    return JsonResponse({'status': 'error', 'message': 'Route name is required.'}, status=400)

                if source_id == dest_id:
                    return JsonResponse({'status': 'error', 'message': 'Source and Destination cannot be the same.'}, status=400)

                if arrival_offset_minutes <= 0:
                    return JsonResponse({'status': 'error', 'message': 'Arrival offset must be greater than 0 minutes.'}, status=400)

                with transaction.atomic():
                    route = Route.objects.create(
                        name=name,
                        source_id=source_id,
                        destination_id=dest_id
                    )

                    # Source stop (always 0 minutes from start)
                    RouteStop.objects.create(
                        route=route,
                        location_id=source_id,
                        stop_order=0,
                        time_offset_minutes=0
                    )

                    # Destination stop (total duration from source)
                    RouteStop.objects.create(
                        route=route,
                        location_id=dest_id,
                        stop_order=1,
                        time_offset_minutes=arrival_offset_minutes
                    )

                # Optional but recommended: generate route segment price skeleton immediately
                sync_route_prices(route)

                return JsonResponse({'status': 'success', 'message': 'Route created successfully!'})
            
            # --- ACTION: EDIT ---
            elif action == 'edit':
                r_id = data.get('id')
                new_name = (data.get('name') or '').strip()
                new_source_id = int(data.get('source_id'))
                new_dest_id = int(data.get('destination_id'))
                arrival_offset_minutes = int(data.get('arrival_offset_minutes') or 0)
                
                route = get_object_or_404(Route, id=r_id)
                
                # Validation
                if not new_name:
                    return JsonResponse({'status': 'error', 'message': 'Route name is required.'}, status=400)

                if new_source_id == new_dest_id:
                    return JsonResponse({'status': 'error', 'message': 'Source and Destination cannot be the same.'}, status=400)

                if arrival_offset_minutes <= 0:
                    return JsonResponse({'status': 'error', 'message': 'Destination arrival offset must be greater than 0 minutes.'}, status=400)

                old_source_id = route.source_id
                old_dest_id = route.destination_id
                
                with transaction.atomic():
                    # 1. Update Route metadata
                    route.name = new_name
                    route.source_id = new_source_id
                    route.destination_id = new_dest_id
                    route.save()

                    # 2. Sync Source stop (always stop_order=0)
                    source_stop = RouteStop.objects.select_for_update().get(route=route, stop_order=0)
                    source_stop.location_id = new_source_id
                    source_stop.time_offset_minutes = 0
                    source_stop.save(update_fields=['location_id', 'time_offset_minutes'])

                    # 3. Sync Destination stop (always the last stop)
                    dest_stop = RouteStop.objects.select_for_update().filter(route=route).order_by('-stop_order').first()
                    if not dest_stop:
                        return JsonResponse({'status': 'error', 'message': 'Route has no destination stop.'}, status=400)

                    # Validate destination offset against all previous stops
                    max_previous_offset = RouteStop.objects.filter(
                        route=route,
                        stop_order__lt=dest_stop.stop_order
                    ).aggregate(max_offset=models.Max('time_offset_minutes'))['max_offset'] or 0

                    if arrival_offset_minutes <= max_previous_offset:
                        return JsonResponse(
                            {
                                'status': 'error',
                                'message': f'Destination time offset must be greater than {max_previous_offset} minutes.'
                            },
                            status=400
                        )

                    dest_stop.location_id = new_dest_id
                    dest_stop.time_offset_minutes = arrival_offset_minutes
                    dest_stop.save(update_fields=['location_id', 'time_offset_minutes'])

                return JsonResponse({'status': 'success', 'message': 'Route updated successfully!'})
        
            # --- ACTION: DELETE ---
            elif action == 'delete':
                r_id = data.get('id')
                route = get_object_or_404(Route, id=r_id)
                try:
                    with transaction.atomic():
                        # 1. Force delete all Trips attached to this route first
                        # (This will also CASCADE delete any customer bookings for these trips)
                        Trip.objects.filter(route=route).delete()
                        
                        # 2. Now the route is unprotected and can be safely deleted
                        route.delete()
                        
                    return JsonResponse({'status': 'success', 'message': 'Route and all associated trips deleted successfully!'})
                except Exception as e:
                    return JsonResponse({'status': 'error', 'message': f'Error deleting route: {str(e)}'}, status=400)
                
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    # GET Request
    routes = Route.objects.select_related('source', 'destination').prefetch_related('stops').all().order_by('name')

    # Attach timing metadata for edit modal
    for r in routes:
        route_stops = list(r.stops.all().order_by('stop_order'))

        if len(route_stops) >= 1:
            # Destination is always the last stop logically
            dest_stop = route_stops[-1]
            r.destination_stop_offset = dest_stop.time_offset_minutes or 0

            # Highest offset among all previous stops (source + intermediates)
            previous_stops = route_stops[:-1]
            r.max_non_dest_offset = max((s.time_offset_minutes or 0) for s in previous_stops) if previous_stops else 0
        else:
            r.destination_stop_offset = 0
            r.max_non_dest_offset = 0
            
    locations = Location.objects.all().order_by('name')
    return render(request, 'admin_panel/routes/routes.html', {
        'routes': routes,
        'locations': locations
    })

@login_required
def route_details(request, route_id):
    route = get_object_or_404(Route, id=route_id)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            action = data.get('action')
            
            # --- ACTION: ADD STOP ---
            if action == 'add_stop':
                location_id = data.get('location_id')
                
                # 1. Validation
                if route.stops.filter(location_id=location_id).exists():
                    return JsonResponse({'status': 'error', 'message': 'Location already exists in this route!'}, status=400)

                with transaction.atomic():
                    # Get Destination Stop to determine insertion point
                    dest_stop = get_object_or_404(RouteStop, route=route, location=route.destination)
                    insert_index = dest_stop.stop_order
                    
                    # 2. SAFE SHIFT: Move existing stops down to make space.
                    # We fetch them ordered by '-stop_order' (DESCENDING)
                    # so we move the last one first, avoiding collisions.
                    stops_to_shift = route.stops.filter(
                        stop_order__gte=insert_index
                    ).order_by('-stop_order')

                    for stop in stops_to_shift:
                        stop.stop_order = stop.stop_order + 1
                        stop.save()
                    
                    # 3. Insert new stop in the now-empty slot
                    RouteStop.objects.create(
                        route=route, 
                        location_id=location_id, 
                        stop_order=insert_index,
                        time_offset_minutes=data.get('time_offset', 0)
                    )

                sync_route_prices(route)
                return JsonResponse({'status': 'success', 'message': 'Stop added successfully!'})

            # --- ACTION: DELETE STOP ---
            elif action == 'delete_stop':
                stop_id = data.get('id')
                stop = get_object_or_404(RouteStop, id=stop_id, route=route)
                
                # Validation: Don't delete Source/Dest
                if stop.location_id in [route.source.id, route.destination.id]:
                    return JsonResponse({'status': 'error', 'message': 'Cannot delete start or end points.'}, status=400)
                
                stop.delete() 
                # The signal automatically runs now. 
                # It detects the delete -> finds stops 3, 4, 5 -> updates them to 2, 3, 4.
                
                sync_route_prices(route)
                return JsonResponse({'status': 'success', 'message': 'Stop deleted successfully!'})

            # --- ACTION: REORDER STOPS ---
            elif action == 'reorder_stops':
                raw_ordered_ids = data.get('ordered_ids', [])
                ordered_ids = [int(x) for x in raw_ordered_ids if str(x).isdigit()]
                
                with transaction.atomic():
                    # 1. Integrity Check
                    current_stops = list(route.stops.values_list('id', flat=True))
                    if set(ordered_ids) != set(current_stops):
                         return JsonResponse({'status': 'error', 'message': 'Stop list mismatch. Please refresh.'}, status=400)

                    # 2. Logic Check: Ensure Source is first and Dest is last
                    # (Mapping IDs to verify logic)
                    id_to_loc = dict(route.stops.values_list('id', 'location_id'))
                    
                    if id_to_loc[ordered_ids[0]] != route.source.id:
                        return JsonResponse({'status': 'error', 'message': 'Source must remain the first stop.'}, status=400)
                    if id_to_loc[ordered_ids[-1]] != route.destination.id:
                         return JsonResponse({'status': 'error', 'message': 'Destination must remain the last stop.'}, status=400)

                    # 3. Two-Step Update (To avoid UniqueConstraint collisions on stop_order)
                    stops_to_update = []
                    
                    # Step A: Temporarily move to high numbers
                    for index, stop_id in enumerate(ordered_ids):
                         stop = RouteStop.objects.get(id=stop_id)
                         stop.stop_order = 10000 + index
                         stops_to_update.append(stop)
                    RouteStop.objects.bulk_update(stops_to_update, ['stop_order'])
                    
                    # Step B: Set to correct 0-indexed sequence
                    for index, stop in enumerate(stops_to_update):
                        stop.stop_order = index
                    RouteStop.objects.bulk_update(stops_to_update, ['stop_order'])
                    
                sync_route_prices(route)
                        
                return JsonResponse({'status': 'success', 'message': 'Sequence updated!'})
            
            elif action == 'save_prices':
                updates = data.get('prices', [])
                for item in updates:
                    RouteSegmentPricing.objects.filter(id=item.get('id'), route=route).update(price=item.get('price'))
                return JsonResponse({'status': 'success', 'message': 'Prices saved!'})
                
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    # Query Optimization: Fetch District info to avoid N+1 queries in the template
    # Query Optimization
    stops = route.stops.select_related('location', 'location__district').order_by('stop_order')
    locations = Location.objects.all().order_by('name')
    
    # --- PRICING LOGIC STARTS HERE ---
    segment_prices = RouteSegmentPricing.objects.filter(route=route)\
        .select_related('from_stop__location', 'to_stop__location', 'seat_category')\
        .order_by('from_stop__stop_order', 'to_stop__stop_order', 'seat_category__name')

    # 1. Check for unset prices (triggers the flash)
    prices_exist = segment_prices.exists()
    has_unset_prices = segment_prices.filter(
        price__lte=0, 
        seat_category__is_bookable=True
    ).exists()

    # If no prices exist but we have stops, run sync once to generate them
    if not prices_exist and stops.count() >= 2:
        sync_route_prices(route)
        has_unset_prices = True 
        # Refetch to show in modal immediately
        segment_prices = RouteSegmentPricing.objects.filter(route=route)\
            .select_related('from_stop__location', 'to_stop__location', 'seat_category')

    # 2. Organize data for the Modal
    price_matrix = {}
    for obj in segment_prices:
        label = f"{obj.from_stop.location.name} → {obj.to_stop.location.name}"
        
        if label not in price_matrix:
            price_matrix[label] = {
                'prices': [],
                'needs_attention': False
            }
        
        price_matrix[label]['prices'].append(obj)
        
        # If any bookable price in this segment is 0, mark the whole segment
        if obj.seat_category.is_bookable and obj.price <= 0:
            price_matrix[label]['needs_attention'] = True

    return render(request, 'admin_panel/routes/route_details.html', {
        'route': route,
        'stops': stops,
        'locations': locations,
        'price_matrix': price_matrix,       # <--- New context variable
        'has_unset_prices': has_unset_prices # <--- New context variable
    })
    
    
@login_required    
def trip_schedule_list(request):
    # Fetching all schedules with related ship and route data for performance
    schedules = TripSchedule.objects.select_related(
        'ship', 
        'route__source', 
        'route__destination'
    ).all().order_by('-id')
    
    context = {
        'schedules': schedules
    }
    return render(request, 'admin_panel/trips/trip_schedule_list.html', context)

    
def check_exact_time_conflict(ship_id, departure_dt):
    """
    HARD CONFLICT: Checks if the ship already has a trip starting at the EXACT same time.
    This should generally always be blocked to prevent duplicate identical dispatch times.
    """
    return Trip.objects.filter(
        ship_id=ship_id,
        departure_datetime=departure_dt
    ).select_related('route').first()

def check_overlap_conflict(ship_id, departure_dt, arrival_dt):
    """
    SOFT CONFLICT: Checks if the ship is scheduled for another route during this time window.
    Depending on business logic, this can be strictly blocked or overridden by an admin.
    """
    return Trip.objects.filter(
        ship_id=ship_id,
        departure_datetime__lt=arrival_dt,
        arrival_datetime__gt=departure_dt
    ).select_related('route').order_by('departure_datetime').first()

# --- MAIN VIEW ---

@login_required
def save_trip_schedule(request):
    if request.method == "POST":
        ship_id = request.POST.get('ship_id')
        route_id = request.POST.get('route_id')
        departure_time_str = request.POST.get('departure_time')
        date_list_str = request.POST.get('date_range')
        booking_close_offset_str = (request.POST.get('booking_close_offset_minutes') or '').strip()
        is_active = request.POST.get('is_active') == 'on'
        
        # 1. Capture the frontend override signal
        force_save = request.POST.get('force_save') == 'true'
        
        # 2. THE TOGGLE SWITCH:
        # Set to True -> Reverts to your old functionality (blocks ALL overlaps immediately).
        # Set to False -> New functionality (warns on overlaps, allows override via modal).
        STRICT_CONFLICT_CHECK = False 

        try:
            if not ship_id or not route_id or not departure_time_str:
                return JsonResponse({'success': False, 'message': 'Ship, route, and departure time are required.'}, status=400)

            dep_time_obj = datetime.strptime(departure_time_str, "%I:%M %p").time()

            booking_close_offset_minutes = None
            if booking_close_offset_str:
                try:
                    booking_close_offset_minutes = int(booking_close_offset_str)
                except (TypeError, ValueError):
                    return JsonResponse({'success': False, 'message': 'Booking close offset must be a valid whole number.'}, status=400)

                if booking_close_offset_minutes < 0:
                    return JsonResponse({'success': False, 'message': 'Booking close offset cannot be negative.'}, status=400)

                if booking_close_offset_minutes > 10080:
                    return JsonResponse({'success': False, 'message': 'Booking close offset is too large. Maximum allowed is 10080 minutes.'}, status=400)

            route = get_object_or_404(Route, id=route_id)
            dest_stop = RouteStop.objects.filter(route=route).order_by('-stop_order').first()

            if not dest_stop:
                return JsonResponse({'success': False, 'message': 'Route has no stops configured.'}, status=400)

            destination_offset_minutes = dest_stop.time_offset_minutes or 0
            if destination_offset_minutes <= 0:
                return JsonResponse({'success': False, 'message': 'Destination time offset must be greater than 0.'}, status=400)

            selected_dates = [d.strip() for d in (date_list_str or '').split(',') if d.strip()]
            if not selected_dates:
                return JsonResponse({'success': False, 'message': 'Please select at least one trip date.'}, status=400)

            proposed_trip_windows = []
            hard_conflict_messages = []
            soft_conflict_messages = []

            for date_str in selected_dates:
                date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
                departure_dt = datetime.combine(date_obj, dep_time_obj)
                arrival_dt = departure_dt + timedelta(minutes=destination_offset_minutes)

                # Check 1: Exact Time (Hard limit)
                exact_conflict = check_exact_time_conflict(ship_id, departure_dt)
                if exact_conflict:
                    hard_conflict_messages.append(
                        f"{date_str}: Exact departure time ({departure_dt.strftime('%I:%M %p')}) already exists for this ship."
                    )
                    continue # Skip adding to proposed windows

                # Check 2: Time Window Overlap (Soft or Hard limit based on toggle)
                overlap_conflict = check_overlap_conflict(ship_id, departure_dt, arrival_dt)
                if overlap_conflict:
                    msg = (f"{date_str}: Ship assigned to {overlap_conflict.route.name} "
                           f"({overlap_conflict.departure_datetime.strftime('%I:%M %p')} - "
                           f"{overlap_conflict.arrival_datetime.strftime('%I:%M %p')})")
                    
                    if STRICT_CONFLICT_CHECK:
                        hard_conflict_messages.append(msg)
                        continue # Skip adding to proposed windows
                    else:
                        soft_conflict_messages.append(msg)
                
                # If we get here, it's either clean, or it's a soft conflict we are allowing
                proposed_trip_windows.append((departure_dt, arrival_dt))

            # Handle response based on validation results
            if hard_conflict_messages:
                return JsonResponse({
                    'success': False,
                    'message': "Cannot create schedule. Hard conflict found:<br>" + "<br>".join(hard_conflict_messages)
                }, status=400)

            # If there are soft conflicts and the user hasn't clicked "Override" yet
            if soft_conflict_messages and not force_save:
                return JsonResponse({
                    'success': False,
                    'requires_confirmation': True,
                    'message': "<strong>Overlap Warning:</strong><br>" + "<br>".join(soft_conflict_messages) + "<br><br>Do you want to ignore this and save anyway?"
                }, status=200)

            # --- Validation Passed or Overridden. Save Data ---
            base_departure_dt = datetime.combine(datetime.today().date(), dep_time_obj)
            derived_arrival_dt = base_departure_dt + timedelta(minutes=destination_offset_minutes)

            schedule = TripSchedule.objects.create(
                ship_id=ship_id,
                route_id=route_id,
                departure_time=dep_time_obj,
                arrival_time=derived_arrival_dt.time(),
                is_active=is_active,
                booking_close_offset_minutes=booking_close_offset_minutes
            )

            for departure_dt, arrival_dt in proposed_trip_windows:
                Trip.objects.create(
                    schedule=schedule,
                    ship_id=ship_id,
                    route_id=route_id,
                    departure_datetime=departure_dt,
                    arrival_datetime=arrival_dt,
                    is_published=True,
                    booking_close_offset_minutes=booking_close_offset_minutes
                )

            return JsonResponse({'success': True, 'message': 'Schedule and trips created successfully!'})

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)

    # GET context remains exactly the same
    ships = Ship.objects.all()
    routes = Route.objects.all()
    days_mapping = [
        ('monday', 'Mon'), ('tuesday', 'Tue'), ('wednesday', 'Wed'),
        ('thursday', 'Thu'), ('friday', 'Fri'), ('saturday', 'Sat'), ('sunday', 'Sun')
    ]

    context = {
        'ships': ships,
        'routes': routes,
        'days_mapping': days_mapping
    }

    return render(request, 'admin_panel/trips/create_trip_schedule.html', context)


@login_required
def update_trip_schedule(request, schedule_id):
    schedule = get_object_or_404(TripSchedule, id=schedule_id)

    if request.method == "POST":
        ship_id = request.POST.get('ship_id')
        route_id = request.POST.get('route_id')
        departure_time_str = request.POST.get('departure_time')
        date_list_str = request.POST.get('date_range')
        is_active = request.POST.get('is_active') == 'on'

        try:
            if not ship_id or not route_id or not departure_time_str:
                return JsonResponse({'success': False, 'message': 'Ship, route, and departure time are required.'}, status=400)

            dep_time_obj = datetime.strptime(departure_time_str, "%I:%M %p").time()

            # ✅ Get destination stop and compute destination_offset_minutes
            route = get_object_or_404(Route, id=route_id)
            dest_stop = RouteStop.objects.filter(route=route).order_by('-stop_order').first()

            if not dest_stop:
                return JsonResponse({'success': False, 'message': 'Route has no stops configured.'}, status=400)

            destination_offset_minutes = dest_stop.time_offset_minutes or 0
            if destination_offset_minutes <= 0:
                return JsonResponse(
                    {'success': False, 'message': 'Destination time offset must be greater than 0 for this route.'},
                    status=400
                )

            # ✅ Parse selected dates
            selected_dates = [d.strip() for d in (date_list_str or '').split(',') if d.strip()]
            if not selected_dates:
                return JsonResponse({'success': False, 'message': 'Please select at least one trip date.'}, status=400)

            # ✅ Build windows + validate conflicts (excluding this schedule)
            proposed_trip_windows = []
            conflict_messages = []

            for date_str in selected_dates:
                date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
                departure_dt = datetime.combine(date_obj, dep_time_obj)
                arrival_dt = departure_dt + timedelta(minutes=destination_offset_minutes)

                conflicting_trip = Trip.objects.filter(
                    ship_id=ship_id,
                    departure_datetime__lt=arrival_dt,
                    arrival_datetime__gt=departure_dt
                ).exclude(
                    schedule=schedule
                ).select_related('route').order_by('departure_datetime').first()

                if conflicting_trip:
                    conflict_messages.append(
                        f"{date_str}: Ship already assigned to {conflicting_trip.route.name} "
                        f"({conflicting_trip.departure_datetime.strftime('%Y-%m-%d %I:%M %p')} - "
                        f"{conflicting_trip.arrival_datetime.strftime('%Y-%m-%d %I:%M %p')})"
                    )
                else:
                    proposed_trip_windows.append((departure_dt, arrival_dt))

            if conflict_messages:
                return JsonResponse({
                    'success': False,
                    'message': "Cannot update schedule. Ship timing conflict found: " + " | ".join(conflict_messages)
                }, status=400)

            # ✅ Save schedule metadata
            base_departure_dt = datetime.combine(datetime.today().date(), dep_time_obj)
            derived_arrival_dt = base_departure_dt + timedelta(minutes=destination_offset_minutes)

            schedule.ship_id = ship_id
            schedule.route_id = route_id
            schedule.departure_time = dep_time_obj
            schedule.arrival_time = derived_arrival_dt.time()
            schedule.is_active = is_active
            schedule.save()

            # ✅ Rebuild generated trips
            schedule.generated_trips.all().delete()

            for departure_dt, arrival_dt in proposed_trip_windows:
                Trip.objects.create(
                    schedule=schedule,
                    ship_id=ship_id,
                    route_id=route_id,
                    departure_datetime=departure_dt,
                    arrival_datetime=arrival_dt,
                    is_published=True
                )

            return JsonResponse({'success': True, 'message': 'Schedule and trips updated successfully!'})

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=400)

    existing_dates = schedule.generated_trips.values_list('departure_datetime__date', flat=True)
    formatted_dates = ", ".join([d.strftime('%Y-%m-%d') for d in existing_dates])

    return render(request, 'admin_panel/trips/update_trip_schedule.html', {
        'schedule': schedule,
        'ships': Ship.objects.all(),
        'routes': Route.objects.all(),
        'formatted_dates': formatted_dates,
    })

@login_required
@require_POST
def delete_trip_schedule(request, pk):
    try:
        schedule = get_object_or_404(TripSchedule, pk=pk)
        schedule.delete()
        return JsonResponse({'status': 'success', 'message': 'Schedule deleted successfully.'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
        
@login_required
def trip_list(request):
    # Optimizing to get Ship and Route names in one go
    trips = Trip.objects.select_related(
        'ship', 
        'route__source', 
        'route__destination',
        'schedule'
    ).all().order_by('-departure_datetime')
    
    return render(request, 'admin_panel/trips/trip_list.html', {'trips': trips})
    
@login_required
def individual_trip_management(request):
    # Use select_related here too for performance
    trips = Trip.objects.select_related(
        'ship', 
        'route__source', 
        'route__destination'
    ).all().order_by('-departure_datetime')

    date_range = request.GET.get('date_range')

    if date_range:
        # Splits "2026-01-07, 2026-01-08" into ['2026-01-07', '2026-01-08']
        selected_dates = [d.strip() for d in date_range.split(',') if d.strip()]
        # Backend filter on the date portion of the datetime field
        trips = trips.filter(departure_datetime__date__in=selected_dates)

    context = {
        'trips': trips,
        'date_range_value': date_range
    }
    
    # FIXED: Changed from 'your_app/trip_list.html' to the correct path below
    return render(request, 'admin_panel/trips/trip_list.html', context)

@login_required
def update_trip(request, trip_id):
    trip = get_object_or_404(Trip.objects.select_related('ship', 'route'), id=trip_id)

    route_stops = RouteStop.objects.filter(route=trip.route).order_by('stop_order')

    base_segments = RouteSegmentPricing.objects.filter(route=trip.route).select_related(
        'seat_category', 'from_stop__location', 'to_stop__location'
    )

    has_bookings = trip.tickets.filter(status__in=['BOOKED', 'LOCKED']).exists()

    if request.method == 'POST':
        try:
            with transaction.atomic():
                action = request.POST.get('action', 'save')
                force_confirm = str(request.POST.get('force_confirm', '')).lower() in ['1', 'true', 'yes']

                # ---------- ACTION: DELETE TRIP ----------
                if action == 'delete_trip':
                    if has_bookings and not force_confirm:
                        return JsonResponse({
                            'status': 'warning',
                            'origin': 'Booking Warning',
                            'message': 'This trip has booked/locked tickets. Deleting it may affect existing bookings. Do you want to continue?',
                            'requires_confirmation': True,
                            'confirm_action': 'delete_trip'
                        }, status=200)

                    trip.delete()
                    return JsonResponse({
                        'status': 'success',
                        'message': 'Trip deleted successfully.',
                        'redirect_url': reverse('trip_list')  # change if your URL name differs
                    })

                # ---------- ACTION: TOGGLE STATUS ----------
                if action == 'toggle_status':
                    requested_status = str(request.POST.get('is_published', '')).lower() in ['1', 'true', 'on']

                    if has_bookings and not force_confirm:
                        return JsonResponse({
                            'status': 'warning',
                            'origin': 'Booking Warning',
                            'message': f'This trip has booked/locked tickets. Are you sure you want to {"activate" if requested_status else "deactivate"} this trip?',
                            'requires_confirmation': True,
                            'confirm_action': 'toggle_status'
                        }, status=200)

                    trip.is_published = requested_status
                    trip.save(update_fields=['is_published', 'updated_at'])

                    return JsonResponse({
                        'status': 'success',
                        'message': f'Trip {"activated" if trip.is_published else "deactivated"} successfully.',
                        'is_published': trip.is_published
                    })

                # ---------- ACTION: SAVE (default) ----------
                new_date_str = request.POST.get('departure_datetime')
                new_multiplier = request.POST.get('price_multiplier', 1.0)
                booking_close_offset_str = (request.POST.get('booking_close_offset_minutes') or '').strip()

                # Pre-check booking warning for updates (date/time, offsets, pricing, etc.)
                if has_bookings and not force_confirm:
                    return JsonResponse({
                        'status': 'warning',
                        'origin': 'Booking Warning',
                        'message': 'This trip has booked/locked tickets. Updating date/time, itinerary offsets, or pricing may affect issued tickets. Do you want to continue?',
                        'requires_confirmation': True,
                        'confirm_action': 'save'
                    }, status=200)

                # --- VALIDATION: Date Change ---
                if new_date_str:
                    new_date = parse_datetime(new_date_str)
                    if not new_date:
                        return JsonResponse({
                            'status': 'error',
                            'origin': 'Validation',
                            'message': 'Invalid departure date/time format.'
                        }, status=400)

                    trip.departure_datetime = new_date
                    
                # --- VALIDATION: Trip-specific booking close offset ---
                booking_close_offset_minutes = None
                if booking_close_offset_str:
                    try:
                        booking_close_offset_minutes = int(booking_close_offset_str)
                    except (TypeError, ValueError):
                        return JsonResponse({
                            'status': 'error',
                            'origin': 'Validation',
                            'message': 'Booking close offset must be a valid whole number.'
                        }, status=400)

                    if booking_close_offset_minutes < 0:
                        return JsonResponse({
                            'status': 'error',
                            'origin': 'Validation',
                            'message': 'Booking close offset cannot be negative.'
                        }, status=400)

                    if booking_close_offset_minutes > 10080:
                        return JsonResponse({
                            'status': 'error',
                            'origin': 'Validation',
                            'message': 'Booking close offset is too large. Maximum allowed is 10080 minutes (7 days).'
                        }, status=400)

                # --- UPDATE CORE FIELDS ---
                trip.price_multiplier = new_multiplier
                trip.booking_close_offset_minutes = booking_close_offset_minutes
                trip.is_published = True  # keep existing behavior for save if you want
                trip.save()

                # --- UPDATE ITINERARY OFFSETS (ROUTE-LEVEL; affects all trips on route) ---
                # NOTE: This is route master data, not trip-specific.
                for stop in route_stops:
                    offset_val = request.POST.get(f'offset_{stop.id}')
                    if offset_val is not None:
                        if stop.stop_order == 0:
                            stop.time_offset_minutes = 0
                        else:
                            stop.time_offset_minutes = int(offset_val)
                        stop.save()

                # Recompute this trip's arrival based on destination offset
                destination_stop = route_stops.order_by('-stop_order').first()
                if destination_stop:
                    trip.arrival_datetime = trip.departure_datetime + timedelta(minutes=(destination_stop.time_offset_minutes or 0))
                    trip.save(update_fields=['arrival_datetime', 'updated_at'])

                # --- UPDATE PRICING ---
                for segment in base_segments:
                    price_val = request.POST.get(f'price_override_{segment.id}')
                    if price_val and price_val.strip() != "":
                        TripPricing.objects.update_or_create(
                            trip=trip,
                            seat_category=segment.seat_category,
                            from_stop=segment.from_stop,
                            to_stop=segment.to_stop,
                            defaults={'price': price_val}
                        )
                    else:
                        TripPricing.objects.filter(
                            trip=trip,
                            seat_category=segment.seat_category,
                            from_stop=segment.from_stop,
                            to_stop=segment.to_stop
                        ).delete()

                return JsonResponse({
                    'status': 'success',
                    'message': 'Trip updated successfully.'
                })

        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'origin': 'System Server',
                'message': str(e)
            }, status=500)

    # --- PREPARE DATA FOR DISPLAY ---
    itinerary = []
    last_stop = route_stops.last()
    for stop in route_stops:
        arrival_time = trip.departure_datetime + timedelta(minutes=stop.time_offset_minutes)
        itinerary.append({
            'stop_id': stop.id,
            'location': stop.location.name,
            'time': arrival_time,
            'offset': stop.time_offset_minutes,
            'is_start': stop.stop_order == 0,
            'is_end': (last_stop and stop.id == last_stop.id)
        })

    current_overrides = TripPricing.objects.filter(trip=trip)
    override_map = {(p.seat_category_id, p.from_stop_id, p.to_stop_id): p.price for p in current_overrides}

    for segment in base_segments:
        key = (segment.seat_category_id, segment.from_stop_id, segment.to_stop_id)
        segment.existing_override = override_map.get(key)
        segment.current_total_price = segment.existing_override if segment.existing_override else (segment.price * trip.price_multiplier)
        segment.is_fixed = bool(segment.existing_override)

    return render(request, 'admin_panel/trips/update_trip.html', {
        'trip': trip,
        'base_segments': base_segments,
        'has_bookings': has_bookings,
        'itinerary': itinerary
    })
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    













@login_required
def site_identity_view(request):
    identity = SiteIdentity.objects.first()
    
    if request.method == 'POST':
        # Upload new logo
        if request.FILES.get('logo'):
            if not identity:
                identity = SiteIdentity.objects.create()
            
            identity.logo = request.FILES.get('logo')
            identity.save()
            messages.success(request, "Logo Updated Successfully!")
        
        return redirect('site_identity')

    return render(request, 'admin_panel/home/identity.html', {'identity': identity})


@login_required
def banner(request):
    if request.method == 'POST':
        try:
            # We get data from request.POST because we are using FormData for images
            action = request.POST.get('action')

            if action == 'add':
                HomeBanner.objects.create(
                    title=request.POST.get('title'),
                    description=request.POST.get('description'),
                    # logo line removed here
                    background_image=request.FILES.get('background_image'),
                    is_active=request.POST.get('is_active') == 'on'
                )
                return JsonResponse({'status': 'success', 'message': 'Banner added successfully!'})

            elif action == 'edit':
                banner_id = request.POST.get('id')
                banner = HomeBanner.objects.get(id=banner_id)
                
                banner.title = request.POST.get('title')
                banner.description = request.POST.get('description')
                
                # Logo update block removed here
                
                # Only update background image if a new one is uploaded
                if request.FILES.get('background_image'):
                    banner.background_image = request.FILES.get('background_image')
                
                banner.is_active = request.POST.get('is_active') == 'on'
                banner.save()
                return JsonResponse({'status': 'success', 'message': 'Banner updated successfully!'})

            elif action == 'delete':
                banner_id = request.POST.get('id')
                HomeBanner.objects.get(id=banner_id).delete()
                return JsonResponse({'status': 'success', 'message': 'Banner deleted successfully!'})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    # GET Request: Render the list
    banners = HomeBanner.objects.all().order_by('-updated_at')
    return render(request, 'admin_panel/home/banner.html', {'banners': banners})




from .forms import CompanyOverviewForm
@login_required
def overview(request):
    """
    Manages the 'Company Overview' section on the Home Page.
    Acts as a singleton editor (always edits the first object).
    """
    
    # 1. Try to get the existing record (we only want one 'About Us' section)
    obj = CompanyOverview.objects.first()

    if request.method == 'POST':
        # 2. If data is sent, bind it to the form (and the object if it exists)
        form = CompanyOverviewForm(request.POST, request.FILES, instance=obj)
        
        if form.is_valid():
            form.save()
            messages.success(request, "Home Page Overview updated successfully!")
            return redirect('overview') # Reload the page to show changes
        else:
            messages.error(request, "Please correct the errors below.")
            
    else:
        # 3. If GET request, show the form with existing data (if any)
        form = CompanyOverviewForm(instance=obj)

    # 4. Prepare context for the template
    context = {
        'form': form,
        # Pass the image separately so we can show a preview in the HTML
        'existing_image': obj.image if obj else None 
    }
    
    return render(request, 'admin_panel/home/overview.html', context)



from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from .models import ContactBanner # Import the model created in Part 1
@login_required
def contact_banner_view(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        
        try:
            # --- ADD BANNER ---
            if action == 'add':
                title = request.POST.get('title')
                image = request.FILES.get('background_image')
                # Checkbox sends 'on' if checked, None if not.
                is_active = request.POST.get('is_active') == 'on' 

                if not title or not image:
                    return JsonResponse({'status': 'error', 'message': 'Title and Image are required.'})

                # Ensure only one banner is active if this one is set to active
                if is_active:
                    ContactBanner.objects.update(is_active=False)

                ContactBanner.objects.create(
                    title=title,
                    background_image=image,
                    is_active=is_active
                )
                return JsonResponse({'status': 'success', 'message': 'Banner added successfully!'})

            # --- EDIT BANNER ---
            elif action == 'edit':
                banner_id = request.POST.get('id')
                banner = get_object_or_404(ContactBanner, id=banner_id)
                
                banner.title = request.POST.get('title')
                
                # Update image only if a new one is uploaded
                if 'background_image' in request.FILES:
                    banner.background_image = request.FILES['background_image']
                
                is_active = request.POST.get('is_active') == 'on'
                
                # Logic to handle active state
                if is_active:
                    ContactBanner.objects.exclude(id=banner.id).update(is_active=False)
                
                banner.is_active = is_active
                banner.save()
                
                return JsonResponse({'status': 'success', 'message': 'Banner updated successfully!'})

            # --- DELETE BANNER ---
            elif action == 'delete':
                banner_id = request.POST.get('id')
                banner = get_object_or_404(ContactBanner, id=banner_id)
                banner.delete()
                return JsonResponse({'status': 'success', 'message': 'Banner deleted successfully!'})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    # --- GET REQUEST (Render Page) ---
    banners = ContactBanner.objects.all().order_by('-id')
    context = {
        'banners': banners
    }
    return render(request, 'admin_panel/contactus/banner.html', context)


@login_required
def contact_messages_view(request):
    if request.method == 'POST':
        action = request.POST.get('action')

        try:
            # --- DELETE MESSAGE ---
            if action == 'delete':
                msg_id = request.POST.get('id')
                message = get_object_or_404(ContactMessage, id=msg_id)
                message.delete()
                return JsonResponse({'status': 'success', 'message': 'Message deleted successfully!'})
            
            # --- MARK AS READ (Optional, triggered via JS when viewing) ---
            elif action == 'mark_read':
                msg_id = request.POST.get('id')
                message = get_object_or_404(ContactMessage, id=msg_id)
                message.is_read = True
                message.save()
                return JsonResponse({'status': 'success'})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    # --- GET REQUEST (Render List) ---
    # Show newest messages first
    messages = ContactMessage.objects.all().order_by('-created_at')
    
    context = {
        'messages': messages
    }
    return render(request, 'admin_panel/contactus/messages.html', context)

@login_required
def contact_info_cards_view(request):
    if request.method == 'POST':
        action = request.POST.get('action')

        try:
            # --- ADD CARD ---
            if action == 'add':
                title = request.POST.get('title')
                description = request.POST.get('description')
                contact_info = request.POST.get('contact_info')
                icon = request.FILES.get('icon')
                is_active = request.POST.get('is_active') == 'on'

                ContactInfoCard.objects.create(
                    title=title, description=description, 
                    contact_info=contact_info, icon=icon, is_active=is_active
                )
                return JsonResponse({'status': 'success', 'message': 'Info Card added successfully!'})

            # --- EDIT CARD ---
            elif action == 'edit':
                card_id = request.POST.get('id')
                card = get_object_or_404(ContactInfoCard, id=card_id)

                card.title = request.POST.get('title')
                card.description = request.POST.get('description')
                card.contact_info = request.POST.get('contact_info')
                card.is_active = request.POST.get('is_active') == 'on'
                
                if request.FILES.get('icon'):
                    card.icon = request.FILES.get('icon')

                card.save()
                return JsonResponse({'status': 'success', 'message': 'Info Card updated successfully!'})

            # --- DELETE CARD ---
            elif action == 'delete':
                card_id = request.POST.get('id')
                ContactInfoCard.objects.filter(id=card_id).delete()
                return JsonResponse({'status': 'success', 'message': 'Info Card deleted successfully!'})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    # --- GET REQUEST ---
    cards = ContactInfoCard.objects.all()
    return render(request, 'admin_panel/contactus/info_cards.html', {'cards': cards})

@login_required
def contact_map_view(request):
    # Fetch the existing map object (if any)
    map_obj = ContactMap.objects.first()

    if request.method == 'POST':
        embed_code = request.POST.get('map_embed_code')
        is_active = request.POST.get('is_active') == 'on'

        if map_obj:
            # Update existing
            map_obj.map_embed_code = embed_code
            map_obj.is_active = is_active
            map_obj.save()
            messages.success(request, "Map updated successfully!")
        else:
            # Create new
            ContactMap.objects.create(map_embed_code=embed_code, is_active=is_active)
            messages.success(request, "Map created successfully!")
        
        return redirect('contact_map')

    return render(request, 'admin_panel/contactus/map.html', {'map_obj': map_obj})

@login_required
def contact_faq_view(request):
    # Fetch Settings (Create if doesn't exist to avoid errors)
    section_settings = ContactFAQSection.objects.first()
    if not section_settings:
        section_settings = ContactFAQSection.objects.create(title="Have questions?")

    if request.method == 'POST':
        action = request.POST.get('action')

        try:
            # --- 1. UPDATE SECTION SETTINGS ---
            if action == 'update_settings':
                section_settings.title = request.POST.get('title')
                section_settings.is_active = request.POST.get('is_active') == 'on'
                
                if request.FILES.get('side_image'):
                    section_settings.side_image = request.FILES.get('side_image')
                
                section_settings.save()
                return JsonResponse({'status': 'success', 'message': 'Settings updated!'})

            # --- 2. ADD FAQ ITEM ---
            elif action == 'add_item':
                ContactFAQItem.objects.create(
                    question=request.POST.get('question'),
                    answer=request.POST.get('answer'),
                    order=request.POST.get('order', 0),
                    is_active=request.POST.get('is_active') == 'on'
                )
                return JsonResponse({'status': 'success', 'message': 'FAQ added!'})

            # --- 3. EDIT FAQ ITEM ---
            elif action == 'edit_item':
                item = get_object_or_404(ContactFAQItem, id=request.POST.get('id'))
                item.question = request.POST.get('question')
                item.answer = request.POST.get('answer')
                item.order = request.POST.get('order', 0)
                item.is_active = request.POST.get('is_active') == 'on'
                item.save()
                return JsonResponse({'status': 'success', 'message': 'FAQ updated!'})

            # --- 4. DELETE FAQ ITEM ---
            elif action == 'delete_item':
                ContactFAQItem.objects.filter(id=request.POST.get('id')).delete()
                return JsonResponse({'status': 'success', 'message': 'FAQ deleted!'})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    # GET Request
    faq_items = ContactFAQItem.objects.all().order_by('order')
    return render(request, 'admin_panel/contactus/faq.html', {
        'settings': section_settings,
        'faq_items': faq_items
    })
    
@login_required    
def about_banner_view(request):
    banner = AboutBanner.objects.first()
    
    if request.method == 'POST':
        title = request.POST.get('title')
        is_active = request.POST.get('is_active') == 'on'
        
        if banner:
            banner.title = title
            banner.is_active = is_active
            if request.FILES.get('background_image'):
                banner.background_image = request.FILES.get('background_image')
            banner.save()
            messages.success(request, "About Banner Updated!")
        else:
            img = request.FILES.get('background_image')
            AboutBanner.objects.create(title=title, background_image=img, is_active=is_active)
            messages.success(request, "About Banner Created!")
        return redirect('about_banner')
        
    return render(request, 'admin_panel/about/banner.html', {'banner': banner})

from .forms import AboutStoryForm
@login_required
def about_story_view(request):
    story = AboutStory.objects.first()

    if request.method == 'POST':
        form = AboutStoryForm(request.POST, request.FILES, instance=story)
        if form.is_valid():
            form.save()
            messages.success(request, "About Us content updated!")
            return redirect('about_story')
    else:
        form = AboutStoryForm(instance=story)

    return render(request, 'admin_panel/about/story.html', {'form': form, 'story': story})






# --- 1. MAIN GALLERY ADMIN VIEW ---
@login_required
def gallery_main_view(request):
    settings = GallerySection.objects.first()
    if not settings:
        settings = GallerySection.objects.create()

    if request.method == 'POST':
        action = request.POST.get('action')

        # Update Settings
        if action == 'update_settings':
            settings.subtitle = request.POST.get('subtitle')
            settings.title = request.POST.get('title')
            settings.description = request.POST.get('description')
            settings.save()
            messages.success(request, "Settings Updated")

        # Add Image
        elif action == 'add_image':
            if request.FILES.get('image'):
                GalleryImage.objects.create(
                    image=request.FILES.get('image'),
                    order=request.POST.get('order', 0)
                )
                messages.success(request, "Image Added")

        # Delete Image
        elif action == 'delete_image':
            GalleryImage.objects.filter(id=request.POST.get('id')).delete()
            messages.success(request, "Image Deleted")
            
        return redirect('gallery_main')

    images = GalleryImage.objects.all()
    return render(request, 'admin_panel/gallery/main.html', {'settings': settings, 'images': images})


# --- 2. SEASONAL TOURS ADMIN VIEW ---
@login_required
def gallery_seasonal_view(request):
    settings = SeasonalSection.objects.first()
    if not settings:
        settings = SeasonalSection.objects.create()

    if request.method == 'POST':
        action = request.POST.get('action')

        # Update Settings
        if action == 'update_settings':
            settings.subtitle = request.POST.get('subtitle')
            settings.title = request.POST.get('title')
            settings.save()
            messages.success(request, "Settings Updated")

        # Add Tour
        elif action == 'add_tour':
            SeasonalTour.objects.create(
                title=request.POST.get('title'),
                link=request.POST.get('link'),
                image=request.FILES.get('image'),
                order=request.POST.get('order', 0)
            )
            messages.success(request, "Tour Added")

        # --- NEW: EDIT TOUR LOGIC ---
        elif action == 'edit_tour':
            tour_id = request.POST.get('id')
            tour = SeasonalTour.objects.get(id=tour_id)
            
            tour.title = request.POST.get('title')
            tour.link = request.POST.get('link')
            tour.order = request.POST.get('order')
            
            # Only update image if a new one is selected
            if request.FILES.get('image'):
                tour.image = request.FILES.get('image')
            
            tour.save()
            messages.success(request, "Tour Updated Successfully")
        
        
        # Delete Tour
        elif action == 'delete_tour':
            SeasonalTour.objects.filter(id=request.POST.get('id')).delete()
            messages.success(request, "Tour Deleted")

        return redirect('gallery_seasonal')

    tours = SeasonalTour.objects.all()
    return render(request, 'admin_panel/gallery/seasonal.html', {'settings': settings, 'tours': tours})


@login_required
def blog_banner_update(request):
    # Try to get the first existing banner, or create one if none exists
    banner = BlogBanner.objects.first()
    
    if request.method == 'POST':
        form = BlogBannerForm(request.POST, request.FILES, instance=banner)
        if form.is_valid():
            form.save()
            messages.success(request, "Blog Banner Updated Successfully!")
            return redirect('admin_blog_banner')
    else:
        # If banner doesn't exist yet, we pass None so it creates a new one
        form = BlogBannerForm(instance=banner)

    return render(request, 'admin_panel/blog/banner_form.html', {
        'form': form,
        'current_banner': banner
    })





#team
from .forms import TeamMemberForm
@login_required
def manage_team(request):
    """
    Allows the admin to:
    1. See a list of all members.
    2. Add a new member (Name, Designation, Description, Image).
    """
    # Get all current members to show in the table
    members = TeamMember.objects.all().order_by('-created_at')
    
    # Handle the "Add New Member" Form
    if request.method == 'POST':
        form = TeamMemberForm(request.POST, request.FILES)
        if form.is_valid():
            form.save() # This saves Name, Designation, Description, and Image
            messages.success(request, "New Team Member Added Successfully!")
            return redirect('manage_team')
        else:
            messages.error(request, "Error adding member. Please check the form.")
    else:
        form = TeamMemberForm()

    context = {
        'members': members,
        'form': form
    }
    return render(request, 'admin_panel/team/manage_team.html', context)

# ==========================================
# 3. DELETE VIEW (Admin Action)
# ==========================================
@login_required
def delete_team_member(request, pk):
    """
    Deletes a specific team member by their ID (pk).
    """
    member = get_object_or_404(TeamMember, pk=pk)
    member.delete()
    messages.success(request, "Team member deleted.")
    return redirect('manage_team')







# 1. LIST VIEW
@login_required
def blog_list(request):
    blogs = BlogPost.objects.all().order_by('-date')
    return render(request, 'admin_panel/blog/list.html', {'blogs': blogs})

# 2. ADD VIEW
@login_required
def blog_add(request):
    if request.method == 'POST':
        form = BlogPostForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Blog Post Added Successfully!")
            return redirect('admin_blog_list')
    else:
        form = BlogPostForm()
    
    return render(request, 'admin_panel/blog/form.html', {'form': form, 'title': 'Add New Blog'})

# 3. EDIT VIEW
@login_required
def blog_edit(request, slug):
    blog = get_object_or_404(BlogPost, slug=slug)
    
    if request.method == 'POST':
        form = BlogPostForm(request.POST, request.FILES, instance=blog)
        if form.is_valid():
            form.save()
            messages.success(request, "Blog Updated Successfully!")
            return redirect('admin_blog_list')
    else:
        form = BlogPostForm(instance=blog)

    return render(request, 'admin_panel/blog/form.html', {'form': form, 'title': 'Edit Blog', 'blog': blog})

# 4. DELETE VIEW
@login_required
def blog_delete(request, slug):
    blog = get_object_or_404(BlogPost, slug=slug)
    blog.delete()
    messages.success(request, "Blog Deleted Successfully!")
    return redirect('admin_blog_list')

from django.db.models import Count
@login_required
def admin_comment_list(request):
    posts = BlogPost.objects.annotate(
        total_comments=Count('comments')
    ).filter(total_comments__gt=0).prefetch_related('comments').order_by('-date')
    
    context = {
        'posts': posts, # We are sending POSTS now
        'title': 'Blog Comments'
    }
    return render(request, 'admin_panel/blog/comments.html', context)

@login_required
def admin_comment_delete(request, id):
    comment = get_object_or_404(BlogComment, id=id)
    comment.delete()
    messages.success(request, "Comment deleted successfully!")
    # Redirect back to the comments list
    return redirect('admin_comment_list')

@login_required
def get_search_locations(request):
    # Fetch all locations, selecting only necessary fields for speed
    locations = list(Location.objects.values('id', 'name', 'code'))
    return JsonResponse({'status': 'success', 'data': locations})

@login_required
def admin_user_list(request):
    # Fetch all users, newest first
    users = User.objects.exclude(user_type=1).order_by('-created_at')
    
    context = {
        'users': users,
        'title': 'User Management'
    }
    return render(request, 'admin_panel/users/user_list.html', context)


from django.contrib.auth.models import Permission
from accounts.models import User
@login_required
def user_add(request):
    
    # 3. SAFETY CHECK
    if not hasattr(request.user, 'user_type'):
         messages.error(request, "Error: Your account information is incomplete.")
         return redirect('home')

    # 4. ADMIN CHECK
    if not request.user.is_superuser and request.user.user_type != 0:
         messages.error(request, "Access denied. Admins only.")
         return redirect('home')

    if request.method == 'POST':
        form = AdminUserAddForm(request.POST)
        
        if form.is_valid():
            try:
                # 1. Save the new user first
                new_user = form.save()
                
                # 2. Permission Saving Logic
                # We look for the custom checkboxes named 'permissions[]'
                selected_permission_ids = request.POST.getlist('permissions[]')
                
                # Convert strings to integers safely
                selected_ids = [int(p_id) for p_id in selected_permission_ids if p_id.isdigit()]
                
                if selected_ids:
                    # Fetch permission objects and assign to the NEW user
                    permissions = Permission.objects.filter(id__in=selected_ids)
                    new_user.user_permissions.set(permissions)
                
                messages.success(request, f"User {new_user.email} created successfully!")
                return redirect('admin_user_list')
            except IntegrityError:
                # Fallback in case a duplicate slips past form validation during save.
                form.add_error(None, "A user with this email or phone number already exists. Please correct the duplicate field and try again.")
            except Exception as e:
                # Surface unexpected errors on the form so the admin knows what to fix.
                form.add_error(None, f"Something went wrong while creating the user: {str(e)}")
        else:
            # If form is invalid, we need to reload the permission form
            perm_form = AdminUserPermissionsForm(instance=User()) 
    else:
        form = AdminUserAddForm()
        # Initialize permission form with an empty User instance (so no boxes are checked by default)
        perm_form = AdminUserPermissionsForm(instance=User())

    context = {
        'form': form,
        'perm_form': perm_form # Passing this to template is crucial
    }

    return render(request, 'admin_panel/users/add.html', context)




from .decorators import admin_only

@admin_only  # <--- Ensures only Admins can access this page
def admin_user_edit(request, id):
    user_obj = get_object_or_404(User, id=id)
    
    if request.method == 'POST':
        # 1. Standard Info Form (Existing logic)
        form = AdminUserEditForm(request.POST, instance=user_obj)
        
        # 2. We don't necessarily need to bind POST data to perm_form for saving 
        # because we are manually reading the checkboxes below.
        
        if form.is_valid():
            user = form.save()
            
            # --- START PERMISSION SAVING LOGIC ---
            # We look for the custom checkboxes named 'permissions[]' from the HTML
            selected_permission_ids = request.POST.getlist('permissions[]')
            
            # Convert the list of strings ['1', '5'] to integers [1, 5]
            selected_ids = [int(p_id) for p_id in selected_permission_ids if p_id.isdigit()]
            
            if selected_ids:
                # Fetch the actual Permission objects and assign them
                permissions = Permission.objects.filter(id__in=selected_ids)
                user.user_permissions.set(permissions)
            else:
                # If no boxes were checked, clear all permissions for this user
                user.user_permissions.clear()
            # --- END PERMISSION SAVING LOGIC ---

            messages.success(request, f"User '{user.email}' and permissions updated successfully.")
            return redirect('admin_user_list')
        else:
            messages.error(request, "Please correct the errors below.")
            
            # If there's an error, we need to re-initialize the perm_form so the
            # checkboxes still show up in the view
            perm_form = AdminUserPermissionsForm(instance=user_obj)

    else:
        # GET Request: Load both forms
        form = AdminUserEditForm(instance=user_obj)
        
        # This triggers the "Smart Grouping" logic in your form's __init__
        perm_form = AdminUserPermissionsForm(instance=user_obj)

    context = {
        'form': form,
        'perm_form': perm_form,  # <--- PASS THIS TO TEMPLATE
        'user_obj': user_obj,
        'title': f'Edit User: {user_obj.first_name}'
    }
    return render(request, 'admin_panel/users/edit.html', context)


@login_required
def admin_user_delete(request, id):
    user_to_delete = get_object_or_404(User, id=id)
    
    # SAFETY: Prevent deleting yourself (the currently logged-in admin)
    if user_to_delete == request.user:
        messages.error(request, "You cannot delete your own admin account!")
        return redirect('admin_user_list')

    # Delete the user
    user_to_delete.delete()
    messages.success(request, "User deleted successfully.")
    return redirect('admin_user_list')


@login_required
def profile_view(request):
    # Ensure a UserProfile exists for the user so we can access the picture without errors
    profile, created = UserProfile.objects.get_or_create(user=request.user)
    
    context = {
        'profile': profile,
    }
    return render(request, 'admin_panel/profile/edit_profile.html', context)

@login_required
def update_profile(request):
    if request.method == 'POST':
        user = request.user
        
        # 1. Update First & Last Name
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        
        user.first_name = first_name
        user.last_name = last_name
        user.save()

        # 2. Update Profile Picture
        # Get or create profile just in case
        profile, created = UserProfile.objects.get_or_create(user=user)
        
        if 'profile_pic' in request.FILES:
            profile.profile_picture = request.FILES['profile_pic']
            profile.save()

        messages.success(request, 'Profile updated successfully.')
        return redirect('admin_profile')
        
    return redirect('admin_profile')

@login_required
def change_password(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            current_password = data.get('current_password')
            new_password = data.get('new_password')
            
            user = request.user

            # 1. Verify current password
            if not user.check_password(current_password):
                return JsonResponse({'success': False, 'message': 'Incorrect current password.'})

            # 2. Validate new password requirements (backend validation)
            if len(new_password) < 8:
                return JsonResponse({'success': False, 'message': 'Password must be at least 8 characters long.'})
            if not any(char.isupper() for char in new_password):
                return JsonResponse({'success': False, 'message': 'Password must contain at least one uppercase letter.'})
            if not any(char.islower() for char in new_password):
                return JsonResponse({'success': False, 'message': 'Password must contain at least one lowercase letter.'})
            if not any(char.isdigit() for char in new_password):
                return JsonResponse({'success': False, 'message': 'Password must contain at least one number.'})

            # 3. Save new password
            user.set_password(new_password)
            user.save()
            
            # Keep the user logged in after password change
            update_session_auth_hash(request, user)

            return JsonResponse({'success': True, 'message': 'Password changed successfully.'})

        except Exception as e:
            return JsonResponse({'success': False, 'message': 'An error occurred while processing your request.'})

    return JsonResponse({'success': False, 'message': 'Invalid request method.'})



@login_required
def tcktbook(request):
    search_results = []
    form = TripSearchForm(request.GET or None) # Bind data if it exists in URL
    
    # These will be useful for the template to know what we searched for
    searched_from = None
    searched_to = None

    if form.is_valid():
        source = form.cleaned_data['from_location']
        destination = form.cleaned_data['to_location']
        date = form.cleaned_data['journey_date']
        
        searched_from = source
        searched_to = destination

        # --- LOGIC START ---
        
        # 1. Find Routes that have BOTH the source and destination
        # We find RouteStops matching source, and RouteStops matching destination
        # Then we check which Routes appear in BOTH lists.
        
        routes_with_source = RouteStop.objects.filter(location=source).values_list('route_id', 'stop_order')
        routes_with_dest = RouteStop.objects.filter(location=destination).values_list('route_id', 'stop_order')
        
        # Convert to dictionaries for easier lookup: {route_id: stop_order}
        source_map = {r_id: order for r_id, order in routes_with_source}
        dest_map = {r_id: order for r_id, order in routes_with_dest}
        
        valid_route_ids = []
        
        # 2. Compare Orders: Source must be BEFORE Destination
        for route_id, source_order in source_map.items():
            if route_id in dest_map:
                dest_order = dest_map[route_id]
                if source_order < dest_order:
                    valid_route_ids.append(route_id)
        
        # 3. Find TRIPS for these valid routes on the specific date
        if valid_route_ids:
            search_results = Trip.objects.filter(
                route_id__in=valid_route_ids,
                departure_datetime__date=date
            ).select_related('ship', 'route')
            
        # --- LOGIC END ---

    context = {
        'form': form,
        'trips': search_results,
        'searched_from': searched_from,
        'searched_to': searched_to,
    }
    return render(request, 'admin_panel/book/book.html', context)



@login_required
def select_seats(request, trip_id):
    trip = get_object_or_404(Trip, id=trip_id)
    
    # 1. Get Route Segment Info from GET params
    from_loc_id = request.GET.get('from_loc')
    to_loc_id = request.GET.get('to_loc')
    
    if not from_loc_id or not to_loc_id:
        messages.error(request, "Please select source and destination first.")
        return redirect('admin_search_trips') # Replace with your search url name

    # 2. Get the RouteStop objects to determine order (Stop 1 -> Stop 5)
    try:
        from_stop = RouteStop.objects.get(route=trip.route, location_id=from_loc_id)
        to_stop = RouteStop.objects.get(route=trip.route, location_id=to_loc_id)
    except RouteStop.DoesNotExist:
        messages.error(request, "Invalid route stops.")
        return redirect('admin_home')

    # 3. Calculate Pricing for each Category (Map ID -> Price)
    # We'll create a dictionary to pass to JS: { 'AC Cabin': 1500, 'Deck': 300 }
    category_prices = {}
    
    # We fetch all unique categories used in this ship's layout
    # This avoids calculating prices for categories that don't exist on this ship
    layout_categories = LayoutObject.objects.filter(deck__ship=trip.ship).values_list('category', flat=True).distinct()
    
    for cat_id in layout_categories:
        # We need the actual category object to pass to get_price if your logic expects objects
        # Or if get_price expects ID, pass ID. Assuming it takes the object:
        from .models import SeatCategory
        category = SeatCategory.objects.get(id=cat_id)
        price = trip.get_price(category, from_stop, to_stop)
        category_prices[cat_id] = float(price) # Convert Decimal to float for JSON

    # 4. Determine Availability (The Overlap Logic)
    # Fetch ALL unavailable tickets in one optimized query
    unavailable_tickets = Ticket.objects.filter(
        trip=trip,
        status__in=['BOOKED', 'CONFIRMED', 'LOCKED', 'PENDING']
    ).filter(
        Q(from_stop__stop_order__lt=to_stop.stop_order) & 
        Q(to_stop__stop_order__gt=from_stop.stop_order)
    ).prefetch_related('passenger')

    booked_seat_ids = []
    pending_seat_ids = []
    locked_seat_ids = [] # NEW: Dedicated list for locked seats
    seat_gender_map = {}

    # Sort tickets into the correct lists to keep frontend styles accurate
    for ticket in unavailable_tickets:
        if ticket.status == 'PENDING':
            pending_seat_ids.append(ticket.seat_object_id)
        elif ticket.status == 'LOCKED':
            locked_seat_ids.append(ticket.seat_object_id)
        else:
            # Treats BOOKED and CONFIRMED as fully booked
            booked_seat_ids.append(ticket.seat_object_id)
            
            # Map gender only for actual bookings
            gender = 0   # Default to Male
            if getattr(ticket, 'passenger_id', None):
                try:
                    if ticket.passenger and ticket.passenger.gender is not None:
                        gender = ticket.passenger.gender
                except Exception:
                    pass # Failsafe if the passenger record is missing/corrupted
            
            seat_gender_map[ticket.seat_object_id] = gender

    # 4.5) Active holds for overlapping segment (so admin UI stays in sync)
    holder_id = get_holder_id(request)

    active_holds = SeatHold.objects.filter(
        trip=trip,
        expires_at__gt=timezone.now()
    ).filter(
        Q(from_stop__stop_order__lt=to_stop.stop_order) &
        Q(to_stop__stop_order__gt=from_stop.stop_order)
    )

    held_seat_ids = list(active_holds.values_list('seat_object_id', flat=True).distinct())

    held_by_map = {
        h.seat_object_id: h.holder_id
        for h in active_holds.only('seat_object_id', 'holder_id')
    }
    
    # ========== Debugging Logs ==========
    print("\n" + "=" * 80)
    print("ADMIN SELECT_SEATS DEBUG")
    print(f"Trip ID: {trip.id}")
    print(f"Route: {trip.route.name}")
    print(f"Requested segment: {from_stop.location.name} ({from_stop.stop_order}) -> {to_stop.location.name} ({to_stop.stop_order})")
    print(f"Admin holder_id (current user): {holder_id}")
    
    print(f"============================================= Locked seat IDs ==========================================\n: {locked_seat_ids}")

    # Show ALL active holds on this trip (before overlap filter)
    all_trip_holds = SeatHold.objects.filter(
        trip=trip,
        expires_at__gt=timezone.now()
    ).select_related('from_stop', 'to_stop', 'seat_object')

    print(f"All active holds on trip: {all_trip_holds.count()}")
    for h in all_trip_holds:
        print(
            f"  HOLD seat_id={h.seat_object_id} label={getattr(h.seat_object, 'label', None)} "
            f"holder_id={h.holder_id} "
            f"segment={h.from_stop.location.name}({h.from_stop.stop_order})->{h.to_stop.location.name}({h.to_stop.stop_order}) "
            f"expires_at={h.expires_at}"
        )

    # Show only overlapping holds (the ones admin UI should display)
    overlap_holds = active_holds.select_related('from_stop', 'to_stop', 'seat_object')
    print(f"Overlapping holds for requested segment: {overlap_holds.count()}")
    for h in overlap_holds:
        print(
            f"  OVERLAP seat_id={h.seat_object_id} label={getattr(h.seat_object, 'label', None)} "
            f"holder_id={h.holder_id} "
            f"segment={h.from_stop.location.name}({h.from_stop.stop_order})->{h.to_stop.location.name}({h.to_stop.stop_order})"
        )

    print(f"held_seat_ids passed to template: {held_seat_ids}")
    print(f"held_by_map passed to template: {held_by_map}")
    print("=" * 80 + "\n")
    # =========== End of Debugging Logs ===========

    # 5. Fetch Layout grouped by Deck
    decks = trip.ship.decks.all().order_by('level_order')
    
    context = {
        'trip': trip,
        'from_stop': from_stop,
        'to_stop': to_stop,
        'decks': decks,
        'booked_seat_ids': booked_seat_ids,
        'pending_seat_ids': pending_seat_ids,
        'locked_seat_ids': locked_seat_ids,
        'category_prices': category_prices,
        'seat_gender_map': seat_gender_map,
        'held_seat_ids': held_seat_ids,
        'held_by_map': held_by_map,
        'holder_id': holder_id,
    }
    return render(request, 'admin_panel/book/select_seats.html', context)



@login_required
def admin_book_confirm(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method.'})
    
    # Check if request is AJAX
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if not is_ajax:
        return JsonResponse({'success': False, 'message': 'Expected AJAX request.'})

    gender = int(request.POST.get('passenger_gender', 0))
    # --- 1. GET BASIC DATA ---
    trip_id = request.POST.get('trip_id')
    seat_ids_str = request.POST.get('selected_seats')
    
    # Action type determines status ('ISSUE' or 'BOOK')
    action_type = request.POST.get('action_type') 
    
    # Customer Data
    c_phone = request.POST.get('customer_phone')
    c_email = request.POST.get('customer_email')
    c_name = request.POST.get('customer_name')
    
    # Route Data
    from_stop_id = request.POST.get('from_stop_id')
    to_stop_id = request.POST.get('to_stop_id')
    
    # Payment Amount
    manual_amount_str = request.POST.get('manual_amount')
    
    if not seat_ids_str:
        return JsonResponse({'success': False, 'message': 'No seats selected.'})

    # --- 2. GET DYNAMIC DURATION ---
    try:
        hold_duration_minutes = int(request.POST.get('hold_duration', 120))
    except (ValueError, TypeError):
        hold_duration_minutes = 120

    paid_amount_input = 0.0
    if manual_amount_str:
        try:
            paid_amount_input = float(manual_amount_str)
        except ValueError:
            paid_amount_input = 0.0

    trip = get_object_or_404(Trip, id=trip_id)
    seat_ids = seat_ids_str.split(',')
    from_stop = get_object_or_404(RouteStop, id=from_stop_id)
    to_stop = get_object_or_404(RouteStop, id=to_stop_id)

    # --- 3. DETERMINE STATUS & EXPIRY BASED ON BUTTON CLICKED ---
    if action_type == 'ISSUE':
        final_status = 'CONFIRMED'
        expiry_time = None
    else: # 'BOOK' clicked
        final_status = 'PENDING'
        expiry_time = timezone.now() + timedelta(minutes=hold_duration_minutes)

    try:
        with transaction.atomic():
            # A. Handle User
            booking_user = request.user 
            if c_name and c_phone:
                user = User.objects.filter(phone_number=c_phone).first()
                if not user and c_email:
                    user = User.objects.filter(email=c_email).first()

                if not user:
                    final_email = c_email if c_email else f"{c_phone}@guest.com"
                    random_pass = get_random_string(length=12)
                    user = User.objects.create_user(
                        email=final_email,
                        username=c_phone,
                        phone_number=c_phone,
                        first_name=c_name,
                        password=random_pass,
                        user_type=1
                    )
                booking_user = user

            paid_amount = paid_amount_input
            
            # B. Create Booking
            operator_counter = get_logged_in_counter(request.user)

            if final_status == 'CONFIRMED':
                issued_by_user = request.user
            else:
                issued_by_user = None
                
            booked_by_user = request.user

            booking = Booking.objects.create(
                user=booking_user,
                trip=trip,
                booking_ref=str(uuid.uuid4())[:12].upper(),
                status=final_status,
                payment_status='UNPAID', # Default, will calculate accurately below
                expiry_at=expiry_time,
                sales_channel='COUNTER' if operator_counter else 'ONLINE',
                counter=operator_counter,
                issued_by=issued_by_user,
                booked_by=booked_by_user,
                total_amount=0,
                paid_amount=paid_amount
            )
            
            # Create Passenger
            passenger = None
            if c_name and c_phone:
                passenger = Passenger.objects.create(
                    booking=booking,
                    user=booking_user,
                    name=c_name,
                    phone=c_phone,
                    email=c_email,
                    gender=gender
                )
            else:
                raise Exception("Passenger name and phone are required")

            # C. Create Tickets
            calculated_total = 0
            booked_seat_labels = [] 
            
            # ---> ADD THIS: Get the exact holder ID <---
            admin_holder_id = get_holder_id(request) 
            
            for seat_id in seat_ids:
                layout_obj = get_object_or_404(LayoutObject, id=seat_id)
                
                # ---> UPDATE THIS: Pass the exclude_holder_id <---
                if not trip.is_seat_available_admin(layout_obj, from_stop, to_stop, exclude_user=request.user, exclude_holder_id=admin_holder_id):
                    raise Exception(f"Seat {layout_obj.label} is currently held by someone else!")

                price = trip.get_price(layout_obj.category, from_stop, to_stop)

                # Determine ticket status based on the final_status of the booking
                ticket_status = 'BOOKED' if final_status == 'CONFIRMED' else 'PENDING'

                Ticket.objects.create(
                    booking=booking,
                    passenger=passenger,   
                    trip=trip,
                    seat_object=layout_obj,
                    from_stop=from_stop,
                    to_stop=to_stop,
                    passenger_name=c_name if c_name else "Walk-in Guest",
                    fare_amount=price,
                    status=ticket_status, # <-- Now dynamically assigned
                    lock_expires_at=timezone.now(),
                )
                
                # ---> BROADCAST to CHANNEL <---
                channel_layer = get_channel_layer()
                admin_holder_id = get_holder_id(request) # Get the exact holder ID
                
                # Determine the correct WebSocket action
                ws_action = 'booked' if final_status == 'CONFIRMED' else 'pending'
                
                async_to_sync(channel_layer.group_send)(
                    f"seats_{trip_id}",  # 1. FIXED: Group name must match portal
                    {
                        'type': 'seat_event', # 2. FIXED: Type must match portal consumer
                        'action': ws_action,  # <-- CHANGED: Send dynamic action
                        'seat_id': int(layout_obj.id),
                        'holder_id': admin_holder_id, # 3. Added for consistency
                        'start_order': from_stop.stop_order, # 4. REQUIRED for route logic
                        'end_order': to_stop.stop_order,     # 5. REQUIRED for route logic
                        'gender': gender
                    }
                )
                
                SeatHold.objects.filter(
                    trip=trip,
                    seat_object=layout_obj,
                    from_stop=from_stop,
                    to_stop=to_stop
                ).delete()
                
                calculated_total += price
                booked_seat_labels.append(layout_obj.label)

            # D. Update amounts dynamically
            booking.total_amount = calculated_total

            # Clamp paid amount
            if booking.paid_amount < 0:
                booking.paid_amount = 0
            if booking.paid_amount > booking.total_amount:
                booking.paid_amount = booking.total_amount

            # E. Automatically set PAID vs UNPAID based strictly on amount
            if booking.total_amount > 0 and booking.paid_amount >= booking.total_amount:
                booking.payment_status = 'PAID'
            else:
                booking.payment_status = 'UNPAID'

            booking.save()
            
            # F. Send SMS (Non-blocking)
            if booking_user.phone_number:
                try:
                    send_booking_sms(booking, booked_seat_labels)
                except Exception as e:
                    print(f"SMS Error: {e}")

            # G. TRIGGER AUTO-CANCEL (DYNAMIC TIMER)
            if booking.status == 'PENDING':
                countdown_seconds = hold_duration_minutes * 60 
                auto_cancel_booking.apply_async((booking.id,), countdown=countdown_seconds)

            msg_text = f"Successfully { 'issued' if final_status == 'CONFIRMED' else 'booked' }. Ref: {booking.booking_ref}"
            
            # Return JSON for JS to handle instead of redirect
            return JsonResponse({'success': True, 'message': msg_text})

    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})



#-------------------new api endpoint ----------
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt

@login_required
@csrf_exempt
def check_seat_availability(request):
    """
    API endpoint to check if a seat is available
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Get values
            trip_id = data.get('trip_id')
            seat_id = data.get('seat_id')
            from_stop_id = data.get('from_stop')
            to_stop_id = data.get('to_stop')
            
            print(f"Received: trip={trip_id}, seat={seat_id}, from={from_stop_id}, to={to_stop_id}")
            
            # Validate
            if not all([trip_id, seat_id, from_stop_id, to_stop_id]):
                return JsonResponse({
                    'available': False,
                    'error': 'Missing required fields'
                }, status=400)
            
            # Convert to integers
            trip_id = int(trip_id)
            seat_id = int(seat_id)
            from_stop_id = int(from_stop_id)
            to_stop_id = int(to_stop_id)
            
            # Get objects
            trip = get_object_or_404(Trip, id=trip_id)
            seat = get_object_or_404(LayoutObject, id=seat_id)
            from_stop = get_object_or_404(RouteStop, id=from_stop_id)
            to_stop = get_object_or_404(RouteStop, id=to_stop_id)
            
            # 1. Fetch the holder ID FIRST
            admin_holder_id = get_holder_id(request)
            
            # 2. Pass BOTH exclude_user AND exclude_holder_id
            available = trip.is_seat_available_admin(
                seat, 
                from_stop, 
                to_stop, 
                exclude_user=request.user,
                exclude_holder_id=admin_holder_id
            )
            
            if available:
                return JsonResponse({'available': True})
            
            # If not available, find why
            hold = SeatHold.objects.filter(
                trip=trip,
                seat_object=seat,
                expires_at__gt=timezone.now()
            ).exclude(holder_id=admin_holder_id).first()
            
            if hold:
                return JsonResponse({
                    'available': False,
                    'message': 'This seat is already held by someone else'
                })
            
            # Check Tickets
            ticket = Ticket.objects.filter(
                trip=trip,
                seat_object=seat,
                status__in=['BOOKED', 'CONFIRMED', 'LOCKED']
            ).first()
            
            if ticket:
                return JsonResponse({
                    'available': False,
                    'message': 'This seat is already booked'
                })
            
            return JsonResponse({
                'available': False,
                'message': 'This seat is not available'
            })
            
        except json.JSONDecodeError:
            return JsonResponse({
                'available': False,
                'error': 'Invalid JSON'
            }, status=400)
        except ValueError as e:
            return JsonResponse({
                'available': False,
                'error': f'Invalid ID format: {str(e)}'
            }, status=400)
        except Exception as e:
            print(f"Error in check_seat_availability: {e}")
            return JsonResponse({
                'available': False,
                'error': str(e)
            }, status=500)
    
    return JsonResponse({'error': 'Invalid method'}, status=405)


@login_required
@csrf_exempt
def create_seat_hold(request):
    """
    API endpoint to create a hold on a seat
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            trip_id = data.get('trip_id')
            seat_id = data.get('seat_id')
            from_stop_id = data.get('from_stop')
            to_stop_id = data.get('to_stop')
            
            if not all([trip_id, seat_id, from_stop_id, to_stop_id]):
                return JsonResponse({
                    'success': False,
                    'error': 'Missing required fields'
                }, status=400)
            
            # Convert to integers
            trip_id = int(trip_id)
            seat_id = int(seat_id)
            from_stop_id = int(from_stop_id)
            to_stop_id = int(to_stop_id)
            
            trip = get_object_or_404(Trip, id=trip_id)
            seat = get_object_or_404(LayoutObject, id=seat_id)
            from_stop = get_object_or_404(RouteStop, id=from_stop_id)
            to_stop = get_object_or_404(RouteStop, id=to_stop_id)
            
            # Check availability
            if not trip.is_seat_available_admin(seat, from_stop, to_stop, exclude_user=request.user):
                return JsonResponse({'success': False, 'message': 'Seat is no longer available'})
            
            # Convert admin user ID to string
            admin_user_id = get_holder_id(request)
            
            # 🔴 FIXED: Use holder_id, not holder
            SeatHold.objects.filter(
                trip=trip,
                seat_object=seat,
                holder_id=admin_user_id
            ).delete()
            
            # Create new hold
            hold = SeatHold.objects.create(
                trip=trip,
                seat_object=seat,
                holder_id=admin_user_id,  # Using holder_id
                from_stop=from_stop,
                to_stop=to_stop,
                expires_at=timezone.now() + timedelta(minutes=5)
            )
            
            # ---> ADD THIS BROADCAST <---
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                f"seats_{trip_id}",  # 🔴 Match the portal group name
                {
                    "type": "seat_event", # 🔴 Match the portal event type
                    "action": "hold",
                    "seat_id": seat.id,
                    "holder_id": admin_user_id,
                    "expires_at": hold.expires_at.isoformat(),
                    "start_order": from_stop.stop_order,
                    "end_order": to_stop.stop_order,
                }
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Seat held successfully',
                'expires_at': hold.expires_at.isoformat()
            })
            
        except ValueError as e:
            return JsonResponse({
                'success': False,
                'error': f'Invalid ID format: {str(e)}'
            }, status=400)
        except Exception as e:
            print(f"Error in create_seat_hold: {e}")
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)
    
    return JsonResponse({'error': 'Invalid method'}, status=405)


@login_required
@csrf_exempt
def release_seat_hold(request):
    """
    API endpoint to release a hold on a seat
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            trip_id = data.get('trip_id')
            seat_id = data.get('seat_id')
            
            # ✅ NEW: We need the stops to broadcast the segment correctly!
            from_stop_id = data.get('from_stop')
            to_stop_id = data.get('to_stop')
            
            # Validate
            if not all([trip_id, seat_id, from_stop_id, to_stop_id]):
                return JsonResponse({
                    'success': False, 
                    'error': 'Missing required fields'
                }, status=400)
            
            # Convert to integers
            trip_id = int(trip_id)
            seat_id = int(seat_id)
            from_stop_id = int(from_stop_id)
            to_stop_id = int(to_stop_id)
            
            # Get stops to determine order for the broadcast
            from_stop = get_object_or_404(RouteStop, id=from_stop_id)
            to_stop = get_object_or_404(RouteStop, id=to_stop_id)
            
            # Convert admin user ID to string
            admin_user_id = get_holder_id(request)
            
            deleted_count = SeatHold.objects.filter(
                trip_id=trip_id,
                seat_object_id=seat_id,
                holder_id=admin_user_id
            ).delete()
            
            # ✅ FIXED: Broadcast to the exact same group with the exact same payload shape
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                f"seats_{trip_id}",  # Match portal group name
                {
                    "type": "seat_event", # Match portal event type
                    "action": "release",
                    "seat_id": seat_id,
                    "start_order": from_stop.stop_order,
                    "end_order": to_stop.stop_order,
                }
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Hold released successfully',
                'deleted': deleted_count[0] if deleted_count else 0
            })
            
        except ValueError as e:
            return JsonResponse({
                'success': False, 
                'error': f'Invalid ID format: {str(e)}'
            }, status=400)
        except Exception as e:
            print(f"Error in release_seat_hold: {e}")
            return JsonResponse({
                'success': False, 
                'error': str(e)
            }, status=500)
    
    return JsonResponse({'error': 'Invalid method'}, status=405)



#--------------------------------------------
@login_required
@csrf_exempt
def check_multiple_seats_availability(request):
    """
    API endpoint to check multiple seats availability (for frontend compatibility)
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            trip_id = data.get('trip_id')
            seat_ids = data.get('seat_ids', [])
            from_stop_id = data.get('from_stop')
            to_stop_id = data.get('to_stop')
            
            if not all([trip_id, from_stop_id, to_stop_id]) or not seat_ids:
                return JsonResponse({'error': 'Missing required fields'}, status=400)
            
            # Convert to integers
            trip_id = int(trip_id)
            from_stop_id = int(from_stop_id)
            to_stop_id = int(to_stop_id)
            
            trip = get_object_or_404(Trip, id=trip_id)
            from_stop = get_object_or_404(RouteStop, id=from_stop_id)
            to_stop = get_object_or_404(RouteStop, id=to_stop_id)
            
            admin_user_id = str(request.user.id)
            results = {}
            
            for seat_id in seat_ids:
                seat_id = int(seat_id)
                seat = get_object_or_404(LayoutObject, id=seat_id)
                
                # Check availabili
                available = trip.is_seat_available_admin(seat, from_stop, to_stop, exclude_user=request.user)
                
                if available:
                    results[seat_id] = {'available': True}
                else:
                    # Check if held by someone else
                    hold = SeatHold.objects.filter(
                        trip=trip,
                        seat_object=seat,
                        expires_at__gt=timezone.now()
                    ).exclude(holder_id=admin_user_id).first()
                    
                    if hold:
                        results[seat_id] = {
                            'available': False,
                            'message': 'This seat is held by someone else'
                        }
                    else:
                        ticket = Ticket.objects.filter(
                            trip=trip,
                            seat_object=seat,
                            status__in=['BOOKED', 'CONFIRMED', 'LOCKED']
                        ).first()
                        
                        if ticket:
                            results[seat_id] = {
                                'available': False,
                                'message': 'This seat is already booked'
                            }
                        else:
                            results[seat_id] = {
                                'available': False,
                                'message': 'This seat is not available'
                            }
            
            return JsonResponse({'results': results})
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Invalid method'}, status=405)


@login_required
@csrf_exempt
def create_multiple_seat_holds(request):
    """
    API endpoint to create holds for multiple seats (for frontend compatibility)
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            trip_id = data.get('trip_id')
            from_stop_id = data.get('from_stop')
            to_stop_id = data.get('to_stop')
            seat_ids = data.get('seat_ids', [])
            
            if not all([trip_id, from_stop_id, to_stop_id]) or not seat_ids:
                return JsonResponse({'success': False, 'error': 'Missing required fields'}, status=400)
            
            # Convert to integers
            trip_id = int(trip_id)
            from_stop_id = int(from_stop_id)
            to_stop_id = int(to_stop_id)
            
            trip = get_object_or_404(Trip, id=trip_id)
            from_stop = get_object_or_404(RouteStop, id=from_stop_id)
            to_stop = get_object_or_404(RouteStop, id=to_stop_id)
            
            admin_user_id = str(request.user.id)
            expires_at = timezone.now() + timedelta(minutes=5)
            
            held = []
            rejected = []
            
            for seat_id in seat_ids:
                seat_id = int(seat_id)
                seat = get_object_or_404(LayoutObject, id=seat_id)
                
                # Check availability
                if not trip.is_seat_available_admin(seat, from_stop, to_stop, exclude_user=request.user):
                    rejected.append({'seat_id': seat_id, 'reason': 'unavailable'})
                    continue
                
                # Delete any existing holds by this admin for this seat
                SeatHold.objects.filter(
                    trip=trip,
                    seat_object=seat,
                    holder_id=admin_user_id
                ).delete()
                
                # Create new hold
                SeatHold.objects.create(
                    trip=trip,
                    seat_object=seat,
                    holder_id=admin_user_id,
                    from_stop=from_stop,
                    to_stop=to_stop,
                    expires_at=expires_at
                )
                
                held.append(seat_id)
            
            return JsonResponse({
                'success': True,
                'held': held,
                'rejected': rejected
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Invalid method'}, status=405)


@login_required
@csrf_exempt
def release_multiple_seat_holds(request):
    """
    API endpoint to release multiple seat holds (for frontend compatibility)
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            trip_id = data.get('trip_id')
            from_stop_id = data.get('from_stop')
            to_stop_id = data.get('to_stop')
            seat_ids = data.get('seat_ids', [])
            
            if not all([trip_id, from_stop_id, to_stop_id]) or not seat_ids:
                return JsonResponse({'success': False, 'error': 'Missing required fields'}, status=400)
            
            trip_id = int(trip_id)
            admin_user_id = str(request.user.id)
            
            deleted = SeatHold.objects.filter(
                trip_id=trip_id,
                seat_object_id__in=seat_ids,
                holder_id=admin_user_id
            ).delete()
            
            return JsonResponse({
                'success': True,
                'released': seat_ids
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Invalid method'}, status=405)


#








@require_POST
def extend_booking_time_api(request):
    """
    New API specifically for the 'Add Time' button in the booking list.
    """
    try:
        data = json.loads(request.body)
        booking_id = data.get('booking_id')
        minutes_to_add = int(data.get('minutes', 30))

        booking = Booking.objects.get(id=booking_id)

        # Safety Check: Can only extend Pending or Expired
        if booking.status not in ['PENDING', 'EXPIRED']:
             return JsonResponse({'success': False, 'message': 'Cannot extend a Confirmed or Cancelled booking.'})

        now = timezone.now()

        # Logic: 
        # 1. If currently valid: Add to existing expiry
        # 2. If already expired: Reset to NOW + minutes
        if booking.expiry_at and booking.expiry_at > now:
            booking.expiry_at += timedelta(minutes=minutes_to_add)
        else:
            booking.expiry_at = now + timedelta(minutes=minutes_to_add)
            # If it was marked EXPIRED, revive it to PENDING
            if booking.status == 'EXPIRED':
                booking.status = 'PENDING'
        
        booking.save()

        return JsonResponse({
            'success': True,
            'new_expiry': booking.expiry_at.isoformat(),
            'message': 'Time extended successfully'
        })

    except Booking.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Booking ID not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})



@login_required
@require_POST
def stop_booking_time(request, booking_id):
    """
    Stop the countdown timer for a pending booking.
    Prevents auto-cancellation without confirming the booking.
    """
    try:
        data = json.loads(request.body) if request.body else {}
        booking = get_object_or_404(Booking, id=booking_id)
        
        # Security: Can only stop time for PENDING bookings
        if booking.status != 'PENDING':
            return JsonResponse({
                'success': False, 
                'message': 'Can only stop time for PENDING bookings.'
            })
        
        # Check if time is already stopped
        if booking.time_stopped:
            return JsonResponse({
                'success': False,
                'message': 'Time is already stopped for this booking.'
            })
        
        # Stop the time
        booking.time_stopped = True
        booking.stopped_at = timezone.now()
        booking.stopped_by = request.user
        booking.save()
        
        # Optional: Cancel any pending Celery task
        # You might need to store task_id in the booking model
        # if hasattr(booking, 'celery_task_id') and booking.celery_task_id:
        #     from celery.task.control import revoke
        #     revoke(booking.celery_task_id, terminate=False)
        
        return JsonResponse({
            'success': True,
            'message': 'Timer stopped successfully. Booking will not auto-cancel.',
            'stopped_at': booking.stopped_at.isoformat()
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
@require_POST
def resume_booking_time(request, booking_id):
    """
    Resume the countdown timer for a stopped booking.
    """
    try:
        data = json.loads(request.body) or {}
        minutes_to_add = int(data.get('minutes', 120))  # Default 2 hours
        
        booking = get_object_or_404(Booking, id=booking_id)
        
        if not booking.time_stopped:
            return JsonResponse({
                'success': False,
                'message': 'Time is not stopped for this booking.'
            })
        
        # Resume with new expiry time
        booking.time_stopped = False
        booking.stopped_at = None
        booking.stopped_by = None
        booking.expiry_at = timezone.now() + timedelta(minutes=minutes_to_add)
        booking.save()
        
        # Restart Celery task
        from .tasks import auto_cancel_booking
        countdown_seconds = minutes_to_add * 60
        auto_cancel_booking.apply_async((booking.id,), countdown=countdown_seconds)
        
        return JsonResponse({
            'success': True,
            'message': f'Timer resumed. New expiry: {booking.expiry_at.isoformat()}',
            'expiry_at': booking.expiry_at.isoformat()
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# --- 2. NEW VIEW: QUICK STATUS UPDATE (For the list page) ---
@login_required
def update_booking_status(request, booking_id, new_status):
    booking = get_object_or_404(Booking, id=booking_id)
    
    # 1. Handle CONFIRM
    if new_status == 'CONFIRMED':
        # Retrieve the paid amount entered in the modal (fallback to existing if none)
        try:
            new_paid_amount = float(request.GET.get('paid_amount', booking.paid_amount))
        except ValueError:
            new_paid_amount = float(booking.paid_amount)
            
        booking.paid_amount = new_paid_amount
        
        # Calculate payment_status dynamically
        if booking.paid_amount >= booking.total_amount:
            booking.payment_status = 'PAID'
        else:
            booking.payment_status = 'UNPAID'

        booking.status = 'CONFIRMED'
        booking.expiry_at = None
        booking.time_stopped = False
        booking.stopped_at = None

        # Set issued_by to the user who is confirming it
        booking.issued_by = request.user
        
        booking.save()
        messages.success(request, "Booking Confirmed.")

    # 2. Handle CANCEL (Add this block)
    # elif new_status == 'CANCELLED':
    #     booking.status = 'CANCELLED'
        
    #     # ADDED: Set cancelled_by to the user performing the cancellation
    #     booking.cancelled_by = request.user
        
    #     # Optional: You might want to set payment_status to 'REFUNDED' or 'UNPAID'
    #     booking.save() 
    #     Ticket.objects.filter(booking=booking).update(status='CANCELLED')
    
    # 2. Handle CANCEL (Add this block)
    elif new_status == 'CANCELLED':
        
        # --- ADDED: Authorization Check for Cancellation ---
        is_admin = (request.user.user_type == 0)
        has_perm = request.user.has_perm('admin_panel.delete_booking')
        is_staff_issuer = (request.user.user_type == 2 and booking.issued_by == request.user)
        is_staff_booker = (request.user.user_type == 2 and not booking.issued_by and booking.booked_by == request.user)

        if not (is_admin or has_perm or is_staff_issuer or is_staff_booker):
            messages.error(request, "You do not have permission to cancel this booking.")
            return redirect('booking_issue_list')
        # ---------------------------------------------------

        booking.status = 'CANCELLED'
        
        # ADDED: Set cancelled_by to the user performing the cancellation
        booking.cancelled_by = request.user
        
        # Optional: You might want to set payment_status to 'REFUNDED' or 'UNPAID'
        booking.save() 
        Ticket.objects.filter(booking=booking).update(status='CANCELLED')

    # --- SEND SMS FOR BOTH CASES ---
    try:
        # Get seat labels for the SMS
        tickets = Ticket.objects.filter(booking=booking)
        seat_labels = [t.seat_object.label for t in tickets]
        
        # This will now send the correct SMS based on the new status
        send_booking_sms(booking, seat_labels)
        
    except Exception as e:
        print(f"SMS Failed: {e}")
        messages.error(request, "System error while updating status.")
    # -------------------------------
    
    return redirect('booking_issue_list')



# --- NEW VIEW: HANDLE SEAT DETAILS MODAL ---
@login_required
def get_seat_details(request, trip_id, seat_id):
    try:
        trip = Trip.objects.get(id=trip_id)
        seat = LayoutObject.objects.get(id=seat_id)
        
        data = {
            'seat_number': seat.label,
            'status': 'Available',
            'passenger_name': 'N/A',
            'passenger_phone': 'N/A',
            'amount': 0,
            'payment_status': 'N/A'  # <--- Default value
        }

        ticket = Ticket.objects.filter(
            trip=trip, 
            seat_object=seat,
            status__in=['BOOKED', 'CONFIRMED', 'LOCKED']
        ).first()

        if ticket:
            data['status'] = ticket.status
            data['passenger_name'] = ticket.passenger_name or "Walk-in"
            data['amount'] = ticket.fare_amount
            
            # Check Booking for Phone AND Payment Status
            if ticket.booking:
                if ticket.booking.user:
                    data['passenger_phone'] = ticket.booking.user.phone_number
                else:
                    data['passenger_phone'] = "Guest / Counter"
                
                # <--- NEW: Get Payment Status (Ensure your model field name is correct)
                data['payment_status'] = ticket.booking.payment_status 
            else:
                data['passenger_phone'] = "Counter Manual"
                data['payment_status'] = "Paid (Counter)" # Assume counter sales are paid

        return JsonResponse(data)

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)



@require_POST
@login_required
def toggle_trip_lock(request, trip_id):
    trip = get_object_or_404(Trip, id=trip_id)
    
    # ADDED: Get channel layer for WebSockets
    channel_layer = get_channel_layer() 
    
    # Moved Route Stop logic up so BOTH unlock and lock can use it for the broadcast payload
    route_stops = RouteStop.objects.filter(route=trip.route).order_by('stop_order')
    if not route_stops.exists():
        return JsonResponse({'status': 'error', 'message': 'Route has no stops defined.'})

    start_stop = route_stops.first()
    end_stop = route_stops.last()

    # 1. CHECK: Is it currently locked by Admin?
    locked_booking = Booking.objects.filter(trip=trip, status='LOCKED').first()
    
    if locked_booking:
        # --- UNLOCK ACTION ---
        
        # ADDED: Grab the seat IDs *before* we delete the booking so we can broadcast them
        locked_tickets = Ticket.objects.filter(booking=locked_booking)
        seat_ids_to_release = list(locked_tickets.values_list('seat_object_id', flat=True))

        locked_booking.delete()
        print(f"Unlocked Trip {trip_id} and released all seats.")
        
        # ADDED: Broadcast the 'release' event for every unlocked seat
        for seat_id in seat_ids_to_release:
            async_to_sync(channel_layer.group_send)(
                f"seats_{trip_id}",
                {
                    "type": "seat_event",
                    "action": "release",  # Triggers your existing JS release logic
                    "seat_id": seat_id,
                    "start_order": start_stop.stop_order,
                    "end_order": end_stop.stop_order,
                }
            )

        return JsonResponse({'status': 'unlocked', 'message': 'Seats have been released.'})
        
    else:
        # --- LOCK ACTION ---
        with transaction.atomic():
            # A. Create the "Blocker" Booking
            booking = Booking.objects.create(
                user=request.user,
                trip=trip,
                booking_ref=f"LOCK-{str(uuid.uuid4())[:8].upper()}",
                status='LOCKED',
                payment_status='UNPAID',
                total_amount=0,
                sales_channel='COUNTER'
            )
            
            # C. Find ALL Seats
            all_seats = LayoutObject.objects.filter(
                deck__ship=trip.ship,
                category__is_bookable=True
            )
            
            locked_count = 0
            long_expiry = timezone.now() + timedelta(days=3650)

            for seat in all_seats:
                # D. Check Availability
                if trip.is_seat_available_admin(seat, start_stop, end_stop):
                    Ticket.objects.create(
                        booking=booking,
                        trip=trip,
                        seat_object=seat,
                        from_stop=start_stop,
                        to_stop=end_stop,
                        passenger_name="For Launch",
                        fare_amount=0,
                        status='LOCKED',
                        lock_expires_at=long_expiry 
                    )
                    locked_count += 1
                    
                    # ADDED: Broadcast the 'locked' event for this seat
                    async_to_sync(channel_layer.group_send)(
                        f"seats_{trip_id}",
                        {
                            "type": "seat_event",
                            "action": "locked", # New action to handle in JS
                            "seat_id": seat.id,
                            "holder_id": request.user.id,
                            # "start_order": start_stop.stop_order, 
                            # "end_order": end_stop.stop_order,
                        }
                    )
                    
            print(f"Locked {locked_count} seats for Trip {trip_id}.")
            
            if locked_count == 0:
                booking.delete()
                print(f"No seats were locked for Trip {trip_id}. Blocker booking deleted.")
                return JsonResponse({'status': 'error', 'message': 'No available seats to lock!'})

            return JsonResponse({
                'status': 'locked', 
                'message': f'{locked_count} seats have been locked successfully.'
            })



from django.core.cache import cache  # Ensure this is imported at the top of your file

@require_POST
@login_required
def toggle_single_seat_lock(request):
    try:
        data = json.loads(request.body)
        trip_id = data.get('trip_id')
        seat_id = data.get('seat_id') # LayoutObject ID
        
        trip = get_object_or_404(Trip, id=trip_id)
        seat_obj = get_object_or_404(LayoutObject, id=seat_id)

        channel_layer = get_channel_layer()

        route_stops = RouteStop.objects.filter(route=trip.route).order_by('stop_order')
        if not route_stops.exists():
            return JsonResponse({'success': False, 'message': 'Route error'})
        
        start_stop = route_stops.first()
        end_stop = route_stops.last()

        # =========================================================
        # SECURITY CHECK 1: Is the seat currently ON HOLD in the DB?
        # =========================================================
        active_hold = SeatHold.objects.filter(
            trip=trip,
            seat_object=seat_obj,
            expires_at__gt=timezone.now() # Only check holds that haven't expired yet
        ).first()
        
        if active_hold:
            return JsonResponse({
                'success': False, 
                'message': 'Action Denied: Seat is currently on hold and being viewed by someone else.'
            })

        # =========================================================
        # SECURITY CHECK 2: Is the seat BOOKED or PENDING in the DB?
        # =========================================================
        # We exclude CANCELLED or REFUNDED tickets just to be safe
        existing_ticket = Ticket.objects.filter(
            trip=trip, 
            seat_object=seat_obj
        ).exclude(status__in=['CANCELLED', 'REFUNDED']).first()

        # ==============================
        # CASE A: UNLOCK (Remove Ticket)
        # ==============================
        if existing_ticket:
            if existing_ticket.status == 'LOCKED':
                existing_ticket.delete()
                
                # Broadcast the 'release' event to all listeners
                async_to_sync(channel_layer.group_send)(
                    f"seats_{trip_id}",
                    {
                        "type": "seat_event",
                        "action": "release",
                        "seat_id": seat_id,
                        "start_order": start_stop.stop_order,
                        "end_order": end_stop.stop_order,
                    }
                )
                
                return JsonResponse({'success': True, 'action': 'unlocked'})
            
            # Specific error messages for other DB states
            elif existing_ticket.status == 'PENDING':
                return JsonResponse({'success': False, 'message': 'Action Denied: Seat is currently pending checkout by a customer!'})
            else:
                return JsonResponse({'success': False, 'message': 'Action Denied: Seat is already sold to a customer!'})

        # ==============================
        # CASE B: LOCK (Add Ticket)
        # ==============================
        else:
            with transaction.atomic():
                # 1. Find or Create the Main "Blocker" Booking
                booking = Booking.objects.filter(trip=trip, status='LOCKED').first()
                
                if not booking:
                    booking = Booking.objects.create(
                        user=request.user,
                        trip=trip,
                        booking_ref=f"LOCK-{str(uuid.uuid4())[:8].upper()}",
                        status='LOCKED',
                        payment_status='UNPAID',
                        total_amount=0,
                        sales_channel='COUNTER'
                    )

                long_expiry = timezone.now() + timedelta(days=3650)

                # 3. Create the TICKET
                Ticket.objects.create(
                    booking=booking,
                    trip=trip,
                    seat_object=seat_obj,
                    from_stop=start_stop,
                    to_stop=end_stop,
                    passenger_name="ADMIN_LOCK",
                    fare_amount=0,
                    status='LOCKED',
                    lock_expires_at=long_expiry
                )
                
                # Broadcast the 'locked' event to all listeners
                async_to_sync(channel_layer.group_send)(
                    f"seats_{trip_id}",
                    {
                        "type": "seat_event",
                        "action": "locked",
                        "seat_id": seat_id,
                        "holder_id": request.user.id,
                    }
                )
                
                return JsonResponse({'success': True, 'action': 'locked'})

    except Exception as e:
        print(f"Single Lock Error: {e}")
        return JsonResponse({'success': False, 'message': str(e)})



from django.db.models import Sum, Count, Q

@login_required
def trip_seat_report(request, trip_id):
    trip = get_object_or_404(Trip, id=trip_id)
    
    # 1. Get Total Capacity (using our fixed is_bookable logic)
    total_capacity = LayoutObject.objects.filter(
        deck__ship=trip.ship, 
        category__is_bookable=True
    ).count()

    # 2. Get Locked Count (Admin locks)
    locked_count = Ticket.objects.filter(trip=trip, status='LOCKED').count()

    # 3. Get Sold Tickets (Grouped by Category)
    # assuming 'BOOKED' is your sold status based on previous logs
    sold_stats = Ticket.objects.filter(trip=trip, status='BOOKED').values(
        'seat_object__category__name'
    ).annotate(
        count=Count('id'),
        total_revenue=Sum('fare_amount')
    )

    # 4. Calculate Totals
    total_sold = sum(item['count'] for item in sold_stats)
    total_revenue = sum(item['total_revenue'] for item in sold_stats) or 0
    unsold_count = total_capacity - (total_sold + locked_count)

    data = {
        'breakdown': list(sold_stats), # Converts QuerySet to list for JSON
        'summary': {
            'total_capacity': total_capacity,
            'total_sold': total_sold,
            'total_locked': locked_count,
            'total_unsold': unsold_count,
            'total_revenue': total_revenue
        }
    }
    return JsonResponse(data)





@login_required
def passenger_list(request):
    """
    Display all unique Customers (User type 1) and their total bookings.
    """
    # Filter for Customers and count their related bookings
    customers = User.objects.filter(user_type=1).annotate(
        total_bookings=Count('bookings')
    ).order_by('-created_at')
    
    context = {
        'users': customers,  # Passed as 'users' to match your template loop
        'page_title': 'Customer List',
    }
    # Pointing to the specific template you mentioned
    return render(request, 'admin_panel/book/passenger_list.html', context)


# --- Apply this same query logic to your other views (issue_list, pending_list, etc.) ---
@login_required
def booking_issue_list(request):
    bookings = Booking.objects.filter(status='CONFIRMED').select_related(
        'user', 'trip__ship', 'trip__route__source', 'trip__route__destination'
    ).prefetch_related(
        'tickets__seat_object', 'tickets__from_stop__location', 'tickets__to_stop__location'
    ).order_by('-created_at')
    
    context = {'bookings': bookings, 'page_title': 'Issued (Confirmed) Tickets'}
    return render(request, 'admin_panel/book/booking_list.html', context)


@login_required
def booking_pending_list(request):
    bookings = Booking.objects.filter(status='PENDING').select_related(
        'user', 'trip__ship', 'trip__route__source', 'trip__route__destination'
    ).prefetch_related(
        'tickets__seat_object', 'tickets__from_stop__location', 'tickets__to_stop__location'
    ).order_by('-created_at')

    context = {'bookings': bookings, 'page_title': 'Pending Payment Tickets'}
    return render(request, 'admin_panel/book/booking_list.html', context)


@login_required
def booking_cancel_list(request):
    # 1. The Query: Find bookings that are Cancelled OR have Cancelled tickets
    bookings = Booking.objects.filter(
        Q(status='CANCELLED') | Q(tickets__status='CANCELLED')
    ).select_related(
        'user', 'trip__ship', 'trip__route__source', 'trip__route__destination'
    ).prefetch_related(
        'tickets__seat_object', 'tickets__from_stop__location', 'tickets__to_stop__location'
    ).distinct().order_by('-created_at')

    # 2. The Context: YOU MUST INCLUDE 'is_cancel_page': True HERE
    context = {
        'bookings': bookings, 
        'page_title': 'Cancelled Tickets History',
        'is_cancel_page': True  # <--- THIS IS THE MISSING KEY!
    }
    
    return render(request, 'admin_panel/book/booking_list.html', context)


@login_required
def booking_expired_list(request):
    """
    Shows a list of all bookings that were auto-cancelled (EXPIRED) due to timeout.
    """
    bookings = Booking.objects.filter(status='EXPIRED').select_related(
        'user', 'trip__ship', 'trip__route__source', 'trip__route__destination'
    ).prefetch_related(
        'tickets__seat_object', 'tickets__from_stop__location', 'tickets__to_stop__location'
    ).order_by('-created_at')

    context = {
        'bookings': bookings, 
        'page_title': 'Expired (Unpaid) Tickets'
    }
    return render(request, 'admin_panel/book/booking_list.html', context)


@xframe_options_sameorigin
def admin_booking_invoice(request, booking_ref):
    booking = get_object_or_404(
        Booking.objects.prefetch_related('tickets__seat_object'),
        booking_ref=booking_ref
    )

    # [NEW:QR_ON_INVOICE]
    # Ensure share_token exists (only once)
    if not booking.share_token:
        booking.share_token = uuid.uuid4().hex
        booking.save(update_fields=["share_token"])

    # Shareable public URL (use token)
    public_url = request.build_absolute_uri(f"/ticket/{booking.booking_ref}/{booking.share_token}/")

    # Generate+save QR only if missing
    ensure_booking_qr(booking, public_url)
    booking.refresh_from_db(fields=["qr_image"])
    # [/NEW:QR_ON_INVOICE]

    due_amount = max(0, (booking.total_amount or 0) - (booking.paid_amount or 0))
    return render(request, 'portal/schedules/booking_success.html', {
        'booking': booking,
        'public_url': public_url,
        'due_amount': due_amount,
    })
    
    
@login_required
def ticket_detail(request, pk):
    booking = get_object_or_404(Booking, pk=pk)
    
    # Get all tickets first
    if hasattr(booking, 'tickets'):
        all_tickets = booking.tickets.all()
    else:
        all_tickets = booking.ticket_set.all()
    
    # Separate active and cancelled tickets
    active_tickets = all_tickets.filter(status__in=['BOOKED', 'CONFIRMED', 'LOCKED'])
    cancelled_tickets = all_tickets.filter(status='CANCELLED')

    context = {
        'booking': booking,
        'tickets': active_tickets,  # Only active tickets for main display
        'cancelled_tickets': cancelled_tickets,  # Cancelled tickets for reference
        'seat_count': active_tickets.count(),
        'cancelled_count': cancelled_tickets.count(),
    }
    return render(request, 'admin_panel/book/ticket_detail.html', context)


@login_required
def booking_visual_map(request, booking_id):
    booking_ref = request.GET.get('booking_ref', '').strip()
    if booking_ref:
        matched_booking = Booking.objects.filter(booking_ref__iexact=booking_ref).first()
        if matched_booking:
            return redirect('booking_visual_map', booking_id=matched_booking.id)
        messages.error(request, f"Booking ref '{booking_ref}' not found.")

    booking = get_object_or_404(Booking, id=booking_id)
    trip = booking.trip
    ship = trip.ship
    
    booked_seats_map = {}
    cancelled_seats = []
    
    # Check if this is an expired booking with snapshot
    if booking.status == 'EXPIRED' and booking.seat_snapshot:
        # Parse the snapshot - it might be stored as comma-separated seat labels
        import re
        snapshot = booking.seat_snapshot
        
        # Try to extract seat labels - format could be "A1, B2, C3" or similar
        seat_labels = re.findall(r'[A-Z0-9]+', snapshot)
        
        # Find matching LayoutObjects by label
        for label in seat_labels:
            try:
                # Find the seat object by label (you might need to adjust this query)
                seat_obj = LayoutObject.objects.filter(
                    deck__ship=ship,
                    label__icontains=label
                ).first()
                
                if seat_obj:
                    booked_seats_map[seat_obj.id] = {
                        'passenger': 'Expired Booking',
                        'phone': booking.user.phone_number if booking.user else "N/A",
                        'amount': 0,
                        'status': 'EXPIRED',
                        'payment_status': 'UNPAID',
                        'label': seat_obj.label
                    }
            except Exception as e:
                print(f"Error finding seat {label}: {e}")
                
    else:
        # Get ALL tickets for this booking
        all_tickets = booking.tickets.all()
        
        # Build map with status information
        for t in all_tickets:
            seat_data = {
                'passenger': t.passenger_name,
                'phone': booking.user.phone_number if booking.user else "N/A",
                'amount': t.fare_amount,
                'status': t.status,
                'payment_status': booking.payment_status if t.status in ['BOOKED', 'CONFIRMED', 'PENDING'] else 'N/A'
            }
            booked_seats_map[t.seat_object.id] = seat_data
            
            # Track cancelled seats separately
            if t.status == 'CANCELLED':
                cancelled_seats.append(t.seat_object.id)

    # Build Grid Data
    decks_data = []
    decks = ship.decks.all().order_by('level_order')

    for deck in decks:
        objects = LayoutObject.objects.filter(deck=deck)
        
        max_col_obj = objects.annotate(
            right_edge=F('col_index') + F('col_span')
        ).aggregate(Max('right_edge'))['right_edge__max']
        
        calculated_cols = max_col_obj if max_col_obj else 20

        decks_data.append({
            'deck_name': deck.name,
            'grid_cols': calculated_cols,
            'objects': objects
        })

    context = {
        'booking': booking,
        'decks_data': decks_data,
        'booked_seats_map': booked_seats_map,
        'cancelled_seats': cancelled_seats,
        'is_expired': booking.status == 'EXPIRED'
    }
    
    return render(request, 'admin_panel/book/booking_visual_map.html', context)



@login_required
def cancel_booking(request, booking_id):
    if not request.user.is_staff: # Security check
        messages.error(request, "Access Denied.")
        return redirect('admin_home')
        
    booking = get_object_or_404(Booking, id=booking_id)

    # Prevent cancelling already cancelled bookings
    if booking.status == 'CANCELLED':
        messages.warning(request, "This booking is already cancelled.")
        return redirect('booking_cancel_list')

    try:
        with transaction.atomic():
            # 1. Update Booking Status
            booking.status = 'CANCELLED'
            booking.save()

            # 2. Get Tickets & Extract Labels for SMS
            tickets = Ticket.objects.filter(booking=booking)
            
            # We grab the seat labels (e.g., ['A1', 'B2']) BEFORE updating them
            # This assumes your Ticket model has a foreign key 'seat_object' with a 'label'
            seat_labels = [t.seat_object.label for t in tickets]
            
            count = tickets.count()
            
            # 3. Update Tickets to Cancelled
            # This releases the seats back to the pool
            tickets.update(status='CANCELLED')

            # 4. SEND SMS (The New Part)
            # The booking status is now 'CANCELLED', so utils.py will send the specific cancel message
            send_booking_sms(booking, seat_labels)

            messages.success(request, f"Booking Cancelled. SMS Sent. {count} seats have been released.")
            
    except Exception as e:
        messages.error(request, f"Error cancelling booking: {e}")
        print(f"Cancel Error: {e}") # Print error to terminal for debugging

    return redirect('booking_cancel_list')
    


from .utils import send_booking_sms, send_partial_cancel_sms  # <--- Make sure to import these

@login_required
def cancel_seats(request):
    if request.method != 'POST':
        return redirect('booking_cancel_list')

    booking_id = request.POST.get('booking_id')
    selected_ticket_ids = request.POST.getlist('ticket_ids')
    
    print(f"=======Received cancellation request for Booking ID: {booking_id} with Ticket IDs: {selected_ticket_ids}======")

    booking = get_object_or_404(Booking, id=booking_id)

    if not selected_ticket_ids:
        print("xxxxxxxxxxxxxxxxxxxNo tickets selected for cancellation.xxxxxxxxxxxxxxxx")
        messages.warning(request, "No seats were selected for cancellation.")
        return redirect('booking_cancel_list')

    try:
        with transaction.atomic():
            # 1. Fetch the specific tickets
            tickets_to_cancel = Ticket.objects.filter(
                id__in=selected_ticket_ids, 
                booking=booking,
                status__in=['BOOKED', 'PENDING'] 
            ).select_related('from_stop', 'to_stop', 'seat_object')
            
            if not tickets_to_cancel.exists():
                print(f"xxxxxxxxxxxxxxxxxxxNo valid tickets found for cancellation.xxxxxxxxxxxxxxxx")
                messages.error(request, "Selected tickets are already cancelled or invalid.")
                return redirect('booking_cancel_list')

            # 2. Extract info for SMS/Logs
            cancelled_labels = [t.seat_object.label for t in tickets_to_cancel]
            # cancel_amount = sum(t.fare_amount for t in tickets_to_cancel) # Available if you need it later

            # =========================================================
            # ADDED: Prepare WebSockets for AFTER transaction commits
            # =========================================================
            # Extract the needed data BEFORE calling .update()
            ws_events = [
                {
                    "seat_id": ticket.seat_object_id,
                    "start_order": ticket.from_stop.stop_order if ticket.from_stop else None,
                    "end_order": ticket.to_stop.stop_order if ticket.to_stop else None,
                }
                for ticket in tickets_to_cancel
            ]
            trip_id = booking.trip_id

            def broadcast_releases():
                channel_layer = get_channel_layer()
                group_name = f"seats_{trip_id}"
                for event in ws_events:
                    async_to_sync(channel_layer.group_send)(
                        group_name,
                        {
                            "type": "seat_event",
                            "action": "release",
                            "seat_id": event["seat_id"],
                            "start_order": event["start_order"],
                            "end_order": event["end_order"],
                        }
                    )

            # Queue the broadcast to fire ONLY when the DB transaction is fully completed
            transaction.on_commit(broadcast_releases)

            # 3. Update Tickets to CANCELLED
            tickets_to_cancel.update(status='CANCELLED')

            # 4. Check if ANY active tickets remain
            # Must check for BOOKED or PENDING here as well!
            remaining_tickets = booking.tickets.filter(status__in=['BOOKED', 'PENDING'])
            
            is_full_cancel = False

            if not remaining_tickets.exists():
                # CASE A: Full Cancellation
                booking.status = 'CANCELLED'
                booking.total_amount = 0
                is_full_cancel = True
            else:
                # CASE B: Partial Cancellation
                new_total = remaining_tickets.aggregate(Sum('fare_amount'))['fare_amount__sum'] or 0
                booking.total_amount = new_total
                is_full_cancel = False

            booking.save()

            # 5. Send SMS (Non-blocking because utils.py uses threading)
            try:
                if is_full_cancel:
                    send_booking_sms(booking, cancelled_labels)
                else:
                    send_partial_cancel_sms(booking, cancelled_labels, booking.total_amount)
            except Exception as e:
                print(f"SMS Error: {e}")

            messages.success(request, f"Cancelled {len(cancelled_labels)} seat(s). New Total: {booking.total_amount}")

    except Exception as e:
        print(f"Cancel Error: {e}")
        messages.error(request, "Error processing cancellation.")

    return redirect('booking_cancel_list')
    


from django.core.serializers.json import DjangoJSONEncoder
import json
from .models import Location, Trip, RouteStop, LayoutObject, Ticket  # Ensure imports are correct

@login_required
def pos_trip_select(request):
    """
    Step 1: Search for a Trip based on Source, Dest, and Date.
    """
    stops = Location.objects.all()

    # 1. Get Search Params from the URL (HTML Form)
    source_id = request.GET.get('source')      # e.g., "1" (Dhaka)
    dest_id = request.GET.get('destination')   # e.g., "5" (Barisal)
    date_str = request.GET.get('date')

    search_results = []

    # 2. Perform Logic (Find trips that actually connect these two stops)
    if source_id and dest_id and date_str:
        # A. Find routes having these stops
        routes_with_source = RouteStop.objects.filter(location_id=source_id).values_list('route_id', 'stop_order')
        routes_with_dest = RouteStop.objects.filter(location_id=dest_id).values_list('route_id', 'stop_order')
        
        source_map = {r_id: order for r_id, order in routes_with_source}
        dest_map = {r_id: order for r_id, order in routes_with_dest}
        
        valid_route_ids = []
        
        # B. Ensure Source comes BEFORE Destination
        for route_id, source_order in source_map.items():
            if route_id in dest_map:
                dest_order = dest_map[route_id]
                if source_order < dest_order:
                    valid_route_ids.append(route_id)
        
        # C. Query Trips
        if valid_route_ids:
            search_results = Trip.objects.filter(
                route_id__in=valid_route_ids,
                departure_datetime__date=date_str,
                is_published=True
            ).select_related('ship', 'route', 'route__source', 'route__destination')

    # 3. Context - sending 'selected_source' to HTML to build the link correctly
    context = {
        'trips': search_results,
        'stops': stops,
        'selected_source': int(source_id) if source_id else '',
        'selected_dest': int(dest_id) if dest_id else '',
        'selected_date': date_str if date_str else '',
    }

    return render(request, 'admin_panel/pos/pos_trip_select.html', context)


@login_required
def pos_booking_interface(request, trip_id):
    """
    Step 2: Show Room Selection.
    CRITICAL: Converts Location IDs (from search) to RouteStop IDs (for pricing).
    """
    trip = get_object_or_404(Trip, pk=trip_id)
    
    # 1. Get Location IDs passed from the previous page
    loc_from_id = request.GET.get('from_loc')
    loc_to_id = request.GET.get('to_loc')

    from_stop = None
    to_stop = None

    # 2. THE TRANSLATOR: Convert "Location ID" -> "RouteStop ID" for THIS trip
    if loc_from_id and loc_to_id:
        from_stop = RouteStop.objects.filter(route=trip.route, location_id=loc_from_id).first()
        to_stop = RouteStop.objects.filter(route=trip.route, location_id=loc_to_id).first()

    # 3. Fallback: If conversion fails (or no params), use full route (A to Z)
    if not from_stop or not to_stop:
        from_stop = RouteStop.objects.filter(route=trip.route).order_by('stop_order').first()
        to_stop = RouteStop.objects.filter(route=trip.route).order_by('stop_order').last()

    # 4. Get Seats & Calculate Price
    seats = LayoutObject.objects.filter(deck__ship=trip.ship).select_related('deck', 'category')
    
    booked_ids = Ticket.objects.filter(
        booking__trip=trip,
        booking__status__in=['CONFIRMED', 'PENDING']
    ).values_list('seat_object_id', flat=True)

    pos_data = []
    for seat in seats:
        try:
            # Pricing now uses the SPECIFIC stops (e.g., Dhaka->Chandpur), not just A->Z
            price = trip.get_price(seat.category, from_stop, to_stop)
        except:
            price = 0 

        pos_data.append({
            'id': seat.id,
            'label': seat.label,
            'deck_name': seat.deck.name,
            'category_name': seat.category.name if seat.category else 'General',
            'category_color': '#2563eb',
            'price': float(price),
            'is_booked': seat.id in booked_ids
        })

    # 5. Send RouteStop IDs (from_stop.id) to the template form
    context = {
        'trip': trip,
        'from_id': from_stop.id,  # This is the RouteStop ID
        'to_id': to_stop.id,      # This is the RouteStop ID
        'pos_data_json': json.dumps(pos_data, cls=DjangoJSONEncoder)
    }
    return render(request, 'admin_panel/pos/pos_booking.html', context)


# ==========================================
# 3. CONFIRM: Saving the Ticket Correctly
# ==========================================
from .services import BookingService
@login_required
def pos_book_confirm(request):
    if request.method != "POST":
        return redirect('admin_home')

    try:
        # 1. Extract Data from POST
        trip_id = request.POST.get('trip_id')
        seat_ids_str = request.POST.get('selected_seats')
        
        # Route info
        from_id = request.POST.get('from_id')
        to_id = request.POST.get('to_id')
        
        # Customer Info
        customer_data = {
            'name': request.POST.get('passenger_name'),
            'phone': request.POST.get('passenger_phone'),
            'email': request.POST.get('passenger_email'), # Added email support
        }

        if not seat_ids_str:
            messages.error(request, "No seats selected.")
            return redirect(request.META.get('HTTP_REFERER'))

        seat_ids_list = [int(s) for s in seat_ids_str.split(',') if s.isdigit()]

        # 2. CALL THE SERVICE (The One Source of Truth)
        booking = BookingService.create_booking(
            admin_user=request.user,
            trip_id=trip_id,
            from_id=from_id,
            to_id=to_id,
            seat_ids_list=seat_ids_list,
            customer_data=customer_data
        )

        # 3. Success!
        messages.success(request, f"POS Booking Confirmed! Ref: {booking.booking_ref} - Total: {booking.total_amount}")
        
        # Redirect back to the seat selection page (maintaining the search params)
        return redirect(f"{reverse('pos_booking_interface', args=[trip_id])}?from={from_id}&to={to_id}")

    except Exception as e:
        # Handles "Seat already booked" and other errors
        messages.error(request, f"Error: {str(e)}")
        return redirect(request.META.get('HTTP_REFERER'))
    
    
    
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger    
# 1. The List View (The Cards)
@login_required
def trip_report_list(request):
    # 1. Get Today's Date
    today = timezone.now().date()

    # 2. Start with ONLY Future & Today's trips
    trips = Trip.objects.filter(
        departure_datetime__date__gte=today
    ).order_by('departure_datetime')

    # 3. Apply Search Filters
    from_loc_id = request.GET.get('from_location')
    to_loc_id = request.GET.get('to_location')
    journey_date = request.GET.get('journey_date')

    if from_loc_id:
        trips = trips.filter(route__source_id=from_loc_id)
    
    if to_loc_id:
        trips = trips.filter(route__destination_id=to_loc_id)
    
    if journey_date:
        trips = trips.filter(departure_datetime__date=journey_date)

    # 4. Count Sold Tickets
    trips = trips.select_related(
        'ship', 'route__source', 'route__destination'
    ).annotate(
        sold_count=Count('tickets', filter=Q(tickets__status='BOOKED'))
    )

    # ========================================================
    # 5. PAGINATION & SHOW ENTITIES LOGIC (ADDED)
    # ========================================================
    per_page = request.GET.get('per_page', '10')  # Default to 10
    
    # Handle 'all' case
    if per_page == 'all':
        # If 'all', show total count (or 1 if empty to avoid error)
        paginator_limit = trips.count() if trips.count() > 0 else 1
    else:
        try:
            paginator_limit = int(per_page)
        except ValueError:
            paginator_limit = 10

    paginator = Paginator(trips, paginator_limit)
    page = request.GET.get('page')

    try:
        trips_page = paginator.page(page)
    except PageNotAnInteger:
        trips_page = paginator.page(1)
    except EmptyPage:
        trips_page = paginator.page(paginator.num_pages)

    locations = Location.objects.all().order_by('name')

    context = {
        'trips': trips_page, # Pass the Page object, not the QuerySet
        'locations': locations,
        'page_title': 'Upcoming Trip Reports',
        'per_page': per_page, # Pass this back so the dropdown remembers selection
    }
    return render(request, 'admin_panel/book/trip_list.html', context)



# 2. The Detail View (The Report)
from django.db.models import Sum, Q
# @login_required
# def trip_passenger_manifest(request, trip_id):
#     trip = get_object_or_404(Trip, id=trip_id)

#     tickets = Ticket.objects.filter(
#         trip=trip
#     ).filter(
#         Q(booking__status__in=['CONFIRMED', 'PENDING']) &
#         ~Q(status='CANCELLED')
#     ).select_related(
#         'booking',
#         'booking__user',
#         'seat_object',
#         'from_stop__location',
#         'to_stop__location'
#     ).order_by('booking_id', 'seat_object__label')

#     # Group by booking
#     bookings_dict = {}

#     for ticket in tickets:
#         booking = ticket.booking
#         bid = booking.id

#         if bid not in bookings_dict:
#             paid_amount = booking.paid_amount or 0

#             bookings_dict[bid] = {
#                 'booking': booking,
#                 'tickets': [],
#                 'total_fare': 0,
#                 'paid_amount': paid_amount,
#                 'due_amount': 0,
#                 'payment_label': '',
#             }

#         bookings_dict[bid]['tickets'].append(ticket)
#         bookings_dict[bid]['total_fare'] += (ticket.fare_amount or 0)

#     total_paid = 0
#     total_due = 0

#     # Finalize booking totals
#     for bid, data in bookings_dict.items():
#         booking = data['booking']
#         total_fare = data['total_fare']
#         paid_amount = data['paid_amount']

#         # Confirmed booking = fully paid (for report display)
#         if booking.status == 'CONFIRMED':
#             paid_amount = total_fare
#             due_amount = 0
#         else:
#             # Pending booking can have partial payment
#             if paid_amount < 0:
#                 paid_amount = 0
#             if paid_amount > total_fare:
#                 paid_amount = total_fare

#             due_amount = total_fare - paid_amount

#         data['paid_amount'] = paid_amount
#         data['due_amount'] = due_amount

#         if booking.status == 'CONFIRMED':
#             data['payment_label'] = 'Paid'
#         else:
#             if paid_amount == 0:
#                 data['payment_label'] = 'Unpaid'
#             elif paid_amount < total_fare:
#                 data['payment_label'] = f'Partial ({paid_amount:.0f}/{total_fare:.0f})'
#             else:
#                 data['payment_label'] = 'Fully Paid'

#         total_paid += paid_amount
#         total_due += due_amount

#     # Convert to list (preserve order)
#     booking_groups = list(bookings_dict.values())

#     context = {
#         'trip': trip,
#         'booking_groups': booking_groups,
#         'total_tickets': tickets.count(),
#         'total_paid': total_paid,
#         'total_due': total_due,
#         'print_date': timezone.now(),
#     }
#     return render(request, 'admin_panel/book/passenger_manifest.html', context)


@login_required
@permission_required('admin_panel.can_view_trip_sheet', raise_exception=True)
def trip_passenger_manifest(request, trip_id):
    trip = get_object_or_404(Trip, id=trip_id)

    # 1. Fetch tickets
    tickets = Ticket.objects.filter(
        trip=trip
    ).filter(
        Q(booking__status__in=['CONFIRMED', 'PENDING']) &
        ~Q(status='CANCELLED')
    ).select_related(
        'booking',
        'booking__user',
        'seat_object',
        'from_stop__location',
        'to_stop__location'
    )

    # 2. Pre-calculate booking totals so they apply to all tickets within the same booking
    booking_totals = {}
    for ticket in tickets:
        bid = ticket.booking.id
        if bid not in booking_totals:
            booking_totals[bid] = {
                'total_fare': 0,
                'paid_amount': ticket.booking.paid_amount or 0,
                'status': ticket.booking.status,
                # NEW: keep payment status so confirmed-but-unpaid can show correctly in manifest
                'payment_status': ticket.booking.payment_status,
                'payment_label': '',
            }
        booking_totals[bid]['total_fare'] += (ticket.fare_amount or 0)

    total_paid = 0
    total_due = 0

    # 3. Finalize booking totals (Paid / Due)
    for bid, data in booking_totals.items():
        total_fare = data['total_fare']
        paid_amount = data['paid_amount']

        # Confirmed + paid stays fully paid for report display.
        # Confirmed + unpaid keeps the original paid/due values.
        if data['status'] == 'CONFIRMED' and data['payment_status'] == 'PAID':
            paid_amount = total_fare
            due_amount = 0
        else:
            if paid_amount < 0:
                paid_amount = 0
            if paid_amount > total_fare:
                paid_amount = total_fare
            due_amount = total_fare - paid_amount

        data['paid_amount'] = paid_amount
        data['due_amount'] = due_amount

        if data['status'] == 'CONFIRMED' and data['payment_status'] == 'PAID':
            data['payment_label'] = 'Paid'
        elif data['status'] == 'CONFIRMED':
            if due_amount > 0:
                data['payment_label'] = f'Partial ({paid_amount:.0f}/{total_fare:.0f})'
            else:
                data['payment_label'] = 'Paid'
        else:
            if paid_amount == 0:
                data['payment_label'] = 'Unpaid'
            elif paid_amount < total_fare:
                data['payment_label'] = f'Partial ({paid_amount:.0f}/{total_fare:.0f})'
            else:
                data['payment_label'] = 'Fully Paid'

        # Add to Grand Totals (Added ONCE per booking)
        total_paid += paid_amount
        total_due += due_amount

    # 4. Attach booking data to tickets and extract cabin number for sorting
    ticket_list = []
    for ticket in tickets:
        bid = ticket.booking.id
        label = ticket.seat_object.seat_identifier if ticket.seat_object.seat_identifier else ticket.seat_object.label
        
        # Extract numeric part from cabin name (e.g. "double cabin -118" -> 118)
        match = re.search(r'\d+', label) if label else None
        sort_key = int(match.group()) if match else 999999 # Default to high number if no digits found
        
        # Attach dynamic attributes to the ticket object
        ticket.booking_data = booking_totals[bid]
        ticket.sort_key = sort_key
        ticket.display_label = label
        ticket_list.append(ticket)

    # 5. Sort the flat list by the extracted cabin number
    ticket_list.sort(key=lambda x: x.sort_key)

    context = {
        'trip': trip,
        'tickets': ticket_list, 
        'total_tickets': len(ticket_list),
        'total_paid': total_paid,
        'total_due': total_due,
        'print_date': timezone.now(),
    }
    return render(request, 'admin_panel/book/passenger_manifest.html', context)



# group by booking format
# @login_required
# def export_manifest_xls(request, trip_id):
#     trip = get_object_or_404(Trip, id=trip_id)
    
#     # 1. Fetch Data (Exact same logic as main view)
#     tickets = Ticket.objects.filter(
#         trip=trip
#     ).filter(
#         Q(booking__status__in=['CONFIRMED', 'PENDING']) &
#         ~Q(status='CANCELLED')
#     ).select_related(
#         'booking', 'booking__user', 'seat_object', 'from_stop__location', 'to_stop__location'
#     ).order_by('booking_id', 'seat_object__label')

#     # Group by booking
#     bookings_dict = {}
#     for ticket in tickets:
#         booking = ticket.booking
#         bid = booking.id

#         if bid not in bookings_dict:
#             paid_amount = booking.paid_amount or 0
#             bookings_dict[bid] = {
#                 'booking': booking,
#                 'tickets': [],
#                 'total_fare': 0,
#                 'paid_amount': paid_amount,
#                 'due_amount': 0,
#                 'payment_label': '',
#             }
#         bookings_dict[bid]['tickets'].append(ticket)
#         bookings_dict[bid]['total_fare'] += (ticket.fare_amount or 0)

#     total_paid = 0
#     total_due = 0

#     # Finalize booking totals
#     for bid, data in bookings_dict.items():
#         booking = data['booking']
#         total_fare = data['total_fare']
#         paid_amount = data['paid_amount']

#         if booking.status == 'CONFIRMED':
#             paid_amount = total_fare
#             due_amount = 0
#         else:
#             if paid_amount < 0: paid_amount = 0
#             if paid_amount > total_fare: paid_amount = total_fare
#             due_amount = total_fare - paid_amount

#         data['paid_amount'] = paid_amount
#         data['due_amount'] = due_amount
        
#         if booking.status == 'CONFIRMED':
#             data['payment_label'] = 'Paid'
#         else:
#             if paid_amount == 0: data['payment_label'] = 'Unpaid'
#             elif paid_amount < total_fare: data['payment_label'] = f'Partial ({paid_amount:.0f}/{total_fare:.0f})'
#             else: data['payment_label'] = 'Fully Paid'

#         total_paid += paid_amount
#         total_due += due_amount

#     booking_groups = list(bookings_dict.values())

#     # 2. Create Workbook & Sheet
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = f"Trip-{trip.ship.name}"

#     # --- STYLES ---
#     bold_font = Font(bold=True)
#     center_align = Alignment(horizontal='center', vertical='center')
#     left_align = Alignment(horizontal='left', vertical='center')
#     thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
#                          top=Side(style='thin'), bottom=Side(style='thin'))
#     fill_group = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")

#     # 3. Write TRIP INFO
#     ws.append(['Launch', 'Route', 'Starting Place', 'Journey Date', 'Tickets', 'Seats'])
#     ws.append([
#         trip.ship.name, trip.route.name, trip.route.source.name,
#         trip.departure_datetime.strftime("%Y-%m-%d %H:%M"),
#         tickets.count(), tickets.count()
#     ])
#     ws.append([]) # Empty row

#     # 4. Write PASSENGER TABLE Headers
#     headers_main = [
#         'SL', 'Ticket No', 'Cabin', 'Pass. Name', 'Phone No', 
#         'Paid', 'Discount', 'Due', 'From', 'To', 
#         'Status', 'Booked By', 'Issued By', 'Remarks'
#     ]
#     ws.append(headers_main)
#     for cell in ws[ws.max_row]:
#         cell.font = bold_font
#         cell.alignment = center_align
#         cell.border = thin_border

#     # 5. Write Passenger Rows Grouped by Booking
#     for group in booking_groups:
#         booking = group['booking']
        
#         # Booking Header Row
#         timer_text = " (Timer Stopped)" if booking.status == 'PENDING' and getattr(booking, 'time_stopped', False) else ""
#         header_text = f"Booking #{booking.id} (Ref: {booking.booking_ref}) - Total Fare: {group['total_fare']:.0f} - Paid: {group['paid_amount']:.0f} - Due: {group['due_amount']:.0f} - {booking.status}{timer_text}"
        
#         ws.append([header_text] + [''] * 13)
#         ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=14)
#         header_cell = ws.cell(row=ws.max_row, column=1)
#         header_cell.font = bold_font
#         header_cell.fill = fill_group
#         header_cell.alignment = left_align
#         for cell in ws[ws.max_row]: cell.border = thin_border

#         # Tickets for this booking
#         for index, ticket in enumerate(group['tickets'], start=1):
#             phone = getattr(ticket.passenger, 'phone', None) if hasattr(ticket, 'passenger') else None
#             if not phone:
#                 phone = ticket.booking.user.phone_number or "0"
                
#             row = [
#                 index,
#                 ticket.id,
#                 ticket.seat_object.seat_identifier or ticket.seat_object.label,
#                 ticket.passenger_name or "-",
#                 phone,
#                 group['paid_amount'] if index == 1 else "",  # Show paid only on first row
#                 0, # Discount
#                 group['due_amount'] if index == 1 else "",   # Show due only on first row
#                 ticket.from_stop.location.name,
#                 ticket.to_stop.location.name,
#                 "Confirmed" if booking.status == 'CONFIRMED' else group['payment_label'],
#                 ticket.booking.user.first_name or "Admin",
#                 request.user.first_name or "Admin",
#                 "Timer Stopped" if booking.status == 'PENDING' and getattr(booking, 'time_stopped', False) else "N/A"
#             ]
#             ws.append(row)
#             for cell in ws[ws.max_row]:
#                 cell.border = thin_border
#                 cell.alignment = center_align

#     # 6. Write FOOTER
#     footer_row = ['', '', '', '', '', f"Total: {total_paid:.0f}", '', f"Due: {total_due:.0f}", '', '', '', '', '', '']
#     ws.append(footer_row)
#     for cell in ws[ws.max_row]:
#         cell.font = bold_font
#         cell.border = thin_border
#         cell.alignment = center_align

#     # 7. Response
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     filename = f"Manifest_{trip.ship.name}_{trip.departure_datetime.date()}.xlsx"
#     response['Content-Disposition'] = f'attachment; filename="{filename}"'
#     wb.save(response)
#     return response



@login_required
def export_manifest_xls(request, trip_id):
    trip = get_object_or_404(Trip, id=trip_id)
    
    # 1. Fetch Data
    tickets = Ticket.objects.filter(
        trip=trip
    ).filter(
        Q(booking__status__in=['CONFIRMED', 'PENDING']) &
        ~Q(status='CANCELLED')
    ).select_related(
        'booking', 'booking__user', 'seat_object', 'from_stop__location', 'to_stop__location'
    )

    # 2. Pre-calculate booking totals
    booking_totals = {}
    for ticket in tickets:
        bid = ticket.booking.id
        if bid not in booking_totals:
            booking_totals[bid] = {
                'total_fare': 0,
                'paid_amount': ticket.booking.paid_amount or 0,
                'status': ticket.booking.status,
                # UPDATED: keep payment status for confirmed-but-unpaid PDF display
                'payment_status': ticket.booking.payment_status,
                'payment_label': '',
            }
        booking_totals[bid]['total_fare'] += (ticket.fare_amount or 0)

    total_paid = 0
    total_due = 0

    # 3. Finalize booking totals
    for bid, data in booking_totals.items():
        total_fare = data['total_fare']
        paid_amount = data['paid_amount']

        # ✅ FIXED: only force paid=full when payment_status is actually PAID
        if data['status'] == 'CONFIRMED' and data['payment_status'] == 'PAID':
            paid_amount = total_fare
            due_amount = 0
        else:
            if paid_amount < 0: paid_amount = 0
            if paid_amount > total_fare: paid_amount = total_fare
            due_amount = total_fare - paid_amount

        data['paid_amount'] = paid_amount
        data['due_amount'] = due_amount

        # ✅ FIXED: match web view payment_label logic exactly
        if data['status'] == 'CONFIRMED' and data['payment_status'] == 'PAID':
            data['payment_label'] = 'Paid'
        elif data['status'] == 'CONFIRMED':
            if due_amount > 0:
                data['payment_label'] = f'Partial ({paid_amount:.0f}/{total_fare:.0f})'
            else:
                data['payment_label'] = 'Paid'
        else:
            if paid_amount == 0: data['payment_label'] = 'Unpaid'
            elif paid_amount < total_fare: data['payment_label'] = f'Partial ({paid_amount:.0f}/{total_fare:.0f})'
            else: data['payment_label'] = 'Fully Paid'

        total_paid += paid_amount
        total_due += due_amount

    # 4. Attach booking data to tickets and extract cabin number for sorting
    ticket_list = []
    for ticket in tickets:
        bid = ticket.booking.id
        label = ticket.seat_object.seat_identifier if ticket.seat_object.seat_identifier else ticket.seat_object.label
        
        match = re.search(r'\d+', label) if label else None
        sort_key = int(match.group()) if match else 999999
        
        ticket.booking_data = booking_totals[bid]
        ticket.sort_key = sort_key
        ticket.display_label = label
        ticket_list.append(ticket)

    # Sort the flat list
    ticket_list.sort(key=lambda x: x.sort_key)

    # 5. Create Workbook & Sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Trip-{trip.ship.name}"

    # --- STYLES ---
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # 6. Write TRIP INFO
    ws.append(['Launch', 'Route', 'Starting Place', 'Journey Date', 'Tickets', 'Seats'])
    ws.append([
        trip.ship.name, trip.route.name, trip.route.source.name,
        trip.departure_datetime.strftime("%Y-%m-%d %H:%M"),
        len(ticket_list), len(ticket_list)
    ])
    ws.append([]) # Empty row

    # 7. Write PASSENGER TABLE Headers
    headers_main = [
        'SL', 'Ticket No', 'Booking Ref', 'Cabin', 'Pass. Name', 'Phone No', 
        'Paid', 'Discount', 'Due', 'From', 'To', 
        'Status', 'Booked By', 'Issued By', 'Remarks'
    ]
    ws.append(headers_main)
    for cell in ws[ws.max_row]:
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border

    # 8. Write Sorted Passenger Rows
    for index, ticket in enumerate(ticket_list, start=1):
        booking = ticket.booking
        b_data = ticket.booking_data
        
        # Phone logic
        phone = getattr(ticket.passenger, 'phone', None) if hasattr(ticket, 'passenger') else None
        if not phone:
            phone = booking.user.phone_number or "0"

        # Remarks logic matching the HTML exactly
        if booking.status == 'PENDING' and getattr(booking, 'time_stopped', False):
            remarks = "Timer Stopped"
        elif booking.status == 'PENDING' and b_data['paid_amount'] > 0 and b_data['due_amount'] > 0:
            remarks = "Partial Payment"
        elif booking.status == 'PENDING' and b_data['paid_amount'] == 0:
            remarks = "Unpaid"
        elif booking.status == 'PENDING' and b_data['due_amount'] == 0:
            remarks = "Paid*"
        else:
            remarks = "N/A"

        # Status logic
        if booking.status == 'CONFIRMED':
            status_text = f"Confirmed ({b_data['payment_label']})"
        else:
            status_text = f"Pending ({b_data['payment_label']})"

        # Booked By logic (safely handle None)
        booked_by_name = "-"
        if booking.booked_by:
            booked_by_name = booking.booked_by.get_display_name() or booking.booked_by.phone_number or "-"

        # Issued By logic (safely handle None)
        issued_by_name = "-"
        if booking.issued_by:
            issued_by_name = booking.issued_by.get_display_name() or booking.issued_by.phone_number or "-"
            
        row = [
            index,
            ticket.id,
            booking.booking_ref,
            ticket.display_label,
            ticket.passenger_name or "-",
            phone,
            round(b_data['paid_amount']),  # Shows total paid for booking on every row
            0,                             # Discount
            round(b_data['due_amount']),   # Shows total due for booking on every row
            ticket.from_stop.location.name,
            ticket.to_stop.location.name,
            status_text,
            booked_by_name,
            issued_by_name,
            remarks
        ]
        
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.border = thin_border
            cell.alignment = center_align

    # 9. Write FOOTER
    footer_row = [
        '', '', '', '', '', 'Grand Total:', 
        round(total_paid), '', round(total_due), 
        '', '', '', '', '', ''
    ]
    ws.append(footer_row)
    for cell in ws[ws.max_row]:
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = center_align

    # 10. Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Manifest_{trip.ship.name}_{trip.departure_datetime.date()}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response





from django.template.loader import get_template
from xhtml2pdf import pisa  # Make sure you installed this: pip install xhtml2pdf
import io

# groupo by booking format
# --- NEW PDF VIEW (Does not touch Excel view) ---
# @login_required
# def download_manifest_pdf(request, trip_id):
#     trip = get_object_or_404(Trip, id=trip_id)
    
#     # Fetch exactly like the screen view
#     tickets = Ticket.objects.filter(
#         trip=trip
#     ).filter(
#         Q(booking__status__in=['CONFIRMED', 'PENDING']) &
#         ~Q(status='CANCELLED')
#     ).select_related(
#         'booking', 'booking__user', 'seat_object', 
#         'from_stop__location', 'to_stop__location'
#     ).order_by('booking_id', 'seat_object__label')
    
#     # 1. Group tickets by booking
#     bookings_dict = {}
#     for ticket in tickets:
#         bid = ticket.booking_id
#         if bid not in bookings_dict:
#             bookings_dict[bid] = {'booking': ticket.booking, 'tickets': [], 'total_fare': 0}
#         bookings_dict[bid]['tickets'].append(ticket)
#         bookings_dict[bid]['total_fare'] += (ticket.fare_amount or 0)

#     # 2. Build the "Flat List" for perfect PDF rendering
#     pdf_rows = []
#     sl_counter = 1
#     total_paid = 0
#     total_due = 0

#     for bid, data in bookings_dict.items():
#         booking = data['booking']
#         total_fare = data['total_fare']
#         paid_amount = booking.paid_amount or 0

#         # Screen View Financial Logic
#         if booking.status == 'CONFIRMED':
#             paid_amount = total_fare
#             due_amount = 0
#         else:
#             if paid_amount < 0: paid_amount = 0
#             if paid_amount > total_fare: paid_amount = total_fare
#             due_amount = total_fare - paid_amount

#         # Determine label
#         payment_label = 'Paid' if booking.status == 'CONFIRMED' else (
#             'Unpaid' if paid_amount == 0 else 
#             f'Partial ({paid_amount:.0f}/{total_fare:.0f})' if paid_amount < total_fare else 'Fully Paid'
#         )

#         total_paid += paid_amount
#         total_due += due_amount

#         # Add Booking Header Row
#         pdf_rows.append({
#             'is_header': True,
#             'booking_id': booking.id,
#             'booking_ref': getattr(booking, 'booking_ref', 'N/A'),
#             'status': booking.status,
#             'time_stopped': getattr(booking, 'time_stopped', False),
#             'total_fare': total_fare,
#             'paid_amount': paid_amount,
#             'due_amount': due_amount,
#         })

#         # Add Ticket Rows (with visual merge flags)
#         num_tickets = len(data['tickets'])
#         for index, ticket in enumerate(data['tickets']):
#             booked_by = booking.user.first_name if booking.user else "Admin"
            
#             # Safe parsing
#             pass_name = getattr(ticket, 'passenger_name', '') or "-"
#             phone = "0"
#             if hasattr(ticket, 'passenger') and ticket.passenger and ticket.passenger.phone:
#                 phone = ticket.passenger.phone
#             elif booking.user and booking.user.phone_number:
#                 phone = booking.user.phone_number

#             cabin = ticket.seat_object.seat_identifier if getattr(ticket.seat_object, 'seat_identifier', None) else ticket.seat_object.label

#             pdf_rows.append({
#                 'is_header': False,
#                 'sl': sl_counter,
#                 'ticket_id': ticket.id,
#                 'cabin': cabin,
#                 'pass_name': pass_name,
#                 'phone': phone,
#                 'from_station': ticket.from_stop.location.name if ticket.from_stop else "",
#                 'to_station': ticket.to_stop.location.name if ticket.to_stop else "",
                
#                 # Group Data
#                 'group_paid': paid_amount,
#                 'group_due': due_amount,
#                 'payment_label': payment_label,
#                 'status': booking.status,
#                 'booked_by': booked_by,
#                 'time_stopped': getattr(booking, 'time_stopped', False),

#                 # CSS Flags for "Visual Merging"
#                 'is_single': (num_tickets == 1),
#                 'is_first': (index == 0),
#                 'is_last': (index == num_tickets - 1),
#             })
#             sl_counter += 1

#     context = {
#         'trip': trip,
#         'pdf_rows': pdf_rows,
#         'total_tickets': tickets.count(),
#         'total_paid': total_paid,
#         'total_due': total_due,
#         'current_user': request.user,
#     }

#     # IMPORTANT: Ensure this path matches where your HTML file actually lives
#     template = get_template('admin_panel/book/manifest_pdf.html') 
#     html = template.render(context)
    
#     response = HttpResponse(content_type='application/pdf')
#     filename = f"TripSheet_{trip.ship.name}_{trip.departure_datetime.date()}.pdf"
#     response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
#     pisa_status = pisa.CreatePDF(html, dest=response)
#     if pisa_status.err:
#         return HttpResponse(f'PDF generation error: {pisa_status.err}')
    
#     return response


@login_required
def download_manifest_pdf(request, trip_id):
    trip = get_object_or_404(Trip, id=trip_id)

    tickets = Ticket.objects.filter(
        trip=trip
    ).filter(
        Q(booking__status__in=['CONFIRMED', 'PENDING']) &
        ~Q(status='CANCELLED')
    ).select_related(
        'booking', 'booking__user', 'seat_object',
        'from_stop__location', 'to_stop__location'
    )

    booking_totals = {}
    for ticket in tickets:
        bid = ticket.booking.id
        if bid not in booking_totals:
            booking_totals[bid] = {
                'total_fare': 0,
                'paid_amount': ticket.booking.paid_amount or 0,
                'status': ticket.booking.status,
                # ✅ ADDED: needed to correctly handle confirmed-but-unpaid bookings
                'payment_status': ticket.booking.payment_status,
                'payment_label': '',
            }
        booking_totals[bid]['total_fare'] += (ticket.fare_amount or 0)

    total_paid = 0
    total_due = 0

    for bid, data in booking_totals.items():
        total_fare = data['total_fare']
        paid_amount = data['paid_amount']

        # ✅ FIXED: match web view logic exactly — only force paid=full when payment_status is PAID
        if data['status'] == 'CONFIRMED' and data['payment_status'] == 'PAID':
            paid_amount = total_fare
            due_amount = 0
        else:
            if paid_amount < 0:
                paid_amount = 0
            if paid_amount > total_fare:
                paid_amount = total_fare
            due_amount = total_fare - paid_amount

        data['paid_amount'] = paid_amount
        data['due_amount'] = due_amount

        # ✅ FIXED: match web view payment_label logic exactly
        if data['status'] == 'CONFIRMED' and data['payment_status'] == 'PAID':
            data['payment_label'] = 'Paid'
        elif data['status'] == 'CONFIRMED':
            if due_amount > 0:
                data['payment_label'] = f'Partial ({paid_amount:.0f}/{total_fare:.0f})'
            else:
                data['payment_label'] = 'Paid'
        else:
            if paid_amount == 0:
                data['payment_label'] = 'Unpaid'
            elif paid_amount < total_fare:
                data['payment_label'] = f'Partial ({paid_amount:.0f}/{total_fare:.0f})'
            else:
                data['payment_label'] = 'Fully Paid'

        total_paid += paid_amount
        total_due += due_amount

    ticket_list = []
    for ticket in tickets:
        bid = ticket.booking.id
        label = ticket.seat_object.seat_identifier if getattr(ticket.seat_object, 'seat_identifier', None) else ticket.seat_object.label

        match = re.search(r'\d+', label) if label else None
        sort_key = int(match.group()) if match else 999999

        ticket.booking_data = booking_totals[bid]
        ticket.sort_key = sort_key
        ticket.display_label = label
        ticket_list.append(ticket)

    ticket_list.sort(key=lambda x: x.sort_key)

    context = {
        'trip': trip,
        'tickets': ticket_list,
        'total_tickets': len(ticket_list),
        'total_paid': total_paid,
        'total_due': total_due,
        'request': request,
    }

    template = get_template('admin_panel/book/manifest_pdf.html')
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    filename = f"TripSheet_{trip.ship.name}_{trip.departure_datetime.date()}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse(f'PDF generation error: {pisa_status.err}')

    return response


def _get_counter_sales_queryset(counter, request):
    qs = Booking.objects.filter(counter=counter).select_related(
        'trip',
        'trip__route',
        'user',
        'issued_by'
    ).prefetch_related(
        'tickets__seat_object'
    ).order_by('-created_at')

    # Filters
    date_from = (request.GET.get('date_from') or '').strip()
    date_to = (request.GET.get('date_to') or '').strip()
    status = (request.GET.get('status') or '').strip()
    issued_by_id = (request.GET.get('issued_by') or '').strip()

    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)

    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)

    if status:
        qs = qs.filter(status=status)

    if issued_by_id:
        qs = qs.filter(issued_by_id=issued_by_id)

    return qs


@login_required
def counter_sales_report(request, counter_id):
    counter = get_object_or_404(Counter, id=counter_id)

    bookings_qs = _get_counter_sales_queryset(counter, request)

    # Summary totals
    totals = bookings_qs.aggregate(
        total_bookings=Count('id'),
        total_sales=Coalesce(Sum('total_amount'), Value(0, output_field=DecimalField(max_digits=12, decimal_places=2))),
        total_paid=Coalesce(Sum('paid_amount'), Value(0, output_field=DecimalField(max_digits=12, decimal_places=2))),
    )

    total_sales = totals['total_sales'] or Decimal('0')
    total_paid = totals['total_paid'] or Decimal('0')
    total_due = total_sales - total_paid
    if total_due < 0:
        total_due = Decimal('0')

    # Build display rows (seat labels, due, passenger name, etc.)
    report_rows = []
    for booking in bookings_qs:
        seat_labels = list(
            booking.tickets.all().order_by('seat_object__label').values_list('seat_object__label', flat=True)
        )
        seat_labels_str = ", ".join([str(s) for s in seat_labels]) if seat_labels else "-"

        due_amount = (booking.total_amount or Decimal('0')) - (booking.paid_amount or Decimal('0'))
        if due_amount < 0:
            due_amount = Decimal('0')

        # Prefer passenger object name, fallback to booking.user display
        first_passenger = booking.passengers.first() if hasattr(booking, 'passengers') else None
        if first_passenger and first_passenger.name:
            passenger_name = first_passenger.name
            passenger_phone = first_passenger.phone or (booking.user.phone_number if booking.user else '')
        else:
            passenger_name = booking.user.get_display_name() if booking.user else 'Unknown'
            passenger_phone = booking.user.phone_number if booking.user else ''

        report_rows.append({
            'booking': booking,
            'passenger_name': passenger_name,
            'passenger_phone': passenger_phone,
            'seat_labels': seat_labels_str,
            'due_amount': due_amount,
        })

    # Issued-by filter dropdown options (only users used by this counter)
    issued_by_users = User.objects.filter(
        issued_bookings__counter=counter
    ).distinct().order_by('first_name', 'phone_number')

    context = {
        'counter': counter,
        'report_rows': report_rows,
        'total_bookings': totals['total_bookings'] or 0,
        'total_sales': total_sales,
        'total_paid': total_paid,
        'total_due': total_due,
        'issued_by_users': issued_by_users,
        'filters': {
            'date_from': request.GET.get('date_from', ''),
            'date_to': request.GET.get('date_to', ''),
            'status': request.GET.get('status', ''),
            'issued_by': request.GET.get('issued_by', ''),
        }
    }
    return render(request, 'admin_panel/routes/counter_sales_report.html', context)

@login_required
def counter_sales_report_csv(request, counter_id):
    counter = get_object_or_404(Counter, id=counter_id)
    bookings_qs = _get_counter_sales_queryset(counter, request)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="counter_{counter.id}_sales_report.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'Booking Ref',
        'Created At',
        'Passenger Name',
        'Passenger Phone',
        'Trip',
        'Route',
        'Seats',
        'Status',
        'Payment Status',
        'Total Amount',
        'Paid Amount',
        'Due Amount',
        'Issued By'
    ])

    for booking in bookings_qs:
        seat_labels = list(
            booking.tickets.all().order_by('seat_object__label').values_list('seat_object__label', flat=True)
        )
        seat_labels_str = ", ".join([str(s) for s in seat_labels]) if seat_labels else "-"

        due_amount = (booking.total_amount or Decimal('0')) - (booking.paid_amount or Decimal('0'))
        if due_amount < 0:
            due_amount = Decimal('0')

        first_passenger = booking.passengers.first() if hasattr(booking, 'passengers') else None
        if first_passenger and first_passenger.name:
            passenger_name = first_passenger.name
            passenger_phone = first_passenger.phone or (booking.user.phone_number if booking.user else '')
        else:
            passenger_name = booking.user.get_display_name() if booking.user else 'Unknown'
            passenger_phone = booking.user.phone_number if booking.user else ''

        trip_name = str(booking.trip.ship.name) if booking.trip and booking.trip.ship else ''
        route_name = str(booking.trip.route.name) if booking.trip and booking.trip.route else ''

        issued_by_name = booking.issued_by.get_display_name() if getattr(booking, 'issued_by', None) else ''

        writer.writerow([
            booking.booking_ref,
            booking.created_at.strftime('%Y-%m-%d %H:%M:%S') if booking.created_at else '',
            passenger_name,
            passenger_phone,
            trip_name,
            route_name,
            seat_labels_str,
            booking.status,
            booking.payment_status,
            booking.total_amount,
            booking.paid_amount,
            due_amount,
            issued_by_name
        ])

    return response


@login_required
def vessel_showcase_list(request):
    # Handle the AJAX Delete Request
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            action = data.get("action")
            
            if action == "delete":
                showcase_id = data.get("id")
                showcase = VesselShowcase.objects.get(id=showcase_id)
                showcase.delete()
                
                return JsonResponse({"status": "success", "message": "Vessel showcase deleted successfully."})
                
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)})

    # Handle the standard GET Request
    showcases = VesselShowcase.objects.select_related('ship').all().order_by('-created_at')
    
    context = {
        'showcases': showcases
    }
    return render(request, 'admin_panel/content/vessels/vessel_showcase_list.html', context)

# --- Placeholders for next steps ---
@login_required
def add_vessel_showcase(request):
    if request.method == "POST":
        try:
            name = request.POST.get('name')
            ship_id = request.POST.get('ship_id')
            tagline = request.POST.get('tagline')
            short_description = request.POST.get('short_description')
            full_description = request.POST.get('full_description')
            video_tour_url = request.POST.get('video_tour_url')
            display_capacity = request.POST.get('display_capacity')
            hero_image = request.FILES.get('hero_image')
            banner_image = request.FILES.get('banner_image')

            # Basic Validation (Banner removed from required check)
            if not name or not hero_image:
                return JsonResponse({"status": "error", "message": "Name and Hero Image are required."})

            # --- Backend Image Validation ---
            # 1. File Size Validation (<= 500KB)
            if hero_image.size > 500 * 1024:
                return JsonResponse({"status": "error", "message": "Image size must not exceed 500KB."})

            # 2. Dimensions & Format Validation
            try:
                img = Image.open(hero_image)
                img.verify()  # Verifies it is a valid image file (handles jpg, png, webp, etc.)

                # Re-open the file for size/dimension reading (verify() leaves it in a closed state)
                hero_image.seek(0)
                img = Image.open(hero_image)
                width, height = img.size

                if width != height:
                    return JsonResponse({"status": "error", "message": "Image must be a square (1:1 aspect ratio)."})

                if width > 800 or height > 800:
                    return JsonResponse({"status": "error", "message": "Image dimensions must not exceed 800x800 pixels."})

            except Exception:
                return JsonResponse({"status": "error", "message": "Invalid image file uploaded."})
            # --------------------------------

            # --- Backend Banner Image Validation ---
            # Wrapped in an if statement so it only runs if a banner was actually uploaded
            if banner_image:
                if banner_image.size > 2 * 1024 * 1024:
                    return JsonResponse({"status": "error", "message": "Banner image size must not exceed 2MB."})

                try:
                    b_img = Image.open(banner_image)
                    b_img.verify()

                    banner_image.seek(0)
                    b_img = Image.open(banner_image)
                    b_width, b_height = b_img.size

                    # Validate 4:1 aspect ratio with a 0.05 tolerance margin
                    if abs((b_width / b_height) - 4.0) > 0.05:
                        return JsonResponse({"status": "error", "message": "Banner image must have exactly a 4:1 aspect ratio (e.g., 1920x480)."})

                except Exception:
                    return JsonResponse({"status": "error", "message": "Invalid banner image file uploaded."})
            # --------------------------------

            # Check if Ship is selected and valid
            ship_instance = None
            if ship_id:
                if VesselShowcase.objects.filter(ship_id=ship_id).exists():
                    return JsonResponse({"status": "error", "message": "This ship already has a showcase."})
                ship_instance = Ship.objects.get(id=ship_id)

            # Create the showcase
            showcase = VesselShowcase.objects.create(
                name=name,
                ship=ship_instance,
                tagline=tagline,
                short_description=short_description,
                full_description=full_description,
                video_tour_url=video_tour_url,
                display_capacity=display_capacity,
                hero_image=hero_image,
                banner_image=banner_image
            )

            return JsonResponse({"status": "success", "message": "Vessel showcase created successfully!"})

        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)})

    # GET Request: Fetch ships that don't already have a showcase linked
    available_ships = Ship.objects.filter(showcase__isnull=True)

    context = {
        'available_ships': available_ships
    }
    return render(request, 'admin_panel/content/vessels/add_vessel_showcase.html', context)

@login_required
def edit_vessel_showcase(request, id):
    showcase = get_object_or_404(VesselShowcase, id=id)

    if request.method == "POST":
        try:
            name = request.POST.get('name')
            ship_id = request.POST.get('ship_id')
            tagline = request.POST.get('tagline')
            short_description = request.POST.get('short_description')
            full_description = request.POST.get('full_description')
            video_tour_url = request.POST.get('video_tour_url')
            display_capacity = request.POST.get('display_capacity')
            new_hero_image = request.FILES.get('hero_image')
            new_banner_image = request.FILES.get('banner_image') # Get banner

            # Basic Validation
            if not name:
                return JsonResponse({"status": "error", "message": "Name is required."})

            # --- Backend Hero Image Validation (Only if a new image is provided) ---
            if new_hero_image:
                if new_hero_image.size > 500 * 1024:
                    return JsonResponse({"status": "error", "message": "Hero Image size must not exceed 500KB."})

                try:
                    img = Image.open(new_hero_image)
                    img.verify() # Verifies it is a valid image file

                    # Re-open the file for dimension reading
                    new_hero_image.seek(0)
                    img = Image.open(new_hero_image)
                    width, height = img.size

                    if width != height:
                        return JsonResponse({"status": "error", "message": "Hero Image must be a square (1:1 aspect ratio)."})
                    
                    if width > 800 or height > 800:
                        return JsonResponse({"status": "error", "message": "Hero Image dimensions must not exceed 800x800 pixels."})

                except Exception:
                    return JsonResponse({"status": "error", "message": "Invalid hero image file uploaded."})
                
                # If valid, assign the new image
                showcase.hero_image = new_hero_image
            # ----------------------------------------------------------------

            # --- Backend Banner Image Validation (New, optional) ---
            if new_banner_image:
                # Banner size limit 2MB
                if new_banner_image.size > 2 * 1024 * 1024:
                    return JsonResponse({"status": "error", "message": "Banner Image size must not exceed 2MB."})

                try:
                    b_img = Image.open(new_banner_image)
                    b_img.verify()

                    # Re-open the file for dimension reading
                    new_banner_image.seek(0)
                    b_img = Image.open(new_banner_image)
                    b_width, b_height = b_img.size

                    # Validate 4:1 aspect ratio with a tolerance margin
                    if abs((b_width / b_height) - 4.0) > 0.05:
                        return JsonResponse({"status": "error", "message": "Banner Image must have exactly a 4:1 aspect ratio (e.g., 1920x480)."})

                except Exception:
                    return JsonResponse({"status": "error", "message": "Invalid banner image file uploaded."})
                
                # If valid, assign the new image
                showcase.banner_image = new_banner_image
            # ----------------------------------------------------------------

            # Check if Ship is selected and valid
            ship_instance = None
            if ship_id:
                # Exclude the current showcase ID to allow saving the same ship
                if VesselShowcase.objects.filter(ship_id=ship_id).exclude(id=showcase.id).exists():
                    return JsonResponse({"status": "error", "message": "This ship is already linked to another showcase."})
                ship_instance = Ship.objects.get(id=ship_id)

            # Update the showcase fields
            showcase.name = name
            showcase.ship = ship_instance
            showcase.tagline = tagline
            showcase.short_description = short_description
            showcase.full_description = full_description
            showcase.video_tour_url = video_tour_url
            showcase.display_capacity = display_capacity
            
            showcase.save()

            return JsonResponse({"status": "success", "message": "Vessel showcase updated successfully!"})

        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)})

    # GET Request: Fetch ships that don't already have a showcase, OR the one currently linked
    available_ships = Ship.objects.filter(Q(showcase__isnull=True) | Q(id=showcase.ship_id))
    
    context = {
        'showcase': showcase,
        'available_ships': available_ships
    }
    return render(request, 'admin_panel/content/vessels/edit_vessel_showcase.html', context)


@login_required
def cabin_showcases(request):
    # Handle the AJAX Delete Request
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            action = data.get("action")
            
            if action == "delete":
                cabin_id = data.get("id")
                cabin = CabinShowcase.objects.get(id=cabin_id)
                cabin.delete()
                
                return JsonResponse({"status": "success", "message": "Cabin showcase deleted successfully."})
                
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)})

    # Handle the standard GET Request
    cabins = CabinShowcase.objects.select_related('vessel').all().order_by('-created_at')
    
    context = {
        'cabins': cabins
    }
    return render(request, 'admin_panel/content/cabins/cabin_showcase_list.html', context)


@login_required
def add_cabin_showcase(request):
    if request.method == "POST":
        try:
            # --- Extract Fields ---
            title = request.POST.get('title')
            vessel_id = request.POST.get('vessel_id')
            subtitle = request.POST.get('subtitle', '')
            guest_capacity = request.POST.get('guest_capacity')
            room_size = request.POST.get('room_size', '')
            bed_type = request.POST.get('bed_type', '')
            short_description = request.POST.get('short_description')
            full_description = request.POST.get('full_description')
            cover_image = request.FILES.get('cover_image')
            banner_image = request.FILES.get('banner_image')
            
            # --- Extract M2M Fields ---
            operational_categories = request.POST.getlist('operational_categories')
            features = request.POST.getlist('features')

            # --- Basic Validation ---
            if not title or not guest_capacity or not cover_image:
                return JsonResponse({"status": "error", "message": "Title, Guest Capacity, and Cover Image are required."})

            # --- Backend Image Validation ---
            # 1. File Size Validation (<= 500KB)
            if cover_image.size > 500 * 1024:
                return JsonResponse({"status": "error", "message": "Image size must not exceed 500KB."})

            # 2. Dimensions & Format Validation
            try:
                img = Image.open(cover_image)
                img.verify() # Verifies it is a valid image file (handles jpg, png, webp, etc.)

                # Re-open the file for size/dimension reading (verify() leaves it in a closed state)
                cover_image.seek(0)
                img = Image.open(cover_image)
                width, height = img.size

                if width != height:
                    return JsonResponse({"status": "error", "message": "Image must be a square (1:1 aspect ratio)."})
                
                if width > 800 or height > 800:
                    return JsonResponse({"status": "error", "message": "Image dimensions must not exceed 800x800 pixels."})

            except Exception:
                return JsonResponse({"status": "error", "message": "Invalid image file uploaded."})

            # Cabin showcase update: optional banner image validation, same as vessel showcase
            if banner_image:
                if banner_image.size > 2 * 1024 * 1024:
                    return JsonResponse({"status": "error", "message": "Banner image size must not exceed 2MB."})

                try:
                    b_img = Image.open(banner_image)
                    b_img.verify()

                    banner_image.seek(0)
                    b_img = Image.open(banner_image)
                    b_width, b_height = b_img.size

                    if b_height == 0 or abs((b_width / b_height) - 4) > 0.2:
                        return JsonResponse({"status": "error", "message": "Banner image must be approximately 4:1 ratio."})

                except Exception:
                    return JsonResponse({"status": "error", "message": "Invalid banner image file uploaded."})
            # --------------------------------

            # Handle Foreign Key
            vessel_instance = None
            if vessel_id:
                vessel_instance = VesselShowcase.objects.get(id=vessel_id)

            # Create the showcase
            # Note: Fields omitted from the form (booleans, display_order) will take their default model values
            cabin = CabinShowcase.objects.create(
                vessel=vessel_instance,
                title=title,
                subtitle=subtitle,
                guest_capacity=guest_capacity,
                room_size=room_size,
                bed_type=bed_type,
                short_description=short_description,
                full_description=full_description,
                cover_image=cover_image,
                banner_image=banner_image
            )

            # Set Many-to-Many Relationships
            if operational_categories:
                cabin.operational_categories.set(operational_categories)
            if features:
                cabin.features.set(features)

            return JsonResponse({"status": "success", "message": "Cabin showcase created successfully!"})

        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)})

    # GET Request: Fetch context for dropdowns
    context = {
        'vessels': VesselShowcase.objects.all(),
        'categories': SeatCategory.objects.all(),
        'features': SeatFeature.objects.all(),
    }
    return render(request, 'admin_panel/content/cabins/add_cabin_showcase.html', context)


@login_required
def edit_cabin_showcase(request, pk):
    # Fetch the existing cabin showcase
    cabin = get_object_or_404(CabinShowcase, pk=pk)

    if request.method == "POST":
        try:
            # --- Extract Fields ---
            title = request.POST.get('title')
            vessel_id = request.POST.get('vessel_id')
            subtitle = request.POST.get('subtitle', '')
            guest_capacity = request.POST.get('guest_capacity')
            room_size = request.POST.get('room_size', '')
            bed_type = request.POST.get('bed_type', '')
            short_description = request.POST.get('short_description')
            full_description = request.POST.get('full_description')
            cover_image = request.FILES.get('cover_image')
            banner_image = request.FILES.get('banner_image')
            
            # --- Extract M2M Fields ---
            operational_categories = request.POST.getlist('operational_categories')
            features = request.POST.getlist('features')

            # --- Basic Validation ---
            if not title or not guest_capacity:
                return JsonResponse({"status": "error", "message": "Title and Guest Capacity are required."})

            # --- Backend Image Validation (Only if a NEW image is uploaded) ---
            if cover_image:
                # 1. File Size Validation (<= 500KB)
                if cover_image.size > 500 * 1024:
                    return JsonResponse({"status": "error", "message": "Image size must not exceed 500KB."})

                # 2. Dimensions & Format Validation
                try:
                    img = Image.open(cover_image)
                    img.verify() 

                    cover_image.seek(0)
                    img = Image.open(cover_image)
                    width, height = img.size

                    if width != height:
                        return JsonResponse({"status": "error", "message": "Image must be a square (1:1 aspect ratio)."})
                    
                    if width > 800 or height > 800:
                        return JsonResponse({"status": "error", "message": "Image dimensions must not exceed 800x800 pixels."})

                except Exception:
                    return JsonResponse({"status": "error", "message": "Invalid image file uploaded."})
                
                # Assign new image if it passes validation
                cabin.cover_image = cover_image

            # Cabin showcase update: optional banner image validation for edit
            if banner_image:
                if banner_image.size > 2 * 1024 * 1024:
                    return JsonResponse({"status": "error", "message": "Banner image size must not exceed 2MB."})

                try:
                    b_img = Image.open(banner_image)
                    b_img.verify()

                    banner_image.seek(0)
                    b_img = Image.open(banner_image)
                    b_width, b_height = b_img.size

                    if b_height == 0 or abs((b_width / b_height) - 4) > 0.2:
                        return JsonResponse({"status": "error", "message": "Banner image must be approximately 4:1 ratio."})

                except Exception:
                    return JsonResponse({"status": "error", "message": "Invalid banner image file uploaded."})

                cabin.banner_image = banner_image

            # --- Update Fields ---
            cabin.title = title
            # Handle Foreign Key (set to None if empty string is passed)
            cabin.vessel_id = vessel_id if vessel_id else None
            cabin.subtitle = subtitle
            cabin.guest_capacity = guest_capacity
            cabin.room_size = room_size
            cabin.bed_type = bed_type
            cabin.short_description = short_description
            cabin.full_description = full_description
            
            cabin.save()

            # --- Update Many-to-Many Relationships ---
            cabin.operational_categories.set(operational_categories)
            cabin.features.set(features)

            return JsonResponse({"status": "success", "message": "Cabin showcase updated successfully!"})

        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)})
        
    vessels = VesselShowcase.objects.all()
    logger.info(f"Editing Cabin Showcase ID {pk}: Found {vessels.count()} vessels for dropdown selection.")
    print(f"Editing Cabin Showcase ID {pk}: Found {vessels.count()} vessels for dropdown selection.")

    # GET Request: Fetch context for dropdowns and pre-selected M2M data
    context = {
        'cabin': cabin,
        'vessels': vessels,
        'categories': SeatCategory.objects.all(),
        'features': SeatFeature.objects.all(),
        # Pass lists of IDs for easy checking in the template
        'selected_categories': list(cabin.operational_categories.values_list('id', flat=True)),
        'selected_features': list(cabin.features.values_list('id', flat=True)),
    }
    return render(request, 'admin_panel/content/cabins/edit_cabin_showcase.html', context)


@login_required
def featured_articles(request):
    # Handle the AJAX Delete Request
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            action = data.get("action")
            
            if action == "delete":
                article_id = data.get("id")
                article = FeaturedArticle.objects.get(id=article_id)
                article.delete()
                
                return JsonResponse({"status": "success", "message": "Featured article deleted successfully."})
                
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)})

    # Handle the standard GET Request
    # Relying on the model's default ordering, or explicitly defining it here
    articles = FeaturedArticle.objects.all().order_by('-publication_date', '-created_at')
    
    context = {
        'articles': articles
    }
    return render(request, 'admin_panel/content/featured_articles/featured_article_list.html', context)


def validate_image_upload(image_file, max_size_mb=1):
    """
    Helper function to validate file size and MIME type on the backend.
    Returns an error message string if invalid, or None if valid.
    """
    if not image_file:
        return None

    # 1. Validate File Size
    if image_file.size > max_size_mb * 1024 * 1024:
        return f"Image size must be {max_size_mb}MB or less."

    # 2. Validate MIME Type (prevents someone from uploading a .exe renamed as .jpg)
    allowed_types = ['image/jpeg', 'image/png', 'image/webp']
    if image_file.content_type not in allowed_types:
        return "Invalid file type. Only JPG, PNG, and WEBP are allowed."

    return None

def process_and_compress_logo(uploaded_file, threshold_kb=80, target_kb=50):
    """
    Checks if the image exceeds threshold_kb. If so, scales down the resolution 
    proportionally until the file size is approximately under target_kb.
    """
    if not uploaded_file:
        return None

    # 1. If the file is already under the threshold (80KB), return it as-is
    if uploaded_file.size <= threshold_kb * 1024:
        return uploaded_file

    try:
        # 2. Open the image using Pillow
        img = Image.open(uploaded_file)
        file_format = img.format if img.format else 'PNG'
        
        # Prevent black backgrounds if a transparent image is accidentally saved as JPEG
        if file_format == 'JPEG' and img.mode in ('RGBA', 'P'):
            img = img.convert('RGB')

        output_io = BytesIO()
        quality = 90
        resize_factor = 0.85 # Reduce dimensions by 15% each iteration

        # Initial save to the buffer
        img.save(output_io, format=file_format, quality=quality, optimize=True)

        # 3. Loop: Keep shrinking the image until it falls under the target size (50KB)
        while output_io.tell() > target_kb * 1024:
            output_io.seek(0)
            output_io.truncate()

            # For JPEGs/WEBPs, we can drop quality. For PNGs, quality doesn't reduce file size much, 
            # so we strictly rely on reducing the resolution (dimensions).
            if file_format in ['JPEG', 'WEBP']:
                quality -= 10
                if quality < 30:
                    quality = 30
                    # Reduce resolution proportionally
                    new_width = int(img.width * resize_factor)
                    new_height = int(img.height * resize_factor)
                    img = img.resize((new_width, new_height), Image.Resampling.LANCZOS)
            else:
                # Reduce resolution proportionally for PNGs
                new_width = int(img.width * resize_factor)
                new_height = int(img.height * resize_factor)
                img = img.resize((new_width, new_height), Image.Resampling.LANCZOS)
            
            img.save(output_io, format=file_format, quality=quality, optimize=True)

            # Failsafe: break out if the image becomes ridiculously small to prevent infinite loops
            if img.width < 50 or img.height < 50:
                break

        output_io.seek(0)
        
        # 4. Wrap the compressed BytesIO buffer back into a Django InMemoryUploadedFile
        compressed_file = InMemoryUploadedFile(
            output_io,
            'ImageField',
            uploaded_file.name,
            uploaded_file.content_type,
            sys.getsizeof(output_io),
            None
        )
        return compressed_file
        
    except Exception as e:
        logger.error(f"Error compressing image: {str(e)}")
        return uploaded_file # Fallback to the original file if compression fails


@login_required
def add_featured_article(request):
    if request.method == "POST":
        try:
            name = request.POST.get('name', '').strip()
            organization_name = request.POST.get('organization_name', '').strip()
            url = request.POST.get('url', '').strip()
            publication_date = request.POST.get('publication_date', '').strip()
            description = request.POST.get('description', '').strip()
            is_active = request.POST.get('is_active') == 'true'
            logo = request.FILES.get('logo')

            if not all([name, organization_name, url, description, logo]):
                return JsonResponse({"status": "error", "message": "Please fill in all required fields and upload a logo."})

            validate_url = URLValidator()
            try:
                validate_url(url)
            except ValidationError:
                return JsonResponse({"status": "error", "message": "Please provide a valid URL."})

            # Compress the image if it exceeds 80KB
            processed_logo = process_and_compress_logo(logo, threshold_kb=80, target_kb=50)

            FeaturedArticle.objects.create(
                name=name,
                organization_name=organization_name,
                url=url,
                description=description,
                logo=processed_logo,
                is_active=is_active,
                publication_date=publication_date if publication_date else None 
            )

            return JsonResponse({"status": "success", "message": "Featured Article added successfully!"})

        except Exception as e:
            logger.error(f"Error adding featured article: {str(e)}")
            return JsonResponse({"status": "error", "message": "An unexpected server error occurred."})

    return render(request, 'admin_panel/content/featured_articles/add_featured_article.html')


@login_required
def edit_featured_article(request, article_id):
    article = get_object_or_404(FeaturedArticle, id=article_id)

    if request.method == "POST":
        try:
            name = request.POST.get('name', '').strip()
            organization_name = request.POST.get('organization_name', '').strip()
            url = request.POST.get('url', '').strip()
            publication_date = request.POST.get('publication_date', '').strip()
            description = request.POST.get('description', '').strip()
            is_active = request.POST.get('is_active') == 'true'
            logo = request.FILES.get('logo')

            if not all([name, organization_name, url, description]):
                return JsonResponse({"status": "error", "message": "Please fill in all required fields."})

            validate_url = URLValidator()
            try:
                validate_url(url)
            except ValidationError:
                return JsonResponse({"status": "error", "message": "Please provide a valid URL."})

            # Compress the image if a new one was uploaded
            if logo:
                processed_logo = process_and_compress_logo(logo, threshold_kb=80, target_kb=50)
                article.logo = processed_logo

            article.name = name
            article.organization_name = organization_name
            article.url = url
            article.description = description
            article.is_active = is_active
            article.publication_date = publication_date if publication_date else None 

            article.save()

            return JsonResponse({"status": "success", "message": "Featured Article updated successfully!"})

        except Exception as e:
            logger.error(f"Error editing featured article {article_id}: {str(e)}")
            return JsonResponse({"status": "error", "message": "An unexpected server error occurred."})

    context = {'article': article}
    return render(request, 'admin_panel/content/featured_articles/edit_featured_article.html', context)


# --- 1. Main List & Delete View ---
@login_required
def footer_management(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            action = data.get('action')

            # --- Handle Column AJAX Actions ---
            if action == 'add_column':
                FooterColumn.objects.create(
                    name=data.get('name'),
                    order=data.get('order', 0)
                )
                return JsonResponse({'status': 'success', 'message': 'Column added successfully!'})

            elif action == 'edit_column':
                column = FooterColumn.objects.get(id=data.get('id'))
                column.name = data.get('name')
                column.order = data.get('order', 0)
                column.save()
                return JsonResponse({'status': 'success', 'message': 'Column updated successfully!'})

            elif action == 'delete_column':
                FooterColumn.objects.get(id=data.get('id')).delete()
                return JsonResponse({'status': 'success', 'message': 'Column deleted successfully!'})

            # --- Handle Content AJAX Actions ---
            elif action == 'delete_content':
                FooterContent.objects.get(id=data.get('id')).delete()
                return JsonResponse({'status': 'success', 'message': 'Content deleted successfully!'})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    # GET Request: Show tables
    columns = FooterColumn.objects.all().order_by('order', 'name')
    contents = FooterContent.objects.select_related('column').all().order_by('column__order', 'order')

    return render(request, 'admin_panel/footer/footer_management.html', {
        'columns': columns, 
        'contents': contents,
    })


@login_required
def footer_social_links_view(request):
    social_settings, _ = FooterSocialSettings.objects.get_or_create(id=1)

    if request.method == 'POST':
        social_settings.facebook_url = (request.POST.get('facebook_url') or '').strip()
        social_settings.instagram_url = (request.POST.get('instagram_url') or '').strip()
        social_settings.youtube_url = (request.POST.get('youtube_url') or '').strip()
        social_settings.linkedin_url = (request.POST.get('linkedin_url') or '').strip()
        social_settings.save()
        messages.success(request, "Footer social links updated successfully.")
        return redirect('footer_social_links')

    return render(request, 'admin_panel/footer/footer_social_links.html', {
        'social_settings': social_settings,
    })

# --- Content Add/Edit Views remain unchanged below ---
def process_and_validate_images(image_file, banner_file):
    """
    Helper function to validate and process uploaded images.
    Returns a tuple: (error_message, processed_image_file, processed_banner_file)
    """
    processed_image = image_file
    processed_banner = banner_file

    # --- 1. Content Image Validation & Processing ---
    if image_file:
        try:
            img = Image.open(image_file)
            width, height = img.size
            
            # Validation: Must be a square
            if width != height:
                return "Content image must be a perfect square (1:1 aspect ratio).", None, None

            # Processing: If greater than 500KB, resize and compress
            if image_file.size > 500 * 1024:
                # Resize keeping aspect ratio (which is already 1:1)
                img.thumbnail((800, 800), Image.Resampling.LANCZOS)
                
                # Convert to RGB to avoid issues with saving PNG/RGBA as JPEG
                if img.mode in ("RGBA", "P"):
                    img = img.convert("RGB")
                    
                output = BytesIO()
                # Compress quality to bring file size down (aiming for < 50KB)
                img.save(output, format='JPEG', quality=60, optimize=True)
                output.seek(0)
                
                # Overwrite processed_image with the new InMemoryUploadedFile
                filename = f"{image_file.name.rsplit('.', 1)[0]}.jpg"
                processed_image = InMemoryUploadedFile(
                    output, 'ImageField', filename, 
                    'image/jpeg', sys.getsizeof(output), None
                )
        except Exception as e:
            return f"Invalid content image format. ({str(e)})", None, None

    # --- 2. Banner Image Validation ---
    if banner_file:
        # Validation: Must be under 2MB
        if banner_file.size > 2 * 1024 * 1024:
            return "Banner image must be less than 2MB.", None, None
            
        try:
            b_img = Image.open(banner_file)
            b_width, b_height = b_img.size
            
            # Validation: Aspect ratio must be 4:1
            ratio = round(b_width / b_height, 2)
            if ratio != 4.00:
                return f"Banner image aspect ratio must be exactly 4:1. Current ratio is {ratio}:1.", None, None
        except Exception as e:
            return f"Invalid banner image format. ({str(e)})", None, None

    return None, processed_image, processed_banner


def is_valid_footer_url(url):
    if not url:
        return False
    return url.startswith('/') or url.startswith('http://') or url.startswith('https://')


@login_required
def add_content(request):
    columns = FooterColumn.objects.all().order_by('order')
    
    if request.method == 'POST':
        column_id = request.POST.get('column_id')
        title = request.POST.get('title')
        url = request.POST.get('url')
        description = request.POST.get('description', '')
        order = request.POST.get('order', 0)
        
        raw_image = request.FILES.get('image')
        raw_banner_image = request.FILES.get('banner_image')

        # --- Backend Field Validation ---
        if not column_id or not title or not url:
            return JsonResponse({'status': 'error', 'message': 'Column, Title, and URL are strictly required fields.'})
        
        if not is_valid_footer_url(url):
            return JsonResponse({'status': 'error', 'message': 'URL must start with / or use a full http/https link.'})
        
        # --- Image Validation & Processing ---
        img_error, image, banner_image = process_and_validate_images(raw_image, raw_banner_image)
        if img_error:
            return JsonResponse({'status': 'error', 'message': img_error})
        
        try:
            order = int(order) if order else 0
            column = FooterColumn.objects.get(id=column_id)

            FooterContent.objects.create(
                column=column,
                title=title,
                url=url,
                description=description,
                order=order,
                image=image,
                banner_image=banner_image
            )
            
            return JsonResponse({
                'status': 'success', 
                'message': 'Content added successfully!',
                'redirect_url': reverse('footer_management')
            })
            
        except FooterColumn.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Invalid Column selected.'})
        except ValueError:
            return JsonResponse({'status': 'error', 'message': 'Order must be a valid number.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'An unexpected error occurred: {str(e)}'})

    return render(request, 'admin_panel/footer/footer_content_form.html', {'action': 'Add', 'columns': columns})


@login_required
def edit_content(request, id):
    content = get_object_or_404(FooterContent, id=id)
    columns = FooterColumn.objects.all().order_by('order')
    
    if request.method == 'POST':
        column_id = request.POST.get('column_id')
        title = request.POST.get('title')
        url = request.POST.get('url')
        description = request.POST.get('description', '')
        order = request.POST.get('order', 0)
        
        raw_image = request.FILES.get('image')
        raw_banner_image = request.FILES.get('banner_image')
        
        # --- Backend Field Validation ---
        if not column_id or not title or not url:
            return JsonResponse({'status': 'error', 'message': 'Column, Title, and URL are strictly required fields.'})
        
        if not is_valid_footer_url(url):
            return JsonResponse({'status': 'error', 'message': 'URL must start with / or use a full http/https link.'})

        # --- Image Validation & Processing ---
        img_error, image, banner_image = process_and_validate_images(raw_image, raw_banner_image)
        if img_error:
            return JsonResponse({'status': 'error', 'message': img_error})
        
        try:
            content.order = int(order) if order else 0
            column = FooterColumn.objects.get(id=column_id)
            content.column = column
            content.title = title
            content.url = url
            content.description = description
            
            if image:
                content.image = image
            if banner_image:
                content.banner_image = banner_image
                
            content.save()
            
            return JsonResponse({
                'status': 'success', 
                'message': 'Content updated successfully!',
                'redirect_url': reverse('footer_management')
            })
            
        except FooterColumn.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Invalid Column selected.'})
        except ValueError:
            return JsonResponse({'status': 'error', 'message': 'Order must be a valid number.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'An unexpected error occurred: {str(e)}'})

    return render(request, 'admin_panel/footer/footer_content_form.html', {'action': 'Edit', 'content': content, 'columns': columns})



#----------------------------------End---------------------------------------
#--------------------------#################---------------------------------
